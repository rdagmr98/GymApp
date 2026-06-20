#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analisi Buoni Pasto - Cartellini VERONA
Az. Ospedaliera Universitaria Integrata di Verona

Regole buono pasto (dal lunedì al sabato):
 1. Turni mattina > 6h20m e < 7h
 2. Turni mattina > 6h20m che terminano dopo le 14:45
 3. Turni pomeridiani > 6h20m
 4. Turni notturni (qualsiasi durata purché lavorato)

Domenica e festivi infrasettimanali:
 5. Qualsiasi turno (mattina/pomeriggio/notte) > 6h20m
"""

import io
import os
import re
import sys
import shutil

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ── paths ────────────────────────────────────────────────────────────────────
VERONA_PATH      = r"C:\Users\Gianmarco\Documents\VERONA"
ESTRATTI_PATH    = os.path.join(VERONA_PATH, "_estratti")
CARTELLINI_PATH  = os.path.join(VERONA_PATH, "cartellini")
OUTPUT_EXCEL     = os.path.join(VERONA_PATH, "riepilogo_buoni_pasto_verona.xlsx")

# ── constants ────────────────────────────────────────────────────────────────
MIN_DURATION = 6 * 60 + 20   # 380 min  (strictly greater)
EXIT_1445    = 14 * 60 + 45  # 885 min  threshold for rule 2

MONTHS_IT = {
    'gennaio': 1, 'febbraio': 2, 'marzo': 3,   'aprile': 4,
    'maggio': 5,  'giugno': 6,   'luglio': 7,   'agosto': 8,
    'settembre': 9, 'ottobre': 10, 'novembre': 11, 'dicembre': 12,
}

# Matches a valid timesheet day row:  "13 Ve ! T74 ! ... rest"
DAY_RE = re.compile(
    r'^\s*(\d\s*\d?)\s*(L\s*u|M\s*a|M\s*e|G\s*i|V\s*e|S\s*a|D\s*o)\s*!(.+?)!(.*)',
    re.IGNORECASE
)

# ── text helpers ─────────────────────────────────────────────────────────────

def deduplicate_chars(s: str) -> str:
    """'AANNTTOOLLIINNII' → 'ANTOLINI'  (collapse consecutive identical chars)"""
    result, i = [], 0
    while i < len(s):
        result.append(s[i])
        if i + 1 < len(s) and s[i + 1] == s[i]:
            i += 2
        else:
            i += 1
    return ''.join(result)


def parse_minutes(token: str) -> int:
    """'7,52' → 472  (hours×60+minutes, strip any suffix letters/symbols)"""
    token = re.sub(r'[^0-9,]', '', token)
    if ',' in token:
        parts = token.split(',')
        try:
            return int(parts[0]) * 60 + int(parts[1])
        except ValueError:
            return 0
    return 0


def extract_times(col: str):
    """Extract list of HH,MM time values from a column (handles spaces/underscores)."""
    clean = col.replace('_', '').replace(' ', '')
    return re.findall(r'\d+,\d{2}', clean)


# ── PDF parsing ──────────────────────────────────────────────────────────────

def extract_worker_name(text: str):
    """
    Parse doubled-character name header:
      'CCooggnnoommee AANNTTOOLLIINNII  NNoommee MMIICCHHAAEELL'
    Cognome/Nome can be multi-word ('DDII VVIIVVOO', 'MMAARRIIAA TTEERREESSAA') so the
    value is captured to end-of-line, not just the first token. The separator
    after the label is restricted to spaces/tabs (not '\\n'): in some layouts
    (e.g. 2018-era pages) the field is blank and the label sits at end-of-line,
    so this must fail to match rather than bleeding into the next line's
    unrelated text (was causing a spurious 'Stabil Reparto:' fake worker).
    Returns 'ANTOLINI MICHAEL' or None.
    """
    m_cog = re.search(r'CCooggnnoommee[ \t]+([^\n]+)', text)
    m_nom = re.search(r'NNoommee[ \t]+([^\n]+)', text)
    if m_cog and m_nom:
        cognome = deduplicate_chars(m_cog.group(1).strip())
        nome    = deduplicate_chars(m_nom.group(1).strip())
        return f"{cognome} {nome}"
    return None


def extract_month_year(text: str):
    """
    Parse the RILEVAZIONE header line, in either of two observed formats:
      - spaced-out:  'R I L E V A Z I O N E ... d e l m e s e d i N o v e m b r e 2 0 1 5'
      - contiguous:  'RILEVAZIONE PRESENZE del mese di Gennaio 2016'
    Returns (month_int, year_int) or (None, None).
    """
    for line in text.splitlines():
        if 'RILEVAZIONE' in line.replace(' ', '').upper():
            # Contiguous: 'Mese AAAA' as normal text
            m = re.search(r'([A-Za-zàèìòù]+)\s+(\d{4})\b', line)
            if m and 2000 <= int(m.group(2)) <= 2035:
                name = m.group(1).lower()
                if name in MONTHS_IT:
                    return MONTHS_IT[name], int(m.group(2))

            # Spaced: 4 space-separated single digits anywhere in the header
            # e.g. "2 0 1 5" (may have more text after it)
            year_m = re.search(r'(\d)\s+(\d)\s+(\d)\s+(\d)', line)
            if year_m:
                year = int(''.join(year_m.groups()))
                if not (2000 <= year <= 2035):
                    continue
                before = line[:year_m.start()]
                collapsed = re.sub(r'\s+', '', before).lower()
                for name, num in MONTHS_IT.items():
                    if name in collapsed:
                        return num, year
    return None, None


def classify_shift(turno: str, entry_min: int) -> str:
    """Return 'mattina', 'pomeriggio', or 'notte' based on turno code, then entry time."""
    t = turno.upper()
    if t.startswith('X') or t.startswith('N'):          # X99, N* = notte
        return 'notte'
    if t.startswith('T7') or t.startswith('T6') or t.startswith('T8') or t.startswith('B'):
        return 'mattina'                                 # T7x, T6x, B19 = mattina
    if t.startswith('T4') or t.startswith('T5') or t.startswith('V'):
        return 'pomeriggio'                              # T4x, T5x, V8x = pomeriggio
    # Fallback: classify by first entry time
    h = entry_min // 60
    if 5 <= h <= 12:
        return 'mattina'
    if 13 <= h <= 19:
        return 'pomeriggio'
    return 'notte'


def earns_buono(dow: str, is_festivo: bool, shift: str,
                duration: int, exit_min: int) -> bool:
    """Return True when this worked day earns a buono pasto."""
    if is_festivo or dow == 'DO':
        # Sunday or holiday: any shift > 6h20m
        return duration > MIN_DURATION

    if shift == 'mattina':
        # Rule 1: > 6h20m AND strictly < 7h
        if MIN_DURATION < duration < 420:
            return True
        # Rule 2: > 6h20m AND exits after 14:45
        if duration > MIN_DURATION and exit_min > EXIT_1445:
            return True
        return False

    if shift == 'pomeriggio':
        # Rule 3: afternoon > 6h20m
        return duration > MIN_DURATION

    # Night shift (notte): Rule 4 – any worked night
    return duration > 0


def process_page(page_text: str):
    """
    Process one PDF page (one month of presenze).
    Returns dict or None.
    """
    if 'RILEVAZIONE' not in page_text.replace(' ', '').upper():
        return None

    worker_name = extract_worker_name(page_text)
    month, year = extract_month_year(page_text)
    if not (worker_name and month and year):
        return None

    buoni      = 0
    details    = []
    prev_turno = None

    for line in page_text.splitlines():
        m = DAY_RE.match(line)
        if not m:
            continue

        day_str = m.group(1).replace(' ', '').lstrip('0') or '0'
        dow     = m.group(2).replace(' ', '').upper()           # LU/MA/ME/GI/VE/SA/DO
        t_raw   = m.group(3).strip().replace(' ', '')           # turno code (may be spaced)
        rest    = m.group(4)

        # Turno: ditto mark inherits previous
        DITTO = ('"', "'", '\u3003', '\u2033')
        if not t_raw or t_raw in DITTO:
            turno = prev_turno or ''
        else:
            turno = t_raw
            prev_turno = turno

        # Split remaining columns on '!'
        cols     = rest.split('!')
        timb_col = cols[0] if len(cols) > 0 else ''
        pres_col = cols[1] if len(cols) > 1 else ''
        desc_col = cols[-1].strip() if cols else ''

        # Pres. (actual worked hours)
        pres_toks = extract_times(pres_col)
        duration  = parse_minutes(pres_toks[0]) if pres_toks else 0

        # Skip non-worked rows (RRR exit carry-over, FERIE, etc.)
        if duration == 0:
            details.append({
                'day': day_str, 'dow': dow, 'turno': turno,
                'worked': False, 'qualifies': False, 'note': desc_col,
            })
            continue

        # Entry / exit times
        timb = extract_times(timb_col)
        entry_min = parse_minutes(timb[0])  if timb           else 0
        exit_min  = parse_minutes(timb[-1]) if len(timb) > 1 else entry_min

        # Festivo detection
        is_festivo = (
            dow == 'DO' or
            turno == 'FFF' or
            'FESTIVO INFRASETT' in desc_col.upper()
        )

        shift = classify_shift(turno, entry_min)

        qualifies = earns_buono(dow, is_festivo, shift, duration, exit_min)
        if qualifies:
            buoni += 1

        rule = ''
        if qualifies:
            if is_festivo or dow == 'DO':
                rule = 'fest/dom'
            elif shift == 'mattina' and MIN_DURATION < duration < 420:
                rule = 'r1 mattina <7h'
            elif shift == 'mattina' and exit_min > EXIT_1445:
                rule = 'r2 mattina >14:45'
            elif shift == 'pomeriggio':
                rule = 'r3 pomeriggio'
            else:
                rule = 'r4 notte'

        details.append({
            'day': day_str, 'dow': dow, 'turno': turno,
            'entry': f"{entry_min//60}:{entry_min%60:02d}" if entry_min else '',
            'exit':  f"{exit_min//60}:{exit_min%60:02d}"  if exit_min  else '',
            'duration_hm': f"{duration//60}h{duration%60:02d}m",
            'is_festivo': is_festivo, 'shift': shift,
            'worked': True, 'qualifies': qualifies, 'rule': rule,
            'note': desc_col,
        })

    return {
        'worker': worker_name,
        'month': month,
        'year': year,
        'buoni': buoni,
        'details': details,
    }


def process_pdf(pdf_path: str):
    """
    Process all pages of a cumulative VERONA cartellino PDF.
    Returns list of monthly result dicts.
    """
    results = []
    print(f"  Apertura: {os.path.basename(pdf_path)}", flush=True)
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ''
            r = process_page(text)
            if r:
                results.append(r)
                if i % 20 == 0:
                    print(f"    pagine {i}/{total} ...", flush=True)
    print(f"    → {len(results)} mesi elaborati", flush=True)
    return results


# ── cartellini folder ────────────────────────────────────────────────────────

def find_cartellini_pdfs(base: str):
    """
    Find all cumulative cartellino PDFs under base.
    They live directly in a worker subfolder (NOT in a 4-digit year subfolder).
    """
    found = []
    for root, dirs, files in os.walk(base):
        # Skip year subfolders
        if re.match(r'^\d{4}$', os.path.basename(root)):
            dirs.clear()
            continue
        for f in files:
            if f.lower().endswith('.pdf'):
                found.append(os.path.join(root, f))
    return found


def populate_cartellini_folder(cartellini_pdfs):
    """
    Copy each cartellino PDF to CARTELLINI_PATH renamed 'COGNOME NOME.pdf'
    using the worker name extracted from the PDF itself.
    """
    os.makedirs(CARTELLINI_PATH, exist_ok=True)
    copied = []
    for pdf_path in cartellini_pdfs:
        print(f"  Lettura nome da: {os.path.basename(pdf_path)}", flush=True)
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text() or ''
            name = extract_worker_name(text)
        except Exception as e:
            print(f"  ERRORE lettura nome: {e}", flush=True)
            name = None

        if not name:
            # Fallback: use file stem
            name = os.path.splitext(os.path.basename(pdf_path))[0]
            print(f"  (nome non trovato, uso: {name})", flush=True)

        dest = os.path.join(CARTELLINI_PATH, f"{name}.pdf")
        shutil.copy2(pdf_path, dest)
        print(f"  Copiato → {os.path.basename(dest)}", flush=True)
        copied.append((name, dest))
    return copied


# ── Excel output ──────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HDR_FONT   = Font(bold=True, color='FFFFFF')
TOTAL_FILL = PatternFill('solid', fgColor='D6E4F0')
TOTAL_FONT = Font(bold=True)
ALT_FILL   = PatternFill('solid', fgColor='EEF4FB')
THIN       = Side(style='thin', color='CCCCCC')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER     = Alignment(horizontal='center', vertical='center')


def _hcell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CENTER, BORDER


def _cell(ws, row, col, value, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    if bold:
        c.font = TOTAL_FONT
    if fill:
        c.fill = fill
    c.alignment = CENTER
    c.border = BORDER


def create_excel(all_results: list):
    """
    Build the Excel workbook:
    - Sheet 'Riepilogo': rows=workers, cols=years, totale
    - One detail sheet per worker
    """
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = 'Riepilogo'

    # Aggregate: worker → year → month → buoni
    from collections import defaultdict
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    details_by_worker = defaultdict(list)

    for r in all_results:
        data[r['worker']][r['year']][r['month']] += r['buoni']
        details_by_worker[r['worker']].append(r)

    workers = sorted(data.keys())
    years   = sorted({y for w in data.values() for y in w.keys()})

    # ── Summary sheet ─────────────────────────────────────────────────────────
    _hcell(ws_sum, 1, 1, 'Lavoratore')
    for ci, y in enumerate(years, 2):
        _hcell(ws_sum, 1, ci, y)
    _hcell(ws_sum, 1, len(years) + 2, 'TOTALE')

    for ri, worker in enumerate(workers, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _cell(ws_sum, ri, 1, worker, fill=fill)
        total = 0
        for ci, y in enumerate(years, 2):
            val = sum(data[worker][y].values())
            total += val
            _cell(ws_sum, ri, ci, val if val else '', fill=fill)
        _cell(ws_sum, ri, len(years) + 2, total, bold=True,
              fill=PatternFill('solid', fgColor='BDD7EE'))

    # Totals row
    tr = len(workers) + 2
    _cell(ws_sum, tr, 1, 'TOTALE', bold=True, fill=TOTAL_FILL)
    grand = 0
    for ci, y in enumerate(years, 2):
        col_total = sum(sum(data[w][y].values()) for w in workers)
        grand += col_total
        _cell(ws_sum, tr, ci, col_total, bold=True, fill=TOTAL_FILL)
    _cell(ws_sum, tr, len(years) + 2, grand, bold=True,
          fill=PatternFill('solid', fgColor='9DC3E6'))

    # Column widths
    ws_sum.column_dimensions['A'].width = 26
    for ci in range(2, len(years) + 3):
        ws_sum.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = 10

    # ── Detail sheets ─────────────────────────────────────────────────────────
    MONTH_NAMES = {
        1: 'Gen', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'Mag', 6: 'Giu',
        7: 'Lug', 8: 'Ago', 9: 'Set', 10: 'Ott', 11: 'Nov', 12: 'Dic',
    }

    for worker in workers:
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', worker)[:31]
        ws = wb.create_sheet(title=safe_name)

        # Header
        headers = ['Anno', 'Mese', 'Giorno', 'Gg', 'Turno',
                   'Entrata', 'Uscita', 'Ore lav.', 'Fest/Dom',
                   'Tipo turno', 'Buono', 'Regola', 'Note']
        for ci, h in enumerate(headers, 1):
            _hcell(ws, 1, ci, h)

        ri = 2
        month_records = sorted(details_by_worker[worker],
                                key=lambda r: (r['year'], r['month']))
        for rec in month_records:
            for d in rec['details']:
                if not d['worked']:
                    continue
                fill = ALT_FILL if ri % 2 == 0 else None
                vals = [
                    rec['year'],
                    MONTH_NAMES.get(rec['month'], rec['month']),
                    d['day'],
                    d['dow'],
                    d['turno'],
                    d.get('entry', ''),
                    d.get('exit', ''),
                    d.get('duration_hm', ''),
                    'Sì' if d.get('is_festivo') else '',
                    d.get('shift', ''),
                    'Sì' if d['qualifies'] else '',
                    d.get('rule', ''),
                    d.get('note', ''),
                ]
                for ci, v in enumerate(vals, 1):
                    _cell(ws, ri, ci, v, fill=fill)
                ri += 1

            # Month subtotal row
            month_total = rec['buoni']
            label = f"{MONTH_NAMES.get(rec['month'], rec['month'])} {rec['year']}"
            c = ws.cell(row=ri, column=11,
                        value=f"Totale {label}: {month_total}")
            c.font = TOTAL_FONT
            c.fill = TOTAL_FILL
            c.border = BORDER
            ri += 1

        # Column widths
        col_widths = [6, 5, 7, 4, 7, 8, 8, 9, 9, 11, 7, 18, 30]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel salvato: {OUTPUT_EXCEL}", flush=True)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Analisi Buoni Pasto VERONA")
    print("=" * 60)

    # Step 1: find and copy cartellini PDFs
    print("\n[1] Ricerca cartellini PDF ...")
    pdfs = find_cartellini_pdfs(ESTRATTI_PATH)
    print(f"    Trovati {len(pdfs)} PDF cartellini")

    print("\n[2] Copia nella cartella 'cartellini/' ...")
    copied = populate_cartellini_folder(pdfs)

    # Step 2: process each PDF
    print("\n[3] Elaborazione cartellini ...")
    all_results = []
    for name, pdf_path in copied:
        print(f"\n  Lavoratore: {name}")
        results = process_pdf(pdf_path)
        all_results.extend(results)
        total = sum(r['buoni'] for r in results)
        print(f"    Totale buoni pasto maturati: {total}")

    # Step 3: generate Excel
    print("\n[4] Generazione Excel ...")
    create_excel(all_results)

    grand_total = sum(r['buoni'] for r in all_results)
    print(f"\n{'='*60}")
    print(f"Totale buoni pasto maturati (tutti i lavoratori): {grand_total}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
