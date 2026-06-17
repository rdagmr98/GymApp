#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys, zipfile, io, tempfile, re, shutil
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, r"C:\Users\Gianmarco\Documents\tivoli")
import analisi_tivoli, openpyxl, fitz
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DOWNLOADS    = r"C:\Users\Gianmarco\Downloads"
PREV_DEF     = os.path.join(DOWNLOADS, "tivoli12062026_output", "buonipasto DEFINITIVO")
OUT_ROOT     = os.path.join(DOWNLOADS, "tivoli17062026_output")
DEFINITIVO   = os.path.join(OUT_ROOT, "buonipasto DEFINITIVO")
OUT_ZIP      = os.path.join(DOWNLOADS, "tivoli17062026_DEFINITIVO.zip")
BUONO_EURO   = 4.13
BUONO_ORE    = 0.5
MONTH_ABBR   = {'Gen':1,'Feb':2,'Mar':3,'Apr':4,'Mag':5,'Giu':6,
                'Lug':7,'Ago':8,'Set':9,'Ott':10,'Nov':11,'Dic':12}
MESI_IT      = {"GENNAIO":1,"FEBBRAIO":2,"MARZO":3,"APRILE":4,"MAGGIO":5,"GIUGNO":6,
                "LUGLIO":7,"AGOSTO":8,"SETTEMBRE":9,"OTTOBRE":10,"NOVEMBRE":11,"DICEMBRE":12}

os.makedirs(DEFINITIVO, exist_ok=True)
analisi_tivoli.OUTPUT_DIR = DEFINITIVO
analisi_tivoli.LOG_FILE   = os.path.join(OUT_ROOT, "pdf_non_leggibili.log")

# ── Copia Excel esistenti dal precedente DEFINITIVO ──────────────────────────
print("Copio Excel esistenti dal DEFINITIVO precedente...")
copied = 0
for fname in os.listdir(PREV_DEF):
    if fname.endswith(".xlsx") and not fname.startswith("RIEPILOGO"):
        shutil.copy2(os.path.join(PREV_DEF, fname), os.path.join(DEFINITIVO, fname))
        copied += 1
print(f"  Copiati {copied} file.")

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_cid29_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    raw = doc[0].get_text()
    doc.close()
    decoded = "".join(chr(ord(c) + 29) for c in raw)
    m = re.search(r"([A-Z][A-Z\s]+),\s+(\w+)\s+(\d{4})", decoded)
    if not m:
        return None
    worker = m.group(1).strip()
    month  = MESI_IT.get(m.group(2).upper())
    year   = int(m.group(3))
    if not month:
        return None
    m2 = re.search(r"Buoni pasto.(\d+)", decoded)
    maturati = int(m2.group(1)) if m2 else 0
    return {"worker": worker, "year": year, "month": month,
            "maturati": maturati, "erogati": 0, "erogati_found": False, "details": []}

def collect_pdfs_recursive(zip_bytes, depth=0):
    result = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for name in sorted(z.namelist()):
            if name.lower().endswith('.pdf'):
                result.append((name, z.read(name)))
            elif name.lower().endswith('.zip') and depth < 2:
                try:
                    inner = z.read(name)
                    result.extend(collect_pdfs_recursive(inner, depth + 1))
                except Exception as e:
                    print(f"    ZIP interno ignorato ({name}): {e}")
    return result

def read_xlsx_records(xlsx_bytes, worker):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    records = []
    for sn in wb.sheetnames:
        try: year = int(sn)
        except: continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == 'TOTALE MESE':
                mese = row[0]; mat = row[5]; ero = row[6]
                if mese in MONTH_ABBR:
                    records.append({
                        'worker': worker, 'year': year, 'month': MONTH_ABBR[mese],
                        'maturati': int(mat) if isinstance(mat, (int, float)) else 0,
                        'erogati':  int(ero) if isinstance(ero, (int, float)) else 0,
                        'erogati_found': True, 'details': [],
                    })
    wb.close()
    return records

def process_worker(zip_path, worker, xlsx_name, inner_zip_name):
    print(f"\n{'='*55}")
    print(f"Worker: {worker}")

    with zipfile.ZipFile(zip_path) as z:
        xlsx_bytes      = z.read(xlsx_name)
        inner_zip_bytes = z.read(inner_zip_name)

    xlsx_records  = read_xlsx_records(xlsx_bytes, worker)
    existing_keys = {(r['year'], r['month']) for r in xlsx_records}
    print(f"  XLSX: {len(xlsx_records)} mesi  anni={sorted({r['year'] for r in xlsx_records})}")

    pdfs = collect_pdfs_recursive(inner_zip_bytes)
    print(f"  PDF trovati: {len(pdfs)}")

    new_records = []
    for pname, pdata in pdfs:
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
            f.write(pdata); tmp = f.name
        try:
            recs = analisi_tivoli.process_pdf(tmp)
            if not recs:
                rec = parse_cid29_pdf(tmp)
                if rec:
                    rec['worker'] = worker
                    recs = [rec]
        except Exception as e:
            print(f"  ERRORE {os.path.basename(pname)}: {e}")
            recs = []
        finally:
            os.unlink(tmp)
        for r in recs:
            r['worker'] = worker
        new_records.extend(recs)

    print(f"  PDF parsati: {len(new_records)} record")
    truly_new = [r for r in new_records if (r['year'], r['month']) not in existing_keys]
    # deduplicazione: per ogni (year,month) tieni il record con maturati più alto
    seen = {}
    for r in truly_new:
        k = (r['year'], r['month'])
        if k not in seen or r['maturati'] > seen[k]['maturati']:
            seen[k] = r
    truly_new = list(seen.values())
    print(f"  Nuovi da PDF: {sorted((r['year'],r['month']) for r in truly_new)}")

    merged = xlsx_records + truly_new
    merged.sort(key=lambda r: (r['year'], r['month']))

    out_path, _ = analisi_tivoli.create_worker_excel(worker, merged, DEFINITIVO)
    tot_mat = sum(r['maturati'] for r in merged)
    tot_ero = sum(r['erogati'] for r in merged)
    print(f"  Salvato: {out_path}")
    print(f"  Totale: mat={tot_mat} ero={tot_ero} delta={tot_mat-tot_ero} ({(tot_mat-tot_ero)*BUONO_EURO:.2f} EUR)  anni={sorted({r['year'] for r in merged})}")

# ── Processa i 3 nuovi lavoratori ────────────────────────────────────────────

WORKERS = [
    ("romasabrinativoli.zip",        "ROMA SABRINA",      "ROMA SABRINA.xlsx",      "cartellini 17-23.zip"),
    ("iricorsobuonipastotivoli.zip",  "IMPERIALE ILARIA",  "IMPERIALE ILARIA.xlsx",  "cartellini imperiale.zip"),
    ("minghellapatriziativoli.zip",   "MINGHELLA PATRIZIA","MINGHELLA PATRIZIA.xlsx","CARTELLINI 15-22.zip"),
]

for zname, worker, xlsx_name, inner_zip in WORKERS:
    process_worker(os.path.join(DOWNLOADS, zname), worker, xlsx_name, inner_zip)

# ── Rebuild RIEPILOGO ─────────────────────────────────────────────────────────
print(f"\n{'='*55}")
print("Rebuild RIEPILOGO...")

HDR_FILL = PatternFill("solid", fgColor="1F4E79"); HDR_FONT = Font(bold=True, color="FFFFFF")
TOT_FILL = PatternFill("solid", fgColor="BDD7EE"); TOT_FONT = Font(bold=True)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL = PatternFill("solid", fgColor="EEF4FB"); NEG_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="CCCCCC"); BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR  = Alignment(horizontal="center", vertical="center")
LFT  = Alignment(horizontal="left",   vertical="center")

def rh(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CTR, BRD

def rc(ws, r, c, v, bold=False, fill=None, align=None):
    cell = ws.cell(row=r, column=c, value=v)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    cell.alignment = align or CTR
    cell.border = BRD
    return cell

def rf(ws, r, c, formula, bold=False, fill=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=formula)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    cell.alignment = CTR; cell.border = BRD
    return cell

agg = defaultdict(lambda: defaultdict(int))
agg_ero = defaultdict(lambda: defaultdict(int))

xlsx_files = sorted(fn for fn in os.listdir(DEFINITIVO)
                    if fn.endswith(".xlsx") and not fn.startswith("RIEPILOGO"))
for fname in xlsx_files:
    worker2 = fname[:-5]; path2 = os.path.join(DEFINITIVO, fname)
    try:
        wb2 = openpyxl.load_workbook(path2, read_only=True, data_only=True)
        for sn in wb2.sheetnames:
            if sn == "Riepilogo": continue
            try: year2 = int(sn)
            except: continue
            ws2 = wb2[sn]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if row[1] == "TOTALE MESE":
                    mat2 = row[5] if isinstance(row[5], (int, float)) else 0
                    ero2 = row[6] if isinstance(row[6], (int, float)) else 0
                    agg[worker2][year2]     += max(0, int(mat2) - int(ero2))
                    agg_ero[worker2][year2] += int(ero2)
        wb2.close()
    except Exception as e:
        print(f"  ERRORE {fname}: {e}")

workers_all = sorted(agg.keys())
years_all   = sorted({y for w in agg.values() for y in w})
print(f"  Lavoratori: {len(workers_all)}  Anni: {years_all[0]}-{years_all[-1]}")

wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = "Riepilogo Generale"
rh(ws3, 1, 1, "Lavoratore")
for ci, y in enumerate(years_all, 2): rh(ws3, 1, ci, y)
tc = len(years_all) + 2; ec = len(years_all) + 3; oc = len(years_all) + 4
tcl = openpyxl.utils.get_column_letter(tc)
lyl = openpyxl.utils.get_column_letter(len(years_all) + 1)
ocl = openpyxl.utils.get_column_letter(oc)
rh(ws3, 1, tc, "TOTALE Delta"); rh(ws3, 1, ec, "Euro da recuperare"); rh(ws3, 1, oc, "Ore da recuperare")

for ri, w2 in enumerate(workers_all, 2):
    fill = ALT_FILL if ri % 2 == 0 else None
    rc(ws3, ri, 1, w2, fill=fill, align=LFT)
    for ci, y in enumerate(years_all, 2):
        v = agg[w2].get(y, 0)
        fi = fill if v >= 0 else NEG_FILL
        rc(ws3, ri, ci, v if v != 0 else "", fill=fi)
    rf(ws3, ri, tc, f"=SUM(B{ri}:{lyl}{ri})", bold=True, fill=TOT_FILL)
    rf(ws3, ri, ec, f"={tcl}{ri}*{BUONO_EURO}", bold=True, fill=TOT_FILL, fmt='#,##0.00 "EUR"')
    ero_tot = sum(agg_ero[w2].get(y, 0) for y in years_all)
    ore_cell = rc(ws3, ri, oc, ero_tot * BUONO_ORE, bold=True, fill=TOT_FILL)
    ore_cell.number_format = '0.0 "h"'

tr = len(workers_all) + 2
rc(ws3, tr, 1, "TOTALE", bold=True, fill=SUB_FILL)
for ci, y in enumerate(years_all, 2):
    col_l = openpyxl.utils.get_column_letter(ci)
    rf(ws3, tr, ci, f"=SUM({col_l}2:{col_l}{tr-1})", bold=True, fill=SUB_FILL)
rf(ws3, tr, tc, f"=SUM({tcl}2:{tcl}{tr-1})", bold=True, fill=TOT_FILL)
rf(ws3, tr, ec, f"={tcl}{tr}*{BUONO_EURO}", bold=True, fill=TOT_FILL, fmt='#,##0.00 "EUR"')
rf(ws3, tr, oc, f"=SUM({ocl}2:{ocl}{tr-1})", bold=True, fill=TOT_FILL, fmt='0.0 "h"')
ws3.column_dimensions["A"].width = 35
for ci in range(2, len(years_all) + 5):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

riepilogo_path = os.path.join(DEFINITIVO, "RIEPILOGO_BUONI_PASTO.xlsx")
wb3.save(riepilogo_path)
grand_delta = sum(sum(yd.values()) for yd in agg.values())
grand_ero   = sum(sum(yd.values()) for yd in agg_ero.values())
print(f"  TOTALE DELTA  : {grand_delta:,} buoni pasto")
print(f"  VALORE EURO   : {grand_delta * BUONO_EURO:,.2f} EUR")
print(f"  TOTALE EROGATI: {grand_ero:,} buoni pasto")

# ── ZIP finale ────────────────────────────────────────────────────────────────
all_xlsx = sorted(fn for fn in os.listdir(DEFINITIVO) if fn.endswith(".xlsx"))
print(f"\nZIP con {len(all_xlsx)} file...")
with zipfile.ZipFile(OUT_ZIP, "w", zipfile.ZIP_DEFLATED) as zout:
    for fname in all_xlsx:
        zout.write(os.path.join(DEFINITIVO, fname), fname)
        print(f"  + {fname}")
print(f"\nZIP: {OUT_ZIP}")
