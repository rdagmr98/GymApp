#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aggiunge intestazione (NOME/COGNOME/LUOGO) e tabella di riepilogo ANNO|IMPORTO
a tutti i file xlsx in ELABORATI sangiovanni e ELABORATI_ROMA3.
Evidenzia in arancione le celle Medie annue della colonna Sommatoria.
"""
import os, sys, re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SANGIOVANNI = r"C:\Users\Gianmarco\Documents\ELABORATI sangiovanni"
ROMA3       = r"C:\Users\Gianmarco\Documents\ELABORATI_ROMA3"
LUOGO_SGV   = "OSP S GIOVANNI"
LUOGO_R3    = "ASL ROMA 3 OSTIA"

ORANGE  = PatternFill("solid", fgColor="FFC000")
BLUE_H  = PatternFill("solid", fgColor="4472C4")
LIGHT_B = PatternFill("solid", fgColor="DDEBF7")
BOLD_W  = Font(bold=True, color="FFFFFF")
BOLD    = Font(bold=True)
CENTER  = Alignment(horizontal="center", vertical="center")
THIN    = Side(style="thin", color="CCCCCC")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EUR_FMT = '#,##0.00 "€"'


def parse_sgv(filename):
    stem = os.path.splitext(filename)[0]
    parts = stem.split("_", 1)
    return (parts[0], parts[1].replace("_", " ")) if len(parts) >= 2 else (stem, "")


def parse_r3(filename):
    stem = os.path.splitext(filename)[0]
    stem = re.split(r"\s*[-–]\s*", stem)[0].strip()
    parts = stem.split(" ", 1)
    return (parts[0].strip(), parts[1].strip()) if len(parts) >= 2 else (stem.strip(), "")


def process_file(filepath, cognome, nome, luogo):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Salta se già elaborato
    if ws.cell(1, 1).value == "NOME":
        print(f"  SKIP (già elaborato): {os.path.basename(filepath)}")
        wb.close()
        return False

    # Inserisce 3 righe in cima (openpyxl aggiorna le riferimenti delle formule)
    ws.insert_rows(1, 3)

    # Intestazione
    for r, label, val in [(1, "NOME", nome), (2, "COGNOME", cognome), (3, "LUOGO DI LAVORO", luogo)]:
        lc = ws.cell(r, 1, label)
        lc.font = BOLD
        vc = ws.cell(r, 2, val)
        vc.font = BOLD

    # Rileva blocchi anno (ANNO XXXX) e calcola riga Medie annue (header+15)
    year_medie = {}
    for r in range(4, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and isinstance(val, str) and val.startswith("ANNO "):
            try:
                year = int(val.strip().split()[-1])
                year_medie[year] = r + 15
            except ValueError:
                pass

    # Arancione sulle celle Medie annue col G (Sommatoria)
    for mr in year_medie.values():
        c = ws.cell(mr, 7)
        c.fill = ORANGE
        c.font = BOLD

    # Tabella riepilogo in colonne L(12) e M(13) a partire da riga 1
    sorted_years = sorted(year_medie.keys())

    def sc(cell, fill=None, font=None, align=CENTER, fmt=None):
        if fill:  cell.fill = fill
        if font:  cell.font = font
        cell.alignment = align
        cell.border = BORDER
        if fmt: cell.number_format = fmt

    sc(ws.cell(1, 12, "ANNO"),    fill=BLUE_H, font=BOLD_W)
    sc(ws.cell(1, 13, "IMPORTO"), fill=BLUE_H, font=BOLD_W)

    for i, year in enumerate(sorted_years):
        r = 2 + i
        sc(ws.cell(r, 12, year),               fill=LIGHT_B)
        sc(ws.cell(r, 13, f"=G{year_medie[year]}"), fill=LIGHT_B, fmt=EUR_FMT)

    tot_r = 2 + len(sorted_years)
    sc(ws.cell(tot_r, 12, "TOTALE"),                  font=BOLD)
    sc(ws.cell(tot_r, 13, f"=SUM(M2:M{tot_r - 1})"), font=BOLD, fmt=EUR_FMT)

    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 16

    wb.save(filepath)
    wb.close()
    return True


errors = []
done = 0

for folder, luogo, parse_fn in [
    (SANGIOVANNI, LUOGO_SGV, parse_sgv),
    (ROMA3,       LUOGO_R3,  parse_r3),
]:
    print(f"\n{'='*60}")
    print(f"Cartella: {os.path.basename(folder)}")
    print(f"{'='*60}")
    files = sorted([
        f for f in os.listdir(folder)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ])
    for fname in files:
        cognome, nome = parse_fn(fname)
        fpath = os.path.join(folder, fname)
        try:
            ok = process_file(fpath, cognome, nome, luogo)
            if ok:
                done += 1
                print(f"  OK  [{cognome}] [{nome}]: {fname}")
        except Exception as e:
            print(f"  ERR {fname}: {e}")
            errors.append(fname)

print(f"\n{'='*60}")
print(f"Completato: {done} file elaborati")
if errors:
    print(f"Errori ({len(errors)}):")
    for e in errors:
        print(f"  - {e}")
