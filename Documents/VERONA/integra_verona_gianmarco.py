#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Integrazione VERONA: 11 nuovi lavoratori da 'verona gianmarco.zip' (Downloads).
Nessuno dei 11 e' gia' presente nei 99 lavoratori del riepilogo esistente.

A differenza di analizza_verona20062026.py (che rigenera l'intero workbook da
tutte le fonti), qui la chiavetta USB E:\\verona non e' collegata: i dati
grezzi dei 99 lavoratori esistenti non sono disponibili per una rigenerazione
completa. Si integra quindi in modo incrementale: si leggono i totali
anno-per-anno gia' presenti nel foglio 'Riepilogo', si aggiungono gli 11
nuovi lavoratori (con i loro fogli di dettaglio), si riscrive SOLO il foglio
'Riepilogo'. I 99 fogli di dettaglio esistenti restano intatti.
"""
import os, sys, shutil, zipfile
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, r"C:\Users\Gianmarco\Documents\VERONA")
import analisi_verona as av
from analizza_verona20062026 import find_worker_pdfs

import openpyxl

ZIP_SRC   = r"C:\Users\Gianmarco\Downloads\verona gianmarco.zip"
EXTRACT   = os.path.join(av.VERONA_PATH, "_verona_gianmarco_estratto")
ROOT      = os.path.join(EXTRACT, "verona gianmarco")
BACKUP    = os.path.join(av.VERONA_PATH, "riepilogo_buoni_pasto_verona_BACKUP_pre_gianmarco.xlsx")


def process_worker(folder):
    """Stessa logica di dedup semantico usata in analizza_verona20062026.main()
    per i lavoratori USB: trova i PDF (dedup MD5), elabora, dedup per
    (worker, anno, mese) tenendo la versione con piu' buoni in caso di
    conflitto reale."""
    pdfs = find_worker_pdfs(folder)
    worker_results = []
    for pdf_path in pdfs:
        try:
            r = av.process_pdf(pdf_path)
        except Exception as e:
            print(f"    ERRORE {pdf_path}: {e}")
            r = []
        worker_results.extend(r)

    deduped = {}
    conflicts = []
    for r in worker_results:
        key = (r['worker'], r['year'], r['month'])
        sig = tuple(
            (d['day'], d['dow'], d['turno'], d['worked'], d['qualifies'],
             d.get('entry', ''), d.get('exit', ''), d.get('duration_hm', ''),
             d.get('note', ''))
            for d in r['details']
        )
        if key in deduped:
            prev_r, prev_sig = deduped[key]
            if not r['details']:
                continue
            if not prev_r['details']:
                deduped[key] = (r, sig)
                continue
            if prev_sig != sig:
                conflicts.append(key)
                if r['buoni'] > prev_r['buoni']:
                    deduped[key] = (r, sig)
            continue
        deduped[key] = (r, sig)

    return [v[0] for v in deduped.values()], len(pdfs), conflicts


def main():
    if not os.path.exists(BACKUP):
        shutil.copy2(av.OUTPUT_EXCEL, BACKUP)
        print(f"Backup creato: {BACKUP}")

    print("=" * 60)
    print("[1] Estrazione zip")
    print("=" * 60)
    if not os.path.exists(ROOT):
        with zipfile.ZipFile(ZIP_SRC, 'r') as zf:
            zf.extractall(EXTRACT)
        print(f"  Estratto in: {ROOT}")
    else:
        print(f"  Gia' estratto: {ROOT}")

    worker_folders = sorted(
        d for d in os.listdir(ROOT) if os.path.isdir(os.path.join(ROOT, d))
    )
    print(f"  Cartelle lavoratore trovate: {len(worker_folders)}")

    print("\n" + "=" * 60)
    print("[2] Elaborazione 11 nuovi lavoratori")
    print("=" * 60)
    all_results_new = []
    no_data = []
    for i, wf in enumerate(worker_folders, 1):
        wpath = os.path.join(ROOT, wf)
        results, n_pdfs, conflicts = process_worker(wpath)
        if results:
            names = {r['worker'] for r in results}
            total = sum(r['buoni'] for r in results)
            print(f"  [{i}/{len(worker_folders)}] {wf}: {n_pdfs} PDF, "
                  f"{len(results)} mesi, nomi={names}, buoni={total}")
            if conflicts:
                print(f"      CONFLITTI: {conflicts}")
            all_results_new.extend(results)
        else:
            print(f"  [{i}/{len(worker_folders)}] {wf}: NESSUN DATO ({n_pdfs} PDF trovati)")
            no_data.append(wf)

    names_found = sorted({r['worker'] for r in all_results_new})
    print(f"\n  Nomi lavoratore estratti dai PDF: {names_found}")

    print("\n" + "=" * 60)
    print("[3] Lettura totali esistenti dal Riepilogo")
    print("=" * 60)
    wb = openpyxl.load_workbook(av.OUTPUT_EXCEL)
    ws_old = wb['Riepilogo']
    rows = list(ws_old.iter_rows(values_only=True))
    header = rows[0]
    years_existing = [y for y in header[1:-1]]
    year_totals = {}
    for row in rows[1:-1]:
        worker = row[0]
        year_totals[worker] = {y: (v or 0) for y, v in zip(years_existing, row[1:-1])}
    print(f"  Lavoratori esistenti: {len(year_totals)}")
    print(f"  Anni esistenti: {years_existing}")

    print("\n" + "=" * 60)
    print("[4] Merge nuovi lavoratori")
    print("=" * 60)
    from collections import defaultdict
    details_by_worker_new = defaultdict(list)
    for r in all_results_new:
        details_by_worker_new[r['worker']].append(r)

    for worker, recs in details_by_worker_new.items():
        yt = defaultdict(int)
        for r in recs:
            yt[r['year']] += r['buoni']
        year_totals[worker] = dict(yt)
        print(f"  + {worker}: {dict(sorted(yt.items()))} (tot {sum(yt.values())})")

    print("\n" + "=" * 60)
    print("[5] Riscrittura foglio Riepilogo")
    print("=" * 60)
    wb.remove(ws_old)
    ws_sum = wb.create_sheet('Riepilogo', 0)
    wb.active = 0

    workers = sorted(year_totals.keys())
    years = sorted({y for yt in year_totals.values() for y in yt.keys()})

    av._hcell(ws_sum, 1, 1, 'Lavoratore')
    for ci, y in enumerate(years, 2):
        av._hcell(ws_sum, 1, ci, y)
    av._hcell(ws_sum, 1, len(years) + 2, 'TOTALE')

    for ri, worker in enumerate(workers, 2):
        fill = av.ALT_FILL if ri % 2 == 0 else None
        av._cell(ws_sum, ri, 1, worker, fill=fill)
        total = 0
        for ci, y in enumerate(years, 2):
            val = year_totals[worker].get(y, 0)
            total += val
            av._cell(ws_sum, ri, ci, val if val else '', fill=fill)
        av._cell(ws_sum, ri, len(years) + 2, total, bold=True,
                 fill=openpyxl.styles.PatternFill('solid', fgColor='BDD7EE'))

    tr = len(workers) + 2
    av._cell(ws_sum, tr, 1, 'TOTALE', bold=True, fill=av.TOTAL_FILL)
    grand = 0
    for ci, y in enumerate(years, 2):
        col_total = sum(year_totals[w].get(y, 0) for w in workers)
        grand += col_total
        av._cell(ws_sum, tr, ci, col_total, bold=True, fill=av.TOTAL_FILL)
    av._cell(ws_sum, tr, len(years) + 2, grand, bold=True,
             fill=openpyxl.styles.PatternFill('solid', fgColor='9DC3E6'))

    ws_sum.column_dimensions['A'].width = 26
    for ci in range(2, len(years) + 3):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 10

    print(f"  Lavoratori totali: {len(workers)}")
    print(f"  Totale buoni pasto (tutti): {grand}")

    print("\n" + "=" * 60)
    print("[6] Fogli di dettaglio per gli 11 nuovi lavoratori")
    print("=" * 60)
    MONTH_NAMES = {
        1: 'Gen', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'Mag', 6: 'Giu',
        7: 'Lug', 8: 'Ago', 9: 'Set', 10: 'Ott', 11: 'Nov', 12: 'Dic',
    }
    import re as _re
    for worker, recs in details_by_worker_new.items():
        safe_name = _re.sub(r'[\\/*?:\[\]]', '_', worker)[:31]
        ws = wb.create_sheet(title=safe_name)
        headers = ['Anno', 'Mese', 'Giorno', 'Gg', 'Turno',
                   'Entrata', 'Uscita', 'Ore lav.', 'Fest/Dom',
                   'Tipo turno', 'Buono', 'Regola', 'Note']
        for ci, h in enumerate(headers, 1):
            av._hcell(ws, 1, ci, h)

        ri = 2
        for rec in sorted(recs, key=lambda r: (r['year'], r['month'])):
            for d in rec['details']:
                if not d['worked']:
                    continue
                fill = av.ALT_FILL if ri % 2 == 0 else None
                vals = [
                    rec['year'], MONTH_NAMES.get(rec['month'], rec['month']),
                    d['day'], d['dow'], d['turno'],
                    d.get('entry', ''), d.get('exit', ''), d.get('duration_hm', ''),
                    'Sì' if d.get('is_festivo') else '',
                    d.get('shift', ''),
                    'Sì' if d['qualifies'] else '',
                    d.get('rule', ''), d.get('note', ''),
                ]
                for ci, v in enumerate(vals, 1):
                    av._cell(ws, ri, ci, v, fill=fill)
                ri += 1

            month_total = rec['buoni']
            label = f"{MONTH_NAMES.get(rec['month'], rec['month'])} {rec['year']}"
            c = ws.cell(row=ri, column=11, value=f"Totale {label}: {month_total}")
            c.font = av.TOTAL_FONT
            c.fill = av.TOTAL_FILL
            c.border = av.BORDER
            ri += 1

        col_widths = [6, 5, 7, 4, 7, 8, 8, 9, 9, 11, 7, 18, 30]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        print(f"  Foglio creato: {safe_name}")

    wb.save(av.OUTPUT_EXCEL)
    print(f"\nExcel salvato: {av.OUTPUT_EXCEL}")

    if no_data:
        print(f"\nATTENZIONE - lavoratori senza dati validi: {no_data}")


if __name__ == '__main__':
    main()
