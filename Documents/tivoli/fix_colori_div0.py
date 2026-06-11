#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, re, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SANGIOVANNI = r"C:\Users\Gianmarco\Documents\ELABORATI sangiovanni"
ROMA3       = r"C:\Users\Gianmarco\Documents\ELABORATI_ROMA3"

ORANGE   = PatternFill("solid", fgColor="FFC000")  # arancione pieno
ORANGE_L = PatternFill("solid", fgColor="FFE699")  # arancione chiaro per righe anno
BOLD     = Font(bold=True)
EUR_FMT  = '#,##0.00 "€"'
CENTER   = Alignment(horizontal="center", vertical="center")

# 3 cognomi composti Roma3 da correggere: nome_fix, cognome_fix
SURNAME_FIXES = {
    "DE PASCALIS CESARE.xlsx": ("CESARE",    "DE PASCALIS"),
    "DE ROSA FRANCESCO.xlsx":  ("FRANCESCO", "DE ROSA"),
    "DI MARTINO LAURA.xlsx":   ("LAURA",     "DI MARTINO"),
}


def fix_div0(formula):
    """=(Gx/Hx*22.5)  →  =IF(Hx=0,0,Gx/Hx*22.5)"""
    m = re.match(r'^=\(?G(\d+)/H(\d+)\*22\.5\)?$', formula)
    if m and "IF" not in formula:
        g, h = m.group(1), m.group(2)
        return f"=IF(H{h}=0,0,G{g}/H{h}*22.5)"
    return formula


def fix_file(filepath, nome_fix=None, cognome_fix=None):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    if ws.cell(1, 1).value != "NOME":
        wb.close()
        return False

    # ── 1. Correggi cognomi composti ─────────────────────────────────────────
    if nome_fix:
        ws.cell(1, 2).value = nome_fix
        ws.cell(1, 2).font  = BOLD
    if cognome_fix:
        ws.cell(2, 2).value = cognome_fix
        ws.cell(2, 2).font  = BOLD

    # ── 2. Colore arancione righe header 1-3 (label + valore) ────────────────
    for r in range(1, 4):
        for c in range(1, 3):
            cell = ws.cell(r, c)
            cell.fill = ORANGE
            cell.font = BOLD

    # ── 3. Scan righe dati 4+ ────────────────────────────────────────────────
    div0_fixed = 0
    for r in range(4, ws.max_row + 1):
        val_a = ws.cell(r, 1).value

        # Righe ANNO XXXX → arancione pieno su cols A-H
        if val_a and isinstance(val_a, str) and val_a.startswith("ANNO "):
            for c in range(1, 9):  # A-H
                cell = ws.cell(r, c)
                cell.fill = ORANGE
                if c == 1:
                    cell.font = BOLD

        # Fix DIV/0 in col G (Medie annue)
        g_cell = ws.cell(r, 7)
        if isinstance(g_cell.value, str) and g_cell.value.startswith("="):
            new_val = fix_div0(g_cell.value)
            if new_val != g_cell.value:
                g_cell.value = new_val
                div0_fixed += 1

    # ── 4. Colori tabella riepilogo (cols L=12, M=13) ────────────────────────
    hdr_l = ws.cell(1, 12)
    hdr_m = ws.cell(1, 13)
    if hdr_l.value == "ANNO" or hdr_m.value == "IMPORTO":
        # Header (riga 1)
        for c in [12, 13]:
            cell = ws.cell(1, c)
            cell.fill      = ORANGE
            cell.font      = BOLD
            cell.alignment = CENTER
            if c == 13:
                cell.number_format = EUR_FMT

        # Righe anno + TOTALE
        for r in range(2, 20):
            lv = ws.cell(r, 12).value
            mv = ws.cell(r, 13).value
            if lv is None and mv is None:
                break
            is_tot = (lv == "TOTALE")
            fill   = ORANGE if is_tot else ORANGE_L
            for c in [12, 13]:
                cell           = ws.cell(r, c)
                cell.fill      = fill
                cell.font      = BOLD
                cell.alignment = CENTER
                if c == 13:
                    cell.number_format = EUR_FMT

    wb.save(filepath)
    wb.close()
    return div0_fixed


# ── Processa tutto ────────────────────────────────────────────────────────────
total_files  = 0
total_div0   = 0
errors       = 0

for folder, label in [(SANGIOVANNI, "SANGIOVANNI"), (ROMA3, "ROMA3")]:
    files = sorted(f for f in os.listdir(folder) if f.endswith(".xlsx"))
    print(f"\n{label}: {len(files)} file")
    for fname in files:
        fpath      = os.path.join(folder, fname)
        nome_fix   = None
        cognome_fix = None
        if fname in SURNAME_FIXES:
            nome_fix, cognome_fix = SURNAME_FIXES[fname]
        try:
            n = fix_file(fpath, nome_fix, cognome_fix)
            if n is False:
                print(f"  SKIP (no header): {fname}")
                continue
            total_files += 1
            total_div0  += n
            if fname in SURNAME_FIXES:
                print(f"  COGNOME FISSO: {fname} → {cognome_fix} / {nome_fix}")
        except Exception as e:
            print(f"  ERRORE {fname}: {e}")
            errors += 1

print(f"\nFile aggiornati : {total_files}")
print(f"DIV/0 corretti  : {total_div0}")
print(f"Errori          : {errors}")
