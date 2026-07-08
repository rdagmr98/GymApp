#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys, zipfile, io, tempfile, re, shutil
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, r"C:\Users\Gianmarco\Documents\tivoli")
import analisi_tivoli, openpyxl, fitz, py7zr
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DOWNLOADS   = r"C:\Users\Gianmarco\Downloads"
DOCS_TIV    = r"C:\Users\Gianmarco\Documents\tivoli"
PREV_DEF    = os.path.join(DOWNLOADS, "tivoli02072026_output", "buonipasto DEFINITIVO")  # 215 workers
OUT_ROOT    = os.path.join(DOWNLOADS, "tivoli08072026_output")
DEFINITIVO  = os.path.join(OUT_ROOT, "buonipasto DEFINITIVO")
OUT_ZIP     = os.path.join(DOWNLOADS, "tivoli08072026_DEFINITIVO.zip")
BUONO_EURO  = 4.13
BUONO_ORE   = 0.5
MONTH_ABBR  = {'Gen':1,'Feb':2,'Mar':3,'Apr':4,'Mag':5,'Giu':6,
               'Lug':7,'Ago':8,'Set':9,'Ott':10,'Nov':11,'Dic':12}

os.makedirs(DEFINITIVO, exist_ok=True)
analisi_tivoli.OUTPUT_DIR = DEFINITIVO
analisi_tivoli.LOG_FILE   = os.path.join(OUT_ROOT, "pdf_non_leggibili.log")

# ── Helpers ───────────────────────────────────────────────────────────────────

def collect_pdfs_from_7z_bytes(sevenz_bytes, path_filter=None):
    result = []
    with tempfile.NamedTemporaryFile(suffix='.7z', delete=False) as f:
        f.write(sevenz_bytes); tmp7z = f.name
    tmpdir = tempfile.mkdtemp()
    try:
        with py7zr.SevenZipFile(tmp7z, mode='r') as z7:
            z7.extractall(path=tmpdir)
        for root, _, files in os.walk(tmpdir):
            for fname in sorted(files):
                if path_filter and path_filter.lower() not in fname.lower():
                    continue
                if fname.lower().endswith('.pdf'):
                    with open(os.path.join(root, fname), 'rb') as fp:
                        result.append((fname, fp.read()))
    finally:
        os.unlink(tmp7z)
        shutil.rmtree(tmpdir, ignore_errors=True)
    return result

def collect_pdfs_recursive(zip_bytes, depth=0, path_filter=None):
    result = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for name in sorted(z.namelist()):
            if path_filter and path_filter.lower() not in name.lower():
                continue
            if name.lower().endswith('.pdf'):
                result.append((name, z.read(name)))
            elif name.lower().endswith('.zip') and depth < 2:
                try:
                    inner = z.read(name)
                    result.extend(collect_pdfs_recursive(inner, depth + 1, path_filter))
                except Exception as e:
                    print(f"    ZIP interno ignorato ({name}): {e}")
            elif name.lower().endswith('.7z') and depth < 2:
                try:
                    inner = z.read(name)
                    result.extend(collect_pdfs_from_7z_bytes(inner, path_filter))
                except Exception as e:
                    print(f"    7Z interno ignorato ({name}): {e}")
    return result

def read_xlsx_records(xlsx_bytes, worker):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    records = []
    for sn in wb.sheetnames:
        try: year = int(sn)
        except: continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == 'TOTALE MESE':
                mese = row[0]; mat = row[5]; ero = row[6]
                if mese in MONTH_ABBR:
                    records.append({
                        'worker': worker, 'year': year, 'month': MONTH_ABBR[mese],
                        'maturati': int(mat) if isinstance(mat, (int, float)) else 0,
                        'erogati':  int(ero) if isinstance(ero, (int, float)) else 0,
                        'erogati_found': True, 'details': [],
                    })
    wb.close()
    return records

def merge_record_lists(records):
    seen = {}
    for r in records:
        k = (r['year'], r['month'])
        if k not in seen or r['maturati'] > seen[k]['maturati']:
            seen[k] = r
    return sorted(seen.values(), key=lambda r: (r['year'], r['month']))

# ── Carry-forward dal DEFINITIVO 02/07 (215 worker) ──────────────────────────
print("Carry-forward da DEFINITIVO 02/07/2026...")
for fname in os.listdir(PREV_DEF):
    if fname.endswith('.xlsx') and not fname.lower().startswith('riepilogo'):
        shutil.copy2(os.path.join(PREV_DEF, fname), os.path.join(DEFINITIVO, fname))
n_carry = len([f for f in os.listdir(DEFINITIVO) if f.endswith('.xlsx')])
print(f"  Copiati: {n_carry} worker")

# ── Processa nuovi worker ─────────────────────────────────────────────────────

def process_worker(worker, zip_sources, pdf_paths=None, base_records=None):
    print(f"\n{'='*55}")
    print(f"Worker: {worker}")

    xlsx_records = base_records or []
    existing_keys = {(r['year'], r['month']) for r in xlsx_records}
    print(f"  Base: {len(xlsx_records)} mesi  anni={sorted({r['year'] for r in xlsx_records})}")

    pdfs = []
    for zpath, pfilter in (zip_sources or []):
        with open(zpath, 'rb') as f:
            zbytes = f.read()
        found = collect_pdfs_recursive(zbytes, path_filter=pfilter)
        print(f"  {os.path.basename(zpath)} (filtro={pfilter}): {len(found)} pdf")
        pdfs.extend(found)

    for pdf_path in (pdf_paths or []):
        with open(pdf_path, 'rb') as f:
            pdata = f.read()
        pdfs.append((os.path.basename(pdf_path), pdata))
        print(f"  PDF diretto: {os.path.basename(pdf_path)}")

    print(f"  PDF totali: {len(pdfs)}")

    new_records = []
    for pname, pdata in pdfs:
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
            f.write(pdata); tmp = f.name
        try:
            recs = analisi_tivoli.process_pdf(tmp)
        except Exception as e:
            print(f"  ERRORE {os.path.basename(pname)}: {e}")
            recs = []
        finally:
            os.unlink(tmp)
        for r in recs:
            r['worker'] = worker
        new_records.extend(recs)

    print(f"  PDF parsati: {len(new_records)} record")
    truly_new = [r for r in new_records if (r['year'], r['month']) not in existing_keys]
    seen = {}
    for r in truly_new:
        k = (r['year'], r['month'])
        if k not in seen or r['maturati'] > seen[k]['maturati']:
            seen[k] = r
    truly_new = list(seen.values())
    print(f"  Nuovi da PDF: {sorted((r['year'], r['month']) for r in truly_new)}")

    merged = merge_record_lists(xlsx_records + truly_new)
    out_path, _ = analisi_tivoli.create_worker_excel(worker, merged, DEFINITIVO)
    tot_mat = sum(r['maturati'] for r in merged)
    tot_ero = sum(r['erogati'] for r in merged)
    print(f"  Salvato: {out_path}")
    print(f"  Totale: mat={tot_mat} ero={tot_ero} delta={tot_mat-tot_ero} ({(tot_mat-tot_ero)*BUONO_EURO:.2f} EUR)")


DICENSI_ZIP  = os.path.join(DOWNLOADS, "icartellinidicensilorettativoli.zip")
MARIANI_ZIP  = os.path.join(DOWNLOADS, "cartellini alessandro.zip")
TANTARI_PDF  = os.path.join(DOWNLOADS, "3516_2_TANTARI VANESSA.pdf")

# DI CENSI LORETTA (Matricola 0000050129, CHIRURGIA GENER. E D'URGENZA T)
# nuova al bacino: base = xlsx nello zip (2023-2025), poi PDF 72 pagine per gli anni residui
with zipfile.ZipFile(DICENSI_ZIP) as z:
    dicensi_base = read_xlsx_records(z.read("DI CENSI LORETTA.xlsx"), "DI CENSI LORETTA")
process_worker("DI CENSI LORETTA",
               [(DICENSI_ZIP, None)],
               base_records=dicensi_base)

# MARIANI ALESSANDRO (Matricola 0000061163, MEDIC URG P.S.Osser Breve MONTEROTONDO)
# del tutto nuovo: solo PDF (33 mensili 2023-2026 + 1 storico 2015-2023)
process_worker("MARIANI ALESSANDRO",
               [(MARIANI_ZIP, None)])

# TANTARI VANESSA (Matricola 0000081464, MEDICINA INTERNA TIVOLI)
# del tutto nuova: 1 solo PDF diretto (fuori zip)
process_worker("TANTARI VANESSA",
               [],
               pdf_paths=[TANTARI_PDF])

# ── Rebuild RIEPILOGO ─────────────────────────────────────────────────────────
print(f"\n{'='*55}")
print("Rebuild RIEPILOGO...")

HDR_FILL = PatternFill("solid", fgColor="1F4E79"); HDR_FONT = Font(bold=True, color="FFFFFF")
TOT_FILL = PatternFill("solid", fgColor="BDD7EE"); TOT_FONT = Font(bold=True)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL = PatternFill("solid", fgColor="EEF4FB"); NEG_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="CCCCCC"); BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR  = Alignment(horizontal="center", vertical="center")
LFT  = Alignment(horizontal="left",   vertical="center")

def rh(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CTR, BRD

def rc(ws, r, c, v, bold=False, fill=None, align=None):
    cell = ws.cell(row=r, column=c, value=v)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    cell.alignment = align or CTR
    cell.border = BRD
    return cell

def rf(ws, r, c, formula, bold=False, fill=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=formula)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    cell.alignment = CTR; cell.border = BRD
    return cell

agg = defaultdict(lambda: defaultdict(int))
agg_ero = defaultdict(lambda: defaultdict(int))

xlsx_files = sorted(fn for fn in os.listdir(DEFINITIVO)
                    if fn.endswith(".xlsx") and not fn.startswith("RIEPILOGO"))
for fname in xlsx_files:
    worker2 = fname[:-5]; path2 = os.path.join(DEFINITIVO, fname)
    try:
        wb2 = openpyxl.load_workbook(path2, read_only=True, data_only=True)
        for sn in wb2.sheetnames:
            if sn == "Riepilogo": continue
            try: year2 = int(sn)
            except: continue
            ws2 = wb2[sn]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if row[1] == "TOTALE MESE":
                    mat2 = row[5] if isinstance(row[5], (int, float)) else 0
                    ero2 = row[6] if isinstance(row[6], (int, float)) else 0
                    agg[worker2][year2]     += max(0, int(mat2) - int(ero2))
                    agg_ero[worker2][year2] += int(ero2)
        wb2.close()
    except Exception as e:
        print(f"  ERRORE {fname}: {e}")

workers_all = sorted(agg.keys())
years_all   = sorted({y for w in agg.values() for y in w})
print(f"  Lavoratori: {len(workers_all)}  Anni: {years_all[0]}-{years_all[-1]}")

wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = "Riepilogo Generale"
rh(ws3, 1, 1, "Lavoratore")
for ci, y in enumerate(years_all, 2): rh(ws3, 1, ci, y)
tc = len(years_all) + 2; ec = len(years_all) + 3; oc = len(years_all) + 4
tcl = openpyxl.utils.get_column_letter(tc)
lyl = openpyxl.utils.get_column_letter(len(years_all) + 1)
ocl = openpyxl.utils.get_column_letter(oc)
rh(ws3, 1, tc, "TOTALE Delta"); rh(ws3, 1, ec, "Euro da recuperare"); rh(ws3, 1, oc, "Ore da recuperare")

for ri, w2 in enumerate(workers_all, 2):
    fill = ALT_FILL if ri % 2 == 0 else None
    rc(ws3, ri, 1, w2, fill=fill, align=LFT)
    for ci, y in enumerate(years_all, 2):
        v = agg[w2].get(y, 0)
        fi = fill if v >= 0 else NEG_FILL
        rc(ws3, ri, ci, v if v != 0 else "", fill=fi)
    rf(ws3, ri, tc, f"=SUM(B{ri}:{lyl}{ri})", bold=True, fill=TOT_FILL)
    rf(ws3, ri, ec, f"={tcl}{ri}*{BUONO_EURO}", bold=True, fill=TOT_FILL, fmt='#,##0.00 "EUR"')
    ero_tot = sum(agg_ero[w2].get(y, 0) for y in years_all)
    ore_cell = rc(ws3, ri, oc, ero_tot * BUONO_ORE, bold=True, fill=TOT_FILL)
    ore_cell.number_format = '0.0 "h"'

tr = len(workers_all) + 2
rc(ws3, tr, 1, "TOTALE", bold=True, fill=SUB_FILL)
for ci, y in enumerate(years_all, 2):
    col_l = openpyxl.utils.get_column_letter(ci)
    rf(ws3, tr, ci, f"=SUM({col_l}2:{col_l}{tr-1})", bold=True, fill=SUB_FILL)
rf(ws3, tr, tc, f"=SUM({tcl}2:{tcl}{tr-1})", bold=True, fill=TOT_FILL)
rf(ws3, tr, ec, f"={tcl}{tr}*{BUONO_EURO}", bold=True, fill=TOT_FILL, fmt='#,##0.00 "EUR"')
rf(ws3, tr, oc, f"=SUM({ocl}2:{ocl}{tr-1})", bold=True, fill=TOT_FILL, fmt='0.0 "h"')
ws3.column_dimensions["A"].width = 35
for ci in range(2, len(years_all) + 5):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

riepilogo_path = os.path.join(DEFINITIVO, "RIEPILOGO_BUONI_PASTO.xlsx")
wb3.save(riepilogo_path)
grand_delta = sum(sum(yd.values()) for yd in agg.values())
grand_ero   = sum(sum(yd.values()) for yd in agg_ero.values())
print(f"  TOTALE DELTA  : {grand_delta:,} buoni pasto")
print(f"  VALORE EURO   : {grand_delta * BUONO_EURO:,.2f} EUR")
print(f"  TOTALE EROGATI: {grand_ero:,} buoni pasto")

# ── ZIP finale ────────────────────────────────────────────────────────────────
all_xlsx = sorted(fn for fn in os.listdir(DEFINITIVO) if fn.endswith(".xlsx"))
print(f"\nZIP con {len(all_xlsx)} file...")
with zipfile.ZipFile(OUT_ZIP, "w", zipfile.ZIP_DEFLATED) as zout:
    for fname in all_xlsx:
        zout.write(os.path.join(DEFINITIVO, fname), fname)
print(f"\nZIP: {OUT_ZIP}")
