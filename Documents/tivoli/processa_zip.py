#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Legge sangioveroma3.zip, applica tutte le modifiche ai singoli xlsx
(header NOME/COGNOME/LUOGO, tabella riepilogo L-M, colori arancione,
fix riferimenti formula +3, fix DIV/0 Medie annue),
produce sangioveroma3_AGGIORNATO.zip pronto per la mail.
"""
import os, sys, re, io, zipfile, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

INPUT_ZIP  = r"C:\Users\Gianmarco\Downloads\sangioveroma3.zip"
OUTPUT_ZIP = r"C:\Users\Gianmarco\Downloads\sangioveroma3_AGGIORNATO.zip"

LUOGO_SGV = "OSP S GIOVANNI"
LUOGO_R3  = "ASL ROMA 3 OSTIA"

# Prefissi cognomi composti italiani
PREFIXES = {"DE","DI","DA","DEL","DELLA","DELLE","DEGLI","DAL","DALLA","LO","LA"}

ORANGE   = PatternFill("solid", fgColor="FFC000")
ORANGE_L = PatternFill("solid", fgColor="FFE699")
BOLD     = Font(bold=True)
EUR_FMT  = '#,##0.00 "€"'
CENTER   = Alignment(horizontal="center", vertical="center")


# ── Parser nomi ───────────────────────────────────────────────────────────────

def parse_sgv(filename):
    """COGNOME_NOME.xlsx (underscore, anche cognomi composti come DI_COLA_STEFANO)"""
    stem  = os.path.splitext(filename)[0]
    parts = stem.split("_")
    if len(parts) < 2:
        return stem, ""
    if parts[0].upper() in PREFIXES:
        cognome = " ".join(parts[:2])
        nome    = " ".join(parts[2:])
    else:
        cognome = parts[0]
        nome    = " ".join(parts[1:])
    return cognome, nome


def parse_r3(filename):
    """COGNOME NOME.xlsx (spazi, anche 'DE PASCALIS CESARE' e ' - note' in coda)"""
    stem  = os.path.splitext(filename)[0]
    stem  = re.split(r"\s*[-–]\s*", stem)[0].strip()
    parts = stem.split(" ")
    if len(parts) < 2:
        return stem, ""
    if parts[0].upper() in PREFIXES:
        cognome = " ".join(parts[:2])
        nome    = " ".join(parts[2:])
    else:
        cognome = parts[0]
        nome    = " ".join(parts[1:])
    return cognome, nome


# ── Helpers formula ───────────────────────────────────────────────────────────

def shift_formula(formula):
    """Aggiunge +3 a tutti i riferimenti di riga relativi."""
    def rep(m):
        if m.group(3) == "$":   # riga assoluta
            return m.group(0)
        return f"{m.group(1)}{m.group(2)}{int(m.group(4)) + 3}"
    return re.sub(r"(\$?)([A-Z]+)(\$?)(\d+)", rep, formula)


def fix_div0(formula):
    """=(Gx/Hx*22.5) → =IF(Hx=0,0,Gx/Hx*22.5)"""
    m = re.match(r"^=\(?G(\d+)/H(\d+)\*22\.5\)?$", formula)
    if m and "IF" not in formula:
        g, h = m.group(1), m.group(2)
        return f"=IF(H{h}=0,0,G{g}/H{h}*22.5)"
    return formula


# ── Processore singolo xlsx ───────────────────────────────────────────────────

def process_xlsx(data_bytes, cognome, nome, luogo):
    wb = openpyxl.load_workbook(io.BytesIO(data_bytes))
    ws = wb.active

    if ws.cell(1, 1).value == "NOME":
        out = io.BytesIO(); wb.save(out); wb.close()
        return out.getvalue(), False   # già processato

    # 1. Inserisci 3 righe in cima
    ws.insert_rows(1, 3)

    # 2. Header con fill arancione
    for r, label, val in [(1,"NOME",nome),(2,"COGNOME",cognome),(3,"LUOGO DI LAVORO",luogo)]:
        for c, v in [(1, label), (2, val)]:
            cell = ws.cell(r, c, v)
            cell.fill = ORANGE
            cell.font = BOLD

    # 3. Rileva blocchi anno, applica arancione su righe ANNO
    year_medie = {}
    for r in range(4, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and isinstance(val, str) and val.startswith("ANNO "):
            try:
                year = int(val.strip().split()[-1])
                year_medie[year] = r + 15   # riga Medie annue nel blocco
            except ValueError:
                pass
            for c in range(1, 9):   # cols A-H
                cell = ws.cell(r, c)
                cell.fill = ORANGE
                if c == 1: cell.font = BOLD

    # 4. Arancione sulle celle Medie annue (col G)
    for mr in year_medie.values():
        cell = ws.cell(mr, 7)
        cell.fill = ORANGE
        cell.font = BOLD

    # 5. Tabella riepilogo cols L-M
    sorted_years = sorted(year_medie.keys())
    tot_r = 2 + len(sorted_years)

    for c, v in [(12,"ANNO"),(13,"IMPORTO")]:
        cell = ws.cell(1, c, v)
        cell.fill = ORANGE; cell.font = BOLD; cell.alignment = CENTER
        if c == 13: cell.number_format = EUR_FMT

    for i, year in enumerate(sorted_years):
        r  = 2 + i
        mr = year_medie[year]
        for c, v in [(12, year),(13, f"=G{mr}")]:
            cell = ws.cell(r, c, v)
            cell.fill = ORANGE_L; cell.font = BOLD; cell.alignment = CENTER
            if c == 13: cell.number_format = EUR_FMT

    for c, v in [(12,"TOTALE"),(13,f"=SUM(M2:M{tot_r-1})")]:
        cell = ws.cell(tot_r, c, v)
        cell.fill = ORANGE; cell.font = BOLD; cell.alignment = CENTER
        if c == 13: cell.number_format = EUR_FMT

    # 6. Fix riferimenti formula (+3) — solo cols A-K, righe 4+
    for r in range(4, ws.max_row + 1):
        for c in range(1, 12):
            cell = ws.cell(r, c)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                new_val = shift_formula(cell.value)
                if new_val != cell.value:
                    cell.value = new_val

    # 7. Fix DIV/0 nelle Medie annue (col G, righe medie)
    for mr in year_medie.values():
        cell = ws.cell(mr, 7)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = fix_div0(cell.value)

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue(), True


# ── Main ──────────────────────────────────────────────────────────────────────

INNER_CONFIGS = {
    "ELABORATI_RICOSTRUITO.zip": (parse_sgv, LUOGO_SGV),
    "ELABORATI_ROMA3.zip":       (parse_r3,  LUOGO_R3),
}

total_ok  = 0
total_skip = 0
total_err  = 0

with zipfile.ZipFile(INPUT_ZIP, "r") as outer_in, \
     zipfile.ZipFile(OUTPUT_ZIP, "w", zipfile.ZIP_DEFLATED) as outer_out:

    for inner_name, (parse_fn, luogo) in INNER_CONFIGS.items():
        print(f"\n{'='*60}\n{inner_name}")
        inner_data    = outer_in.read(inner_name)
        new_inner_buf = io.BytesIO()

        with zipfile.ZipFile(io.BytesIO(inner_data), "r") as inner_in, \
             zipfile.ZipFile(new_inner_buf, "w", zipfile.ZIP_DEFLATED) as inner_out:

            for entry in inner_in.namelist():
                raw   = inner_in.read(entry)
                fname = entry.split("/")[-1]

                if not fname.endswith(".xlsx"):
                    inner_out.writestr(entry, raw)
                    continue

                cognome, nome = parse_fn(fname)
                try:
                    new_raw, processed = process_xlsx(raw, cognome, nome, luogo)
                    inner_out.writestr(entry, new_raw)
                    if processed:
                        total_ok += 1
                        print(f"  OK  {fname}  [{cognome} / {nome}]")
                    else:
                        total_skip += 1
                        print(f"  SKIP {fname}")
                except Exception as e:
                    inner_out.writestr(entry, raw)
                    print(f"  ERR  {fname}: {e}")
                    total_err += 1

        new_inner_buf.seek(0)
        outer_out.writestr(inner_name, new_inner_buf.read())

print(f"\n{'='*60}")
print(f"OK: {total_ok}  |  Skip: {total_skip}  |  Errori: {total_err}")
print(f"Output: {OUTPUT_ZIP}")
