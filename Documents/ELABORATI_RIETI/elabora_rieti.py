# elabora_rieti.py
# Legge i PDF (anche dentro ZIP e 7z) da ARCHIVIO RIETI e produce un Excel
# per lavoratore con le indennita' accessorie dai cedolini e giorni dai cartellini.
#
# Input  : Documents\ARCHIVIO RIETI\{COGNOME NOME}\  (ricerca ricorsiva + archivi)
# Output : Documents\ELABORATI_RIETI\{COGNOME_NOME}.xlsx
# Log    : Documents\ELABORATI_RIETI\LOG_ELABORAZIONE_RIETI.txt

import io
import os
import re
import sys
import zipfile
import datetime
import tempfile
import shutil
from collections import defaultdict

import pdfplumber
import py7zr
from copy import copy
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm

# ── CONFIGURAZIONE ────────────────────────────────────────────────────────────
PATH_INPUT    = r"C:\Users\Gianmarco\Documents\ARCHIVIO RIETI"
PATH_OUTPUT   = r"C:\Users\Gianmarco\Documents\ELABORATI_RIETI"
MODELLO_EXCEL = r"C:\Users\Gianmarco\Documents\sangiovanni\modello.xlsx"
LOG_PATH      = os.path.join(PATH_OUTPUT, "LOG_ELABORAZIONE_RIETI.txt")
EXCLUDE_DIRS  = {"elaborati_rieti"}

FILL_ARANCIO = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_GIALLO  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
FONT_BOLD    = Font(bold=True)
LUOGO_LAVORO = "OSP GEN PROV RIETI"

# Codici indennita' → colonna Excel (B=2 … F=6)
MAPPATURA_IND = {
    2: ['292C', '291C'],
    3: ['219C', '218C', '1861C', '1862C'],
    4: ['213C', '1011C'],
    5: ['239C', '241C', '233C', '1012C'],
    6: ['294C', '293C'],
}
_TUTTI_CODICI = {c: col for col, codici in MAPPATURA_IND.items() for c in codici}
MAX_INDENNITA = 2500.0

COL_GIORNI = 8   # colonna H
COL_FERIE  = 9   # colonna I

MESI_ITA = ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
            "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]


# ── UTILITY ──────────────────────────────────────────────────────────────────

def r2(v):
    try:
        return round(float(str(v).replace(',', '.')), 2)
    except Exception:
        return 0.0


def mese_precedente(anno, mese):
    if mese > 1:
        return anno, mese - 1
    return anno - 1, 12


# ── PARSING FILENAME ─────────────────────────────────────────────────────────

def parse_cedolino_filename(fname):
    """
    Ritorna (anno, mese) o (anno, None) per bundle, o (None, None).
    Formato: cedolino_AREAS_1_XXXXXXXX_YYYYMM_YYYYMM_..._I.pdf
    """
    if fname.lower().startswith("cedolino"):
        base = os.path.splitext(fname)[0]
        parts = base.split("_")
        if len(parts) >= 5:
            s = parts[4]
            if len(s) == 6 and s.isdigit():
                return int(s[:4]), int(s[4:6])
    m = re.match(r'cedolini[\s_]+(\d{4})', fname, re.IGNORECASE)
    if m:
        return int(m.group(1)), None
    return None, None


def parse_cartellino_filename(fname):
    """
    Ritorna (anno, mese) o (anno, None) per bundle, o (None, None).
    Formati supportati:
      NN_HR01_01_YYYY_MM_...pdf       (es. 10_HR01_01_2020_01_...)
      CARTELLINO_YYYY_M_...pdf
      Cartellino_MMYYYY_ID.pdf        (es. Cartellino_012022_100111)
      CARTELLINO Mese YYYY.pdf
      Mese YYYY.pdf                   (es. agosto 2020.pdf – cartellini Tomassetti)
    """
    # NN_HR01_01_YYYY_MM_...
    m = re.match(r'\d+_HR01_01_(\d{4})_(\d{1,2})_', fname)
    if m:
        return int(m.group(1)), int(m.group(2))
    # CARTELLINO_YYYY_M_... (qualsiasi case) — YYYY prima di MM
    m = re.match(r'cartellino_(\d{4})_(\d{1,2})_', fname, re.IGNORECASE)
    if m:
        return int(m.group(1)), int(m.group(2))
    # Cartellino_MMYYYY_ID.pdf — MM prima di YYYY (es. Cartellino_012022_100111)
    m = re.match(r'cartellino_(\d{2})(\d{4})_', fname, re.IGNORECASE)
    if m:
        mm = int(m.group(1))
        yy = int(m.group(2))
        if 1 <= mm <= 12 and 2015 <= yy <= 2030:
            return yy, mm
    # CARTELLINO Mese YYYY
    m = re.match(r'cartellino\s+(\w+)\s+(\d{4})', fname, re.IGNORECASE)
    if m:
        nome = m.group(1).lower()
        if nome in MESI_ITA:
            return int(m.group(2)), MESI_ITA.index(nome) + 1
    # "agosto 2020.pdf" – solo nome mese + anno (cartellini Tomassetti, ecc.)
    mesi_pat = '|'.join(MESI_ITA)
    m = re.match(rf'^({mesi_pat})\s+(\d{{4}})\b', fname, re.IGNORECASE)
    if m:
        nome = m.group(1).lower()
        return int(m.group(2)), MESI_ITA.index(nome) + 1
    return None, None


# ── PARSING CONTENUTO PDF ────────────────────────────────────────────────────

def leggi_rata_da_testo(text):
    """
    Estrae (anno, mese) dalla riga RATA.
    Ritorna (None, None) per tredicesima o se non trovata.
    """
    m = re.search(
        r'RATA:\s*([A-Za-z\u00e0-\u00fc]+)\s*(\d{4})'
        r'(\s*[-\u2013]\s*(Tredicesima|Tredic|13esima|13\xb0))?',
        text, re.IGNORECASE)
    if m:
        if m.group(3):
            return None, None
        nome = m.group(1).strip().lower()
        try:
            return int(m.group(2)), MESI_ITA.index(nome) + 1
        except ValueError:
            pass
    return None, None


def leggi_mese_cartellino_da_testo(text):
    """Estrae (anno, mese) dal testo di un cartellino RIETI."""
    # Formato: "STAMPA CARTELLINO GENNAIO 2021" oppure "Mese 01/2021"
    m = re.search(r'\bMese\s+(\d{2})/(\d{4})\b', text)
    if m:
        mm, yy = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12 and 2015 <= yy <= 2030:
            return yy, mm
    m = re.search(r'\b(\d{2})/(\d{4})\b', text)
    if m:
        mm, yy = int(m.group(1)), int(m.group(2))
        if 1 <= mm <= 12 and 2015 <= yy <= 2030:
            return yy, mm
    for i, nome in enumerate(MESI_ITA):
        m2 = re.search(rf'\b{nome}\s+(\d{{4}})\b', text, re.IGNORECASE)
        if m2:
            return int(m2.group(1)), i + 1
    return None, None


def determina_tipo_da_testo(text):
    if "RATA:" in text:
        return "cedolino"
    if re.search(r'\b\d{2}/\d{4}\b', text) or "CARTELLINO" in text.upper() or "Matricola" in text:
        return "cartellino"
    return None


# ── ESTRAZIONE VALORI ────────────────────────────────────────────────────────

def clean_amount(line):
    m = re.search(r'(\d+[,\.]\d{2})\s*$', line.strip())
    if m:
        raw = m.group(1).replace('.', '').replace(',', '.')
        return r2(raw)
    return 0.0


def estrai_indennita(text, rata_anno, rata_mese):
    """Estrae {(comp_anno, comp_mese): {col: importo}} dal testo cedolino."""
    result = defaultdict(lambda: defaultdict(float))
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        m_cod = re.match(r'^(\d{1,4}C?)\b', line)
        if not m_cod:
            continue
        codice = m_cod.group(1).upper()
        col = _TUTTI_CODICI.get(codice)
        if col is None:
            continue
        val = clean_amount(line)
        if val == 0.0:
            continue
        rif = re.search(r'Rif\s*(\d{2})[/\s](\d{2,4})', line, re.IGNORECASE)
        if rif:
            comp_m = int(rif.group(1))
            comp_y = int(rif.group(2))
            if comp_y < 100:
                comp_y += 2000
        else:
            comp_y, comp_m = mese_precedente(rata_anno, rata_mese)
        result[(comp_y, comp_m)][col] = r2(result[(comp_y, comp_m)][col] + val)
    return result


def estrai_giorni_ferie(text):
    """Conta giorni lavorati e ferie dal testo cartellino."""
    giorni, ferie = 0, 0
    for line in text.split('\n'):
        if re.match(r'^\d{2}\s+(LU|MA|ME|GI|VE|SA|DO)', line):
            if re.search(r'[EU]\d{2}[:\.]\d{2}', line):
                giorni += 1
            if re.search(r'DFER', line, re.IGNORECASE):
                ferie += 1
    return giorni, ferie


# ── PROCESSO PDF SINGOLO (bytes o path) ────────────────────────────────────

def _processa_pdf(fname, pdf_bytes_or_path, folder_hint, ind_data, cart_data, avvisi, seen=None):
    """Legge e processa un PDF (da bytes in memoria o da percorso file)."""
    if seen is not None:
        if fname in seen:
            return
        seen.add(fname)

    fn_anno_ced,  fn_mese_ced  = parse_cedolino_filename(fname)
    fn_anno_cart, fn_mese_cart = parse_cartellino_filename(fname)

    try:
        if isinstance(pdf_bytes_or_path, (bytes, bytearray)):
            opener = pdfplumber.open(io.BytesIO(pdf_bytes_or_path))
        else:
            opener = pdfplumber.open(pdf_bytes_or_path)
        with opener as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
    except Exception as e:
        avvisi.append(f"  ERRORE lettura {fname}: {e}")
        return

    filename_completo = (
        (fn_anno_ced  is not None and fn_mese_ced  is not None) or
        (fn_anno_cart is not None and fn_mese_cart is not None)
    )
    is_bundle = not filename_completo and (
        (fn_anno_ced is not None and fn_mese_ced is None) or
        (fn_anno_cart is not None and fn_mese_cart is None) or
        (fn_anno_ced is None and fn_anno_cart is None and len(pages_text) > 1)
    )

    if is_bundle:
        for page_text in pages_text:
            tipo = folder_hint or determina_tipo_da_testo(page_text)
            if tipo == "cedolino":
                anno_p, mese_p = leggi_rata_da_testo(page_text)
                if anno_p and mese_p:
                    dati = estrai_indennita(page_text, anno_p, mese_p)
                    for k, vd in dati.items():
                        for col, v in vd.items():
                            ind_data[k][col] = r2(ind_data[k][col] + v)
            elif tipo == "cartellino":
                anno_p, mese_p = leggi_mese_cartellino_da_testo(page_text)
                if anno_p and mese_p and (anno_p, mese_p) not in cart_data:
                    g, f = estrai_giorni_ferie(page_text)
                    cart_data[(anno_p, mese_p)] = (g, f)
        return

    # Documento singolo (anche se multi-pagina, es. cedolino AREAS 3pp)
    full_text = "\n".join(pages_text)
    tipo = None
    anno = None
    mese = None

    if folder_hint == "cedolino" or fn_anno_ced is not None:
        tipo = "cedolino"
        anno, mese = leggi_rata_da_testo(full_text)
        if anno is None:
            anno, mese = fn_anno_ced, fn_mese_ced
    elif folder_hint == "cartellino" or fn_anno_cart is not None:
        tipo = "cartellino"
        anno, mese = leggi_mese_cartellino_da_testo(full_text)
        if anno is None:
            anno, mese = fn_anno_cart, fn_mese_cart
    else:
        tipo = determina_tipo_da_testo(full_text)
        if tipo == "cedolino":
            anno, mese = leggi_rata_da_testo(full_text)
        elif tipo == "cartellino":
            anno, mese = leggi_mese_cartellino_da_testo(full_text)

    if tipo is None or anno is None:
        avvisi.append(f"  SKIP (non classificato): {fname}")
        return
    if mese is None:
        avvisi.append(f"  SKIP (mese non trovato): {fname}")
        return

    if tipo == "cedolino":
        dati = estrai_indennita(full_text, anno, mese)
        for k, vd in dati.items():
            for col, v in vd.items():
                ind_data[k][col] = r2(ind_data[k][col] + v)
        for k, vd in dati.items():
            tot = sum(vd.values())
            if tot > MAX_INDENNITA:
                avvisi.append(f"  ANOMALIA importo elevato {fname}: comp {k[0]}/{k[1]:02d} tot={tot:.2f}")
    elif tipo == "cartellino":
        g, f = estrai_giorni_ferie(full_text)
        if (anno, mese) not in cart_data:
            cart_data[(anno, mese)] = (g, f)


# ── PROCESSO LAVORATORE ───────────────────────────────────────────────────────

def _hint_da_nome(name, default_hint):
    """Deriva il tipo di documento dal nome di cartella o file archivio."""
    u = name.upper()
    if 'CEDOLIN' in u or 'BUSTE PAGA' in u or 'BUSTE' in u:
        return 'cedolino'
    if 'CARTELL' in u or 'TABULAT' in u:
        return 'cartellino'
    return default_hint


def _processa_zip(fpath, fname, folder_hint, ind_data, cart_data, avvisi, depth=0, seen=None):
    """Apre un file ZIP (da percorso o file-like) ed elabora il contenuto, ricorsivo su zip/7z annidati."""
    if depth > 3:
        return
    hint = _hint_da_nome(fname, folder_hint)
    try:
        with zipfile.ZipFile(fpath, 'r') as zf:
            for entry in zf.namelist():
                low = entry.lower()
                if not (low.endswith('.pdf') or low.endswith('.zip') or low.endswith('.7z')):
                    continue
                bname = os.path.basename(entry)
                data  = zf.read(entry)
                if low.endswith('.pdf'):
                    _processa_pdf(bname, data, hint, ind_data, cart_data, avvisi, seen)
                elif low.endswith('.zip'):
                    _processa_zip(io.BytesIO(data), bname, hint, ind_data, cart_data, avvisi, depth + 1, seen)
                else:  # .7z
                    with tempfile.NamedTemporaryFile(suffix='.7z', delete=False) as tmp:
                        tmp.write(data)
                        tmp_path = tmp.name
                    try:
                        _processa_7z(tmp_path, bname, hint, ind_data, cart_data, avvisi, depth + 1, seen)
                    finally:
                        os.unlink(tmp_path)
    except Exception as e:
        avvisi.append(f"  ERRORE ZIP {fname}: {e}")


def _processa_7z(fpath, fname, folder_hint, ind_data, cart_data, avvisi, depth=0, seen=None):
    """Estrae un 7z in cartella temporanea ed elabora il contenuto ricorsivamente."""
    if depth > 3:
        return
    hint = _hint_da_nome(fname, folder_hint)
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            with py7zr.SevenZipFile(fpath, mode='r') as sz:
                sz.extractall(path=tmpdir)
            # Elabora tutto il contenuto estratto
            for root2, dirs2, files2 in os.walk(tmpdir):
                h2 = _hint_da_nome(os.path.basename(root2), hint)
                for fn2 in sorted(files2):
                    fp2  = os.path.join(root2, fn2)
                    ext2 = fn2.lower()
                    if ext2.endswith('.pdf'):
                        _processa_pdf(fn2, fp2, h2, ind_data, cart_data, avvisi, seen)
                    elif ext2.endswith('.zip'):
                        _processa_zip(fp2, fn2, h2, ind_data, cart_data, avvisi, depth + 1, seen)
                    elif ext2.endswith('.7z'):
                        _processa_7z(fp2, fn2, h2, ind_data, cart_data, avvisi, depth + 1, seen)
    except Exception as e:
        avvisi.append(f"  ERRORE 7z {fname}: {e}")


def processa_lavoratore(worker_path, worker_name, log_lines):
    ind_data  = defaultdict(lambda: defaultdict(float))
    cart_data = {}
    avvisi    = []
    # stessa PDF (per basename) raggiungibile da più percorsi ridondanti
    # (zip annidato + copia già estratta su disco): processarla una sola volta.
    seen = set()

    for root, dirs, files in os.walk(worker_path):
        folder_name = os.path.basename(root).upper()
        if "CEDOLIN" in folder_name:
            folder_hint = "cedolino"
        elif "CARTELL" in folder_name or "TABULAT" in folder_name:
            folder_hint = "cartellino"
        else:
            folder_hint = None

        for fname in sorted(files):
            fpath = os.path.join(root, fname)
            ext   = fname.lower()

            if ext.endswith('.pdf'):
                _processa_pdf(fname, fpath, folder_hint, ind_data, cart_data, avvisi, seen)

            elif ext.endswith('.zip'):
                _processa_zip(fpath, fname, folder_hint, ind_data, cart_data, avvisi, seen=seen)

            elif ext.endswith('.7z'):
                # Salta il 7z se esiste già una cartella con lo stesso nome (già estratto)
                stem = os.path.splitext(fname)[0]
                if os.path.isdir(os.path.join(root, stem)):
                    continue
                _processa_7z(fpath, fname, folder_hint, ind_data, cart_data, avvisi, seen=seen)

    if avvisi:
        log_lines.append(f"\n[ {worker_name} ]")
        log_lines.extend(avvisi)

    return ind_data, cart_data


# ── SCRITTURA EXCEL ───────────────────────────────────────────────────────────

def scrivi_excel(worker_name, ind_data, cart_data, log_lines):
    all_keys = set(ind_data.keys()) | set(cart_data.keys())
    if not all_keys:
        log_lines.append(f"\n[ {worker_name} ] NESSUN DATO — file non creato")
        return False

    anni = sorted(set(k[0] for k in all_keys))

    wb_mod = load_workbook(MODELLO_EXCEL)
    ws_mod = wb_mod.active
    wb_new = load_workbook(MODELLO_EXCEL)
    ws_new = wb_new.active
    ws_new.delete_rows(1, ws_new.max_row + 10)

    # Dati anagrafici in alto (riga 1 header, riga 2 valori) — nome cartella "COGNOME NOME"
    cognome, _, nome = worker_name.partition(" ")
    ws_new.cell(row=1, column=1, value="NOME")
    ws_new.cell(row=1, column=2, value="COGNOME")
    ws_new.cell(row=1, column=3, value="LUOGO DI LAVORO")
    ws_new.cell(row=1, column=12, value="ANNO")
    ws_new.cell(row=1, column=13, value="IMPORTO")
    ws_new.cell(row=2, column=1, value=nome)
    ws_new.cell(row=2, column=2, value=cognome)
    ws_new.cell(row=2, column=3, value=LUOGO_LAVORO)
    for c in (1, 2, 3, 12, 13):
        cell = ws_new.cell(row=1, column=c)
        cell.fill, cell.font = FILL_ARANCIO, FONT_BOLD
    for c in (1, 2, 3):
        cell = ws_new.cell(row=2, column=c)
        cell.fill, cell.font = FILL_GIALLO, FONT_BOLD

    curr_row = 3
    for i, anno in enumerate(anni):
        cell = ws_new.cell(row=curr_row, column=1, value=f"ANNO {anno}")
        cell.fill, cell.font = FILL_ARANCIO, FONT_BOLD
        for r in range(1, 16):
            for c in range(1, 11):
                src = ws_mod.cell(row=r, column=c)
                dst = ws_new.cell(row=curr_row + r, column=c)
                if src.has_style:
                    dst.font          = copy(src.font)
                    dst.border        = copy(src.border)
                    dst.fill          = copy(src.fill)
                    dst.number_format = copy(src.number_format)
                    dst.protection    = copy(src.protection)
                    dst.alignment     = copy(src.alignment)
                if isinstance(src.value, str) and src.value.startswith('='):
                    dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
                else:
                    dst.value = src.value
                if 2 <= r <= 13:
                    mese = r - 1
                    col_vals = {1: MESI_ITA[mese - 1].capitalize()}
                    for excel_col in MAPPATURA_IND:
                        col_vals[excel_col] = r2(ind_data.get((anno, mese), {}).get(excel_col, 0))
                    g, f = cart_data.get((anno, mese), (0, 0))
                    if (anno, mese) not in cart_data and (anno, mese) in ind_data:
                        g = 22.5  # cedolino presente ma cartellino mancante: giorni pieni convenzionali
                    col_vals[COL_GIORNI] = g
                    col_vals[COL_FERIE]  = f
                    if c in col_vals:
                        dst.value = col_vals[c]
        # Specchietto riassuntivo: riga block_start+15 = "Medie annue" (colonna G) dell'anno i-esimo
        c_anno    = ws_new.cell(row=2 + i, column=12, value=anno)
        c_importo = ws_new.cell(row=2 + i, column=13, value=f"=G{18 + 17 * i}")
        c_anno.fill, c_anno.font       = FILL_GIALLO, FONT_BOLD
        c_importo.fill, c_importo.font = FILL_GIALLO, FONT_BOLD
        curr_row += 17

    n = len(anni)
    c_tot_lbl = ws_new.cell(row=2 + n, column=12, value="TOTALE")
    c_tot_val = ws_new.cell(row=2 + n, column=13, value=f"=SUM(M2:M{1 + n})")
    c_tot_lbl.fill, c_tot_lbl.font = FILL_ARANCIO, FONT_BOLD
    c_tot_val.fill, c_tot_val.font = FILL_ARANCIO, FONT_BOLD

    safe_name = worker_name.replace(" ", "_").replace("/", "-")
    out_path = os.path.join(PATH_OUTPUT, f"{safe_name}.xlsx")
    wb_new.save(out_path)
    return True


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(PATH_OUTPUT, exist_ok=True)

    # Supporta --solo NOME1 "NOME 2" ... per rielaborare solo certi lavoratori
    solo_set = set()
    args = sys.argv[1:]
    if args and args[0].lower() == '--solo':
        solo_set = {a.upper() for a in args[1:]}

    workers = sorted([
        d for d in os.listdir(PATH_INPUT)
        if os.path.isdir(os.path.join(PATH_INPUT, d))
        and d.lower() not in EXCLUDE_DIRS
        and not d.lower().startswith("elaborati")
        and (not solo_set or d.upper() in solo_set)
    ])

    print(f"Lavoratori trovati: {len(workers)}")
    print(f"Output -> {PATH_OUTPUT}\n")

    log_lines = [
        "LOG ELABORAZIONE ASL RIETI",
        f"Generato: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60,
    ]

    senza_dati = 0
    for w in tqdm(workers, desc="Elaborazione"):
        worker_path = os.path.join(PATH_INPUT, w)
        ind_data, cart_data = processa_lavoratore(worker_path, w, log_lines)
        ok = scrivi_excel(w, ind_data, cart_data, log_lines)
        n_ced  = len(ind_data)
        n_cart = len(cart_data)
        status = "OK" if ok else "NESSUN DATO"
        if not ok:
            senza_dati += 1
        tqdm.write(f"  {w:<40} {status}  ({n_ced} ced, {n_cart} cart)")

    log_lines += [
        "",
        "=" * 60,
        f"COMPLETATO: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Lavoratori elaborati: {len(workers)}",
        f"Senza dati: {senza_dati}",
    ]

    with open(LOG_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    print(f"\nLog -> {LOG_PATH}")


if __name__ == "__main__":
    main()
