#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Corregge i riferimenti di riga nelle formule dei file già modificati
(intestazione 3 righe inserita). Aggiunge +3 a tutti i riferimenti
relativi di riga, escludendo le colonne L-M (tabella riepilogo già corretta).
"""
import os, sys, re
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SANGIOVANNI = r"C:\Users\Gianmarco\Documents\ELABORATI sangiovanni"
ROMA3       = r"C:\Users\Gianmarco\Documents\ELABORATI_ROMA3"

SHIFT = 3  # righe inserite all'inizio


def shift_formula(formula, shift=SHIFT):
    """Aggiunge shift a tutti i riferimenti di riga relativi (no $) in una formula Excel."""
    def replace_ref(m):
        col_abs = m.group(1)  # '$' or ''
        col     = m.group(2)  # lettere colonna
        row_abs = m.group(3)  # '$' or ''
        row     = int(m.group(4))
        if row_abs == '$':
            return m.group(0)  # riferimento assoluto di riga: non modificare
        return f"{col_abs}{col}{row + shift}"
    return re.sub(r'(\$?)([A-Z]+)(\$?)(\d+)', replace_ref, formula)


def fix_formulas(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Verifica che sia stato elaborato (R1 = "NOME")
    if ws.cell(1, 1).value != "NOME":
        print(f"  SKIP (non elaborato): {os.path.basename(filepath)}")
        wb.close()
        return False

    changed = 0
    # Solo righe dati (4+), solo colonne A-K (escludi L=12 e M=13 già corrette)
    for r in range(4, ws.max_row + 1):
        for c in range(1, 12):  # colonne A(1) → K(11)
            cell = ws.cell(r, c)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                new_val = shift_formula(cell.value)
                if new_val != cell.value:
                    cell.value = new_val
                    changed += 1

    wb.save(filepath)
    wb.close()
    return changed


errors = []
total_files = 0
total_changes = 0

for folder in [SANGIOVANNI, ROMA3]:
    print(f"\n{'='*60}")
    print(f"Cartella: {os.path.basename(folder)}")
    print(f"{'='*60}")
    files = sorted([
        f for f in os.listdir(folder)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ])
    for fname in files:
        fpath = os.path.join(folder, fname)
        try:
            n = fix_formulas(fpath)
            if n is not False:
                total_files += 1
                total_changes += n
                print(f"  OK  ({n:3d} formule corrette): {fname}")
        except Exception as e:
            print(f"  ERR {fname}: {e}")
            errors.append(fname)

print(f"\n{'='*60}")
print(f"Completato: {total_files} file, {total_changes} formule corrette")
if errors:
    print(f"Errori: {errors}")
