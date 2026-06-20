# -*- coding: utf-8 -*-
"""
Elaborazione buoni pasto / indennita Regione Lazio - metodologia "Sangiovanni".
Lavoratore: COLETTI AMBRA (Agenzia Regionale di Protezione Civile, Regione Lazio).
Fonte dati: cedolini PDF in E:\REGIONE LAZIO\cedolini 20XX\
Indennita tracciate (stesse categorie concettuali della Polizia Locale, cfr. annistampa.xlsx):
  B: Ind. Di turno diurno          <- AA06/E1GM
  C: Ind rep turno                 <- pattern REPERIB/REP.FEST (stile PL, 0 per Coletti)
  D: IND turno fest/nott           <- AA06/E1GP
  E: IND Turno nott o Fest.        <- AA06/E1GN, AA06/E1HN
  F: ind di rischio HH             <- AA06/E1GS (CONDIZIONILAVORO_2)
  G: Ind pronta disponibilita      <- pattern Dis.PL/ServizioEstPL (stile PL, 0 per Coletti)
  H: IND servizio esterno          <- pattern ServizioEsternoPL (stile PL, 0 per Coletti)
"""
import re
import glob
import os
from collections import defaultdict
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = r"E:\REGIONE LAZIO"
OUT_PATH = r"C:\Users\Gianmarco\Documents\REGIONE LAZIO\COLETTI_AMBRA.xlsx"

COLS = ["B", "C", "D", "E", "F", "G", "H"]
HEADERS = {
    "B": "Ind. Di turno diurno",
    "C": "Ind rep turno",
    "D": "IND turno fest/nott",
    "E": "IND Turno nott o Fest.",
    "F": "ind di rischio HH",
    "G": "Ind pronta disponibilità",
    "H": "IND servizio esterno",
}

CODE_MAP = {
    "AA06/E1GM": "B",
    "AA06/E1GP": "D",
    "AA06/E1GN": "E",
    "AA06/E1HN": "E",
    "AA06/E1GS": "F",
}

TEXT_PATTERNS = {
    "C": re.compile(r"REPERIB|REP\.FEST", re.IGNORECASE),
    "G": re.compile(r"Dis\.PL|ServizioEstPL", re.IGNORECASE),
    "H": re.compile(r"ServizioEsternoPL", re.IGNORECASE),
}

LINE_RE = re.compile(r"^([A-Z0-9]+/[A-Z0-9]+)\s+(.+?)\s+(-?[\d.]+,\d{2})$")
RIF_RE = re.compile(r"(?:Rif|scad)\.(\d{2})/(\d{4})")
FILENAME_RE = re.compile(r"cedolino_SPT_1_\d+_(\d{4})(\d{2})_")


def it_to_float(s):
    return float(s.replace(".", "").replace(",", "."))


def classify(codice, descrizione):
    if codice in CODE_MAP:
        return CODE_MAP[codice]
    for col, pat in TEXT_PATTERNS.items():
        if pat.search(descrizione):
            return col
    return None


def rata_da_filename(path):
    m = FILENAME_RE.search(os.path.basename(path))
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def estrai_da_pdf(path, dati):
    rata_anno, rata_mese = rata_da_filename(path)
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                line = line.strip()
                m = LINE_RE.match(line)
                if not m:
                    continue
                codice, descrizione, importo_s = m.groups()
                col = classify(codice, descrizione)
                if col is None:
                    continue
                rif = RIF_RE.search(descrizione)
                if rif:
                    mese, anno = int(rif.group(1)), int(rif.group(2))
                else:
                    anno, mese = rata_anno, rata_mese
                if anno is None or mese is None:
                    continue
                importo = it_to_float(importo_s)
                dati[(anno, mese)][col] += importo


def raccogli_dati():
    dati = defaultdict(lambda: defaultdict(float))
    cartelle = sorted(glob.glob(os.path.join(BASE_DIR, "cedolini 20*")))
    cartelle = [c for c in cartelle if os.path.isdir(c)]
    n_pdf = 0
    for cartella in cartelle:
        for pdf_path in sorted(glob.glob(os.path.join(cartella, "*.pdf"))):
            estrai_da_pdf(pdf_path, dati)
            n_pdf += 1
    print(f"PDF elaborati: {n_pdf}")
    anni = sorted(set(a for a, m in dati.keys()))
    print(f"Anni con dati: {anni}")
    return dati


# ---------------------------------------------------------------------------
# Generazione Excel
# ---------------------------------------------------------------------------

def col_letter_to_idx(letter):
    return ord(letter) - ord("A") + 1


def genera_excel(dati):
    wb = Workbook()
    ws = wb.active
    ws.title = "COLETTI AMBRA"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header dati personali
    ws["A1"] = "Nome"
    ws["B1"] = "Cognome"
    ws["C1"] = "Luogo lavoro "
    ws["D1"] = " "
    for c in ("A1", "B1", "C1", "D1"):
        ws[c].font = bold
    ws["A2"] = "AMBRA"
    ws["B2"] = "COLETTI"
    ws["C2"] = "REGIONE LAZIO"

    ws["C3"] = "Ferie fruite annue"
    ws["D3"] = 28
    ws["C4"] = "Giorni lavorati"
    ws["D4"] = "=(365-D3)"

    anni = sorted(set(a for a, m in dati.keys()))
    if not anni:
        anni = []

    MESI_IT = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
               "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

    block_rows = {}  # anno -> (header_row, somme_row, medie_row)
    start_row = 6
    for i, anno in enumerate(anni):
        header_row = start_row + i * 16
        mesi_rows = list(range(header_row + 1, header_row + 13))
        somme_row = header_row + 13
        medie_row = header_row + 14
        block_rows[anno] = (header_row, mesi_rows, somme_row, medie_row)

        ws.cell(row=header_row, column=1, value=anno).font = bold
        for col in COLS:
            cell = ws[f"{col}{header_row}"]
            cell.value = HEADERS[col]
            cell.font = bold
            cell.fill = header_fill
        ws.cell(row=header_row, column=col_letter_to_idx("I"), value="Sommatoria precedenti").font = bold
        ws[f"I{header_row}"].fill = header_fill

        for j, mese_row in enumerate(mesi_rows, start=1):
            ws.cell(row=mese_row, column=1, value=MESI_IT[j - 1])
            for col in COLS:
                val = dati.get((anno, j), {}).get(col, 0)
                ws[f"{col}{mese_row}"] = round(val, 2)
            ws[f"I{mese_row}"] = f"=SUM(B{mese_row}:C{mese_row}:D{mese_row}:E{mese_row}:G{mese_row}:H{mese_row})"

        ws.cell(row=somme_row, column=1, value="Somme annuali").font = bold
        for col in COLS + ["I"]:
            refs = ",".join(f"{col}{r}" for r in mesi_rows)
            ws[f"{col}{somme_row}"] = f"=SUM({refs})"
            ws[f"{col}{somme_row}"].font = bold

        ws.cell(row=medie_row, column=1, value="Medie annue").font = bold
        for col in COLS + ["I"]:
            ws[f"{col}{medie_row}"] = f"=({col}{somme_row}/337*28)"
            ws[f"{col}{medie_row}"].font = bold

    # Tabella riassuntiva K/L accanto al primo blocco
    if anni:
        first_header_row = block_rows[anni[0]][0]
        ws.cell(row=first_header_row, column=col_letter_to_idx("K"), value="Anno").font = bold
        ws.cell(row=first_header_row, column=col_letter_to_idx("L"), value="Importo").font = bold
        for i, anno in enumerate(anni):
            r = first_header_row + 1 + i
            medie_row = block_rows[anno][3]
            ws.cell(row=r, column=col_letter_to_idx("K"), value=anno)
            ws.cell(row=r, column=col_letter_to_idx("L"), value=f"=I{medie_row}")
        totale_row = first_header_row + 1 + len(anni)
        ws.cell(row=totale_row, column=col_letter_to_idx("K"), value="Totale:").font = bold
        l_first = first_header_row + 1
        l_last = first_header_row + len(anni)
        ws.cell(row=totale_row, column=col_letter_to_idx("L"),
                value=f"=SUBTOTAL(109,L{l_first}:L{l_last})").font = bold

    ws.column_dimensions["A"].width = 14
    for col in COLS:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 14

    wb.save(OUT_PATH)
    print(f"Salvato: {OUT_PATH}")


if __name__ == "__main__":
    dati = raccogli_dati()
    genera_excel(dati)
