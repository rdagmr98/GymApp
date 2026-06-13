#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Processa TIVOLI 12062026: integra nuovi cartellini negli Excel esistenti (11062026).
Logica:
  - Cerca Excel esistente in tivoli11062026/buonipasto DEFINITIVO/{LAVORATORE}.xlsx
  - Legge i mesi già presenti nell'Excel esistente
  - Processa i PDF nel ZIP (tutti i formati: OLD, OLD_REV, NEW)
  - Aggiunge solo i mesi NUOVI (non già presenti) all'Excel
  - Se nessun Excel esistente: crea da zero con tutti i dati trovati
  - Supporta formato invertito RTL (Meucci 2013-2022)
"""
import os, sys, re, zipfile, io, tempfile
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\Gianmarco\Documents\tivoli")
import analisi_tivoli, openpyxl

DOWNLOADS    = r"C:\Users\Gianmarco\Downloads"
ZIP_DIR      = os.path.join(DOWNLOADS, "tivoli12062026")
EXISTING_DIR = os.path.join(DOWNLOADS, "tivoli11062026", "buonipasto DEFINITIVO")
OUT_ROOT     = os.path.join(DOWNLOADS, "tivoli12062026_output")
DEFINITIVO   = os.path.join(OUT_ROOT, "buonipasto DEFINITIVO")
OUT_ZIP      = os.path.join(DOWNLOADS, "tivoli12062026_DEFINITIVO.zip")

BUONO_ORE  = 0.5
BUONO_EURO = 4.13

MONTH_ABBR = {'Gen':1,'Feb':2,'Mar':3,'Apr':4,'Mag':5,'Giu':6,
              'Lug':7,'Ago':8,'Set':9,'Ott':10,'Nov':11,'Dic':12}

os.makedirs(DEFINITIVO, exist_ok=True)
analisi_tivoli.OUTPUT_DIR = DEFINITIVO
analisi_tivoli.LOG_FILE   = os.path.join(OUT_ROOT, "pdf_non_leggibili.log")
os.makedirs(OUT_ROOT, exist_ok=True)
with open(analisi_tivoli.LOG_FILE, 'w', encoding='utf-8') as f:
    f.write('PDF NON LEGGIBILI - Tivoli 12062026\n\n')


# ── Legge maturati/erogati dall'Excel esistente ───────────────────────────────

def read_existing_excel(xlsx_path, worker_name):
    """Ritorna lista record {year, month, maturati, erogati, details=[]} dal file esistente."""
    records = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            try:
                year = int(sn)
            except ValueError:
                continue
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == 'TOTALE MESE':
                    mese = row[0]
                    mat  = row[5]
                    ero  = row[6]
                    if mese in MONTH_ABBR:
                        records.append({
                            'worker': worker_name,
                            'year': year,
                            'month': MONTH_ABBR[mese],
                            'maturati': int(mat) if isinstance(mat, (int, float)) else 0,
                            'erogati': int(ero) if isinstance(ero, (int, float)) else 0,
                            'erogati_found': True,
                            'details': [],
                        })
        wb.close()
    except Exception as e:
        print(f"  ⚠ Errore lettura Excel esistente {os.path.basename(xlsx_path)}: {e}")
    return records


# ── Legge erogati dall'XLSX dentro il ZIP ────────────────────────────────────

def read_erogati_from_zip_xlsx(xlsx_bytes):
    """Ritorna {(year, month): erogati} dai fogli anno dell'XLSX nel ZIP."""
    result = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        for sn in wb.sheetnames:
            try:
                year = int(sn)
            except ValueError:
                continue
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == 'TOTALE MESE':
                    mese = row[0]
                    ero  = row[6]
                    if mese in MONTH_ABBR and isinstance(ero, (int, float)) and ero > 0:
                        result[(year, MONTH_ABBR[mese])] = int(ero)
        wb.close()
    except Exception:
        pass
    return result


# ── Processa un outer ZIP ─────────────────────────────────────────────────────

def collect_pdfs_from_zip(zf):
    """Raccoglie tutti i PDF da un ZipFile (ricorsivo su un livello di inner ZIP)."""
    names = zf.namelist()
    pdf_bytes = {}
    # PDF diretti
    for n in names:
        if n.lower().endswith('.pdf'):
            pdf_bytes[n] = zf.read(n)
    # Inner ZIP (un solo livello)
    for n in names:
        if n.lower().endswith('.zip'):
            try:
                inner_data = zf.read(n)
                with zipfile.ZipFile(io.BytesIO(inner_data)) as inner:
                    for pn in inner.namelist():
                        if pn.lower().endswith('.pdf'):
                            key = f"{n}/{pn}"
                            pdf_bytes[key] = inner.read(pn)
            except Exception:
                pass
    return pdf_bytes


def process_worker_zip(zip_path):
    fname = os.path.basename(zip_path)
    print(f"\n{'='*60}\n{fname}")

    try:
        with zipfile.ZipFile(zip_path) as outer:
            names = outer.namelist()

            # Nome canonico dal file XLSX nel ZIP
            xlsx_names = [n for n in names if n.lower().endswith('.xlsx')]
            if xlsx_names:
                worker_name = os.path.splitext(os.path.basename(xlsx_names[0]))[0].strip().upper()
                xlsx_data   = outer.read(xlsx_names[0])
                erogati_map = read_erogati_from_zip_xlsx(xlsx_data)
            else:
                worker_name = None
                erogati_map = {}

            pdf_bytes = collect_pdfs_from_zip(outer)

    except Exception as e:
        print(f"  ERRORE apertura ZIP: {e}")
        return None

    if not pdf_bytes:
        print("  SKIP: nessun PDF trovato")
        return None

    print(f"  Lavoratore: {worker_name or '(da PDF)'}  |  PDF: {len(pdf_bytes)}")

    # Processa tutti i PDF
    new_records = []
    for pname, pdata in sorted(pdf_bytes.items()):
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
            f.write(pdata); tmp = f.name
        try:
            recs = analisi_tivoli.process_pdf(tmp)
        except Exception as e:
            print(f"    ERRORE {os.path.basename(pname)}: {e}")
            recs = []
        finally:
            os.unlink(tmp)

        for r in recs:
            if worker_name:
                r['worker'] = worker_name
            # Prendi erogati dall'XLSX se presenti e i PDF non li hanno
            key = (r['year'], r['month'])
            if not r.get('erogati_found') and key in erogati_map:
                r['erogati'] = erogati_map[key]
                r['erogati_found'] = True
        new_records.extend(recs)

    # Se non c'era XLSX, prendi nome dal record più frequente nei PDF
    if not worker_name and new_records:
        from collections import Counter as _C
        worker_name = _C(r['worker'] for r in new_records).most_common(1)[0][0]
        print(f"  Nome da PDF: {worker_name}")

    if not new_records:
        print("  SKIP: nessun dato estratto dai PDF")
        return None

    # Statistiche PDF
    new_mat = sum(r['maturati'] for r in new_records)
    new_anni = sorted({r['year'] for r in new_records})
    print(f"  Da PDF: mat={new_mat}  anni={new_anni}")

    # Cerca Excel esistente in tivoli11062026
    existing_path = os.path.join(EXISTING_DIR, f"{worker_name}.xlsx")
    if os.path.exists(existing_path):
        existing_records = read_existing_excel(existing_path, worker_name)
        existing_keys    = {(r['year'], r['month']) for r in existing_records}

        truly_new = [r for r in new_records if (r['year'], r['month']) not in existing_keys]

        if not truly_new:
            print(f"  SKIP: nessun mese nuovo (esistenti: {sorted(existing_keys)})")
            return None

        merged = existing_records + truly_new
        new_ym = sorted({(r['year'], r['month']) for r in truly_new})
        print(f"  Nuovi mesi aggiunti: {new_ym}")
    else:
        merged = new_records
        print(f"  Nessun Excel esistente → nuovo Excel da zero")

    # Crea Excel integrato
    out_path, _ = analisi_tivoli.create_worker_excel(worker_name, merged, DEFINITIVO)

    total_mat = sum(r['maturati'] for r in merged)
    total_ero = sum(r['erogati'] for r in merged)
    delta     = total_mat - total_ero
    anni      = sorted({r['year'] for r in merged})
    print(f"  Salvato: {os.path.basename(out_path)}")
    print(f"  Totale: mat={total_mat} ero={total_ero} delta={delta} ({delta*BUONO_EURO:.2f}€)  anni={anni}")
    return out_path


# ── Main ─────────────────────────────────────────────────────────────────────

print("=" * 60)
print("TIVOLI 12062026 - Integrazione cartellini")
print("=" * 60)

zip_files = sorted(
    os.path.join(ZIP_DIR, f)
    for f in os.listdir(ZIP_DIR)
    if f.lower().endswith('.zip')
)
print(f"ZIP trovati: {len(zip_files)}")

output_files = []
for zf in zip_files:
    out = process_worker_zip(zf)
    if out:
        output_files.append(out)

# ── RIEPILOGO GENERALE ────────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
TOT_FILL = PatternFill("solid", fgColor="BDD7EE")
TOT_FONT = Font(bold=True)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
NEG_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN     = Side(style="thin", color="CCCCCC")
BORDER_R = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER_R = Alignment(horizontal="center", vertical="center")
LEFT_R   = Alignment(horizontal="left",   vertical="center")

def _rh(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER_R, BORDER_R

def _rc(ws, r, c, v, bold=False, fill=None, align=CENTER_R):
    cell = ws.cell(row=r, column=c, value=v)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    cell.alignment = align
    cell.border = BORDER_R
    return cell

def _rf(ws, r, c, formula, bold=False, fill=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=formula)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    cell.alignment = CENTER_R
    cell.border = BORDER_R
    return cell


def build_riepilogo(src_dir, out_path):
    agg     = defaultdict(lambda: defaultdict(int))
    agg_ero = defaultdict(lambda: defaultdict(int))
    xlsx_files = sorted(
        f for f in os.listdir(src_dir)
        if f.endswith(".xlsx") and not f.startswith("RIEPILOGO")
    )
    print(f"  Lettura {len(xlsx_files)} file da {os.path.basename(src_dir)}...")
    for fname in xlsx_files:
        worker = fname[:-5]
        path   = os.path.join(src_dir, fname)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name == "Riepilogo":
                    continue
                try:
                    year = int(sheet_name)
                except ValueError:
                    continue
                ws_y = wb[sheet_name]
                for row in ws_y.iter_rows(min_row=2, values_only=True):
                    if row[1] == "TOTALE MESE":
                        mat = row[5] if isinstance(row[5], (int, float)) else 0
                        ero = row[6] if isinstance(row[6], (int, float)) else 0
                        agg[worker][year]     += max(0, int(mat) - int(ero))
                        agg_ero[worker][year] += int(ero)
            wb.close()
        except Exception as e:
            print(f"    ERRORE {fname}: {e}")

    workers = sorted(agg.keys())
    years   = sorted({y for w in agg.values() for y in w})
    if not workers:
        print("  Nessun dato per il riepilogo.")
        return None
    print(f"  Lavoratori: {len(workers)},  Anni: {years[0]}–{years[-1]}")

    wb2 = openpyxl.Workbook()
    ws  = wb2.active
    ws.title = "Riepilogo Generale"
    _rh(ws, 1, 1, "Lavoratore")
    for ci, y in enumerate(years, 2):
        _rh(ws, 1, ci, y)
    tot_col   = len(years) + 2
    eur_col   = len(years) + 3
    ore_col   = len(years) + 4
    tot_col_l = openpyxl.utils.get_column_letter(tot_col)
    last_yr_l = openpyxl.utils.get_column_letter(len(years) + 1)
    ore_col_l = openpyxl.utils.get_column_letter(ore_col)
    _rh(ws, 1, tot_col, "TOTALE Δ")
    _rh(ws, 1, eur_col, "€ da recuperare")
    _rh(ws, 1, ore_col, "Ore da recuperare")

    for ri, worker in enumerate(workers, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _rc(ws, ri, 1, worker, fill=fill, align=LEFT_R)
        for ci, y in enumerate(years, 2):
            v = agg[worker].get(y, 0)
            f = fill if v >= 0 else NEG_FILL
            _rc(ws, ri, ci, v if v != 0 else "", fill=f)
        _rf(ws, ri, tot_col, f"=SUM(B{ri}:{last_yr_l}{ri})", bold=True, fill=TOT_FILL)
        _rf(ws, ri, eur_col, f"={tot_col_l}{ri}*{BUONO_EURO}", bold=True, fill=TOT_FILL,
            fmt='0.00 "€"')
        ero_tot  = sum(agg_ero[worker].get(y, 0) for y in years)
        ore_cell = _rc(ws, ri, ore_col, ero_tot * BUONO_ORE, bold=True, fill=TOT_FILL)
        ore_cell.number_format = '0.0 "h"'

    tr = len(workers) + 2
    _rc(ws, tr, 1, "TOTALE", bold=True, fill=SUB_FILL)
    for ci, y in enumerate(years, 2):
        col_l = openpyxl.utils.get_column_letter(ci)
        _rf(ws, tr, ci, f"=SUM({col_l}2:{col_l}{tr-1})", bold=True, fill=SUB_FILL)
    _rf(ws, tr, tot_col, f"=SUM({tot_col_l}2:{tot_col_l}{tr-1})", bold=True, fill=TOT_FILL)
    _rf(ws, tr, eur_col, f"={tot_col_l}{tr}*{BUONO_EURO}",        bold=True, fill=TOT_FILL,
        fmt='0.00 "€"')
    _rf(ws, tr, ore_col, f"=SUM({ore_col_l}2:{ore_col_l}{tr-1})", bold=True, fill=TOT_FILL,
        fmt='0.0 "h"')

    ws.column_dimensions["A"].width = 35
    for ci in range(2, len(years) + 5):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    wb2.save(out_path)
    grand_delta = sum(sum(yd.values()) for yd in agg.values())
    grand_ero   = sum(sum(yd.values()) for yd in agg_ero.values())
    print(f"  Salvato: {out_path}")
    print(f"  TOTALE DELTA  : {grand_delta:,} buoni pasto")
    print(f"  VALORE EURO   : {grand_delta * BUONO_EURO:,.2f} €")
    print(f"  TOTALE EROGATI: {grand_ero:,} buoni pasto")
    return out_path


# ── ZIP finale ────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"File prodotti: {len(output_files)}")

RIEPILOGO_PATH = None
if output_files:
    print("\nCostruzione RIEPILOGO...")
    RIEPILOGO_PATH = build_riepilogo(
        DEFINITIVO,
        os.path.join(DEFINITIVO, "RIEPILOGO_BUONI_PASTO.xlsx")
    )

    print(f"\nCreazione ZIP: {OUT_ZIP}")
    all_xlsx = sorted(
        fp for fp in
        [os.path.join(DEFINITIVO, f) for f in os.listdir(DEFINITIVO) if f.endswith('.xlsx')]
    )
    with zipfile.ZipFile(OUT_ZIP, 'w', zipfile.ZIP_DEFLATED) as zout:
        for fp in all_xlsx:
            zout.write(fp, os.path.basename(fp))
            print(f"  + {os.path.basename(fp)}")
    print(f"ZIP creato: {OUT_ZIP}")
else:
    print("Nessun file da includere nel ZIP.")

print("=" * 60)
print("COMPLETATO.")
