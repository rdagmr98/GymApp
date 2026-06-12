#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analisi Buoni Pasto - Cartellini TIVOLI (ASL RMG)
Formati supportati:
 A) OLD  - "AZIENDA USL RMG - DIPENDENTI Cartellino contratto Sanitario"
 B) NEW  - "Cartellino Orario" (dal 2023 circa)
 C) CART - "STAMPA CARTELLINO" (formato CITTADELLA)

Regola: giorno lavorato >= 6h15m (375 min) -> buono pasto maturato
Delta = maturati - erogati (come da PDF)
Valore euro: 4.13 €/buono
Valore ore:  0.5 h/buono (30 min)
"""

import os, re, sys, shutil
import concurrent.futures
import multiprocessing
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ─── CONFIG ──────────────────────────────────────────────────────────────────
BASE_PATH   = r"C:\Users\Gianmarco\Documents\tivoli"
OUTPUT_DIR  = os.path.join(BASE_PATH, "buonipasto tivoli")
LOG_FILE    = os.path.join(BASE_PATH, "pdf_non_leggibili.log")
THRESHOLD   = 375          # 6h15m in minuti
BUONO_EURO  = 4.13
BUONO_ORE   = 0.5

# Limit workers for test run (None = all)
MAX_WORKERS = None   # None = tutti i lavoratori

# Resume from a specific folder (None = dall'inizio)
START_FROM  = None   # None = dall'inizio

# Timeout per singolo PDF (secondi) - skippa PDF che bloccano il parsing
PDF_TIMEOUT = 45

MONTHS_IT = {
    'gennaio':1,'febbraio':2,'marzo':3,'aprile':4,'maggio':5,'giugno':6,
    'luglio':7,'agosto':8,'settembre':9,'ottobre':10,'novembre':11,'dicembre':12,
}
MONTH_SHORT = {1:'Gen',2:'Feb',3:'Mar',4:'Apr',5:'Mag',6:'Giu',
               7:'Lug',8:'Ago',9:'Set',10:'Ott',11:'Nov',12:'Dic'}

WEEKDAYS = r'(?:Lun|Mar|Mer|Gio|Ven|Sab|Dom|DomD|DomK)'

# ─── UTILITIES ────────────────────────────────────────────────────────────────

def read_worker_deltas_from_xlsx(xlsx_path: str) -> dict:
    """Read per-year total delta from existing individual worker xlsx (for resume mode)."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if 'Riepilogo' not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb['Riepilogo']
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and isinstance(row[0], int):  # year row (not 'TOTALE')
                year = row[0]
                delta = row[13]  # Column N = TOTALE Δ
                if delta is not None and isinstance(delta, (int, float)):
                    result[year] = int(delta)
        wb.close()
        return result
    except Exception as e:
        print(f"  ⚠ Errore lettura xlsx {os.path.basename(xlsx_path)}: {e}")
        return {}

def _log_unreadable(pdf_path: str, error: str):
    """Append an unreadable PDF entry to the log file."""
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        folder = os.path.dirname(pdf_path)
        f.write(f"{pdf_path}\n  Cartella: {folder}\n  Errore: {error}\n\n")

def hhmm_to_min(s: str) -> int:
    """Convert 'HH:MM' or 'HH.MM' or 'H:MM' to minutes."""
    s = s.strip()
    sep = ':' if ':' in s else '.'
    parts = s.split(sep)
    try:
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return 0

def words_by_line(page):
    """Return dict: y_bucket -> list of word dicts (with x0, text)."""
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    lines = defaultdict(list)
    for w in words:
        y = round(w['top'] / 4) * 4
        lines[y].append(w)
    return lines

def find_col_x(lines_dict, col_name_variants):
    """Find x-position of a column header by name."""
    for y in sorted(lines_dict):
        for w in lines_dict[y]:
            if w['text'] in col_name_variants:
                return w['x0']
    return None

def word_at_x(lines_dict, y_bucket, target_x, tol=35):
    """Get text of word at given x-position in a y-bucket."""
    for w in lines_dict.get(y_bucket, []):
        if abs(w['x0'] - target_x) < tol:
            return w['text']
    return None

def word_at_x_nearby(lines_dict, y_bucket, target_x, tol=35, y_spread=20):
    """Search for a word at target_x in y_bucket and adjacent buckets (for split rows)."""
    sorted_buckets = sorted(lines_dict.keys())
    for b in sorted_buckets:
        if y_bucket <= b <= y_bucket + y_spread:
            for w in lines_dict[b]:
                if abs(w['x0'] - target_x) < tol:
                    return w['text'], b
    return None, None

def row_words_nearby(lines_dict, y_bucket, y_spread=20):
    """Collect all words from y_bucket and nearby buckets (for split rows)."""
    result = []
    for b, words in lines_dict.items():
        if y_bucket <= b <= y_bucket + y_spread:
            result.extend(words)
    result.sort(key=lambda w: w['x0'])
    return result

def word_at_x_or_above(lines_dict, y_bucket, target_x, tol=35, max_above=6):
    """Get word at target_x: try current y_bucket first, then just above (split-row fix).
    
    In OLD format PDFs, some rows are split so that the day number/DOW is at y_bucket
    but the timbrature and Lav. value are at y_bucket-4 (slightly above).
    We use max_above=6 to catch splits (gap=4) without reaching the previous row (gap>=8).
    """
    # Try current bucket first
    result = word_at_x(lines_dict, y_bucket, target_x, tol)
    if result is not None:
        return result
    # Search just above (descending order = closest first)
    for b in sorted(lines_dict.keys(), reverse=True):
        if y_bucket - max_above <= b < y_bucket:
            result = word_at_x(lines_dict, b, target_x, tol)
            if result is not None:
                return result
    return None

def has_timbrature_old(lines_dict, y_bucket, max_x=340, max_above=6):
    """True if there are HH.MM time values in the timbrature area (split-row aware)."""
    TIME_RE = re.compile(r'^\d{2}\.\d{2}')
    # Check current bucket
    for w in lines_dict.get(y_bucket, []):
        if w['x0'] < max_x and TIME_RE.match(w['text']):
            return True
    # Check just above (for split rows)
    for b in sorted(lines_dict.keys(), reverse=True):
        if y_bucket - max_above <= b < y_bucket:
            for w in lines_dict.get(b, []):
                if w['x0'] < max_x and TIME_RE.match(w['text']):
                    return True
    return False

def has_timbrature_new(lines_dict, y_bucket, lav_x, tol=35):
    """True if there are HH:MM values in timbrature area (x < lav_x - 30)."""
    TIME_RE = re.compile(r'^\d{1,2}:\d{2}$')
    for w in lines_dict.get(y_bucket, []):
        if w['x0'] < (lav_x - 30) and TIME_RE.match(w['text']):
            if w['text'] not in ('00:00',):
                return True
    return False

def _calc_from_timbrature_old(lines_dict: dict, y_bucket: int, lav_x: float,
                               max_above: int = 6) -> int:
    """Compute worked minutes from entry/exit time pairs in OLD format.
    Used when no Lav. column value is found at the expected x position.
    Times appear as HH.MM in the timbrature area (x < lav_x).
    Pairs are assumed: entry[0], exit[0], entry[1], exit[1], ...
    """
    TIME_RE = re.compile(r'^(\d{2})\.(\d{2})$')
    times_min = []
    # Collect times from y_bucket and just above (split-row aware)
    buckets = sorted(
        [b for b in lines_dict if y_bucket - max_above <= b <= y_bucket]
    )
    for b in buckets:
        for w in lines_dict.get(b, []):
            if w['x0'] < lav_x - 5:
                mt = TIME_RE.match(w['text'])
                if mt:
                    times_min.append(int(mt.group(1)) * 60 + int(mt.group(2)))
    if len(times_min) < 2:
        return 0
    total = 0
    for i in range(0, len(times_min) - 1, 2):
        diff = times_min[i + 1] - times_min[i]
        if 0 < diff < 720:   # sanity: each shift < 12 h
            total += diff
    return total

# ─── FORMAT DETECTION ─────────────────────────────────────────────────────────

def detect_format(text: str) -> str:
    if 'AZIENDA USL RMG' in text and 'Cartellino contratto' in text:
        return 'OLD'
    if 'Cartellino Orario' in text or 'Cartellino orario' in text or 'CartellinoOrario' in text:
        return 'NEW'
    if 'STAMPA CARTELLINO' in text:
        return 'CART'
    # Formato OLD invertito (testo RTL): AZIENDA → ADNEIZA, Cartellino → onilletraC
    if 'ADNEIZA' in text and 'onilletraC' in text:
        return 'OLD_REV'
    return 'UNKNOWN'

# ─── OLD FORMAT PARSER ────────────────────────────────────────────────────────

def parse_old_format(page, text: str):
    """
    Format: AZIENDA USL RMG - Cartellino contratto Sanitario
    Returns dict with worker, year, month, maturati, erogati, details
    """
    # ── Month/year from first line ──
    m_date = re.search(r'mese:\s*(\w+)\s+anno:\s*(\d{4})', text, re.I)
    if not m_date:
        return None
    month_name = m_date.group(1).lower()
    year = int(m_date.group(2))
    month = MONTHS_IT.get(month_name)
    if not month:
        return None

    # ── Worker name from Nominativo column ──
    lines_dict = words_by_line(page)
    # Find "Nominativo" and "Pianificazione" header x positions
    nom_x = find_col_x(lines_dict, ['Nominativo'])
    pia_x = find_col_x(lines_dict, ['Pianificazione'])
    worker = _extract_name_old(lines_dict, nom_x, pia_x)
    if not worker:
        # Fallback: regex on text
        m_name = re.search(r'\d{7,10}\s+([A-ZÀ-ÿ\'\s]+?)\s+[A-Z0-9]{2,8}\s+TURNO', text)
        worker = m_name.group(1).strip() if m_name else 'SCONOSCIUTO'

    # ── Lav. column x position ──
    lav_x = find_col_x(lines_dict, ['Lav.'])
    if lav_x is None:
        lav_x = 359  # fallback

    # ── Buoni erogati ──
    erogati = 0
    erogati_found = False
    m_buoni = re.search(r'Buoni\s+Pasto\s+Salvo\s+Conguaglio\s+(\d+)', text, re.I)
    if m_buoni:
        erogati = int(m_buoni.group(1))
        erogati_found = True

    # ── Day rows ──
    DAY_RE = re.compile(
        r'^\s*(\d{1,2})\s+(' + WEEKDAYS + r')\s*([LFDA]?)\s*(.*)',
        re.IGNORECASE
    )

    maturati = 0
    details  = []

    for y_bucket in sorted(lines_dict):
        row_words = lines_dict[y_bucket]
        row_text  = ' '.join(w['text'] for w in row_words)
        dm = DAY_RE.match(row_text)
        if not dm:
            continue

        day    = dm.group(1)
        dow    = dm.group(2).upper().replace('DOMD', 'DO').replace('DOMK', 'DO')
        flag   = dm.group(3).upper()

        # Get Lav value from column x-position (split-row aware: also checks just above)
        lav_raw = word_at_x_or_above(lines_dict, y_bucket, lav_x)
        if lav_raw is None:
            # No Lav. column value: try computing from entry/exit time pairs
            lav_min = _calc_from_timbrature_old(lines_dict, y_bucket, lav_x)
            worked  = lav_min > 0
        else:
            # Clean suffix letters (CME, AM, HS...)
            lav_clean = re.sub(r'[A-Z]+$', '', lav_raw)
            lav_min = hhmm_to_min(lav_clean) if re.match(r'\d{2}\.\d{2}', lav_clean) else 0
            # If lav_clean is not a valid time, also try from entry/exit pairs
            if lav_min == 0:
                lav_min = _calc_from_timbrature_old(lines_dict, y_bucket, lav_x)
            # Check real work (timbrature present in timbrature area, split-row aware)
            worked = has_timbrature_old(lines_dict, y_bucket, max_x=lav_x - 5)
            if not worked and lav_min > 0:
                worked = True

        qualifies = worked and (lav_min >= THRESHOLD)
        if qualifies:
            maturati += 1

        details.append({
            'day': day, 'dow': dow,
            'worked': worked,
            'lav_min': lav_min,
            'lav_hm': f"{lav_min//60}h{lav_min%60:02d}m" if lav_min else '',
            'qualifies': qualifies,
        })

    return {
        'worker': worker,
        'year': year,
        'month': month,
        'maturati': maturati,
        'erogati': erogati,
        'erogati_found': erogati_found,
        'details': details,
    }

def _extract_name_old(lines_dict, nom_x, pia_x):
    """Extract worker name from Nominativo column x-range."""
    if nom_x is None:
        return None
    x_min = nom_x - 5
    x_max = pia_x - 5 if pia_x else nom_x + 250

    MAT_RE = re.compile(r'^\d{7,10}$')
    for y in sorted(lines_dict):
        row_words = lines_dict[y]
        if any(MAT_RE.match(w['text']) for w in row_words):
            name_words = [
                w['text'] for w in row_words
                if x_min <= w['x0'] <= x_max
            ]
            if name_words:
                return ' '.join(name_words)
    return None

# ─── OLD REVERSED FORMAT PARSER ──────────────────────────────────────────────

def parse_reversed_old_format(page, text: str):
    """
    OLD format PDFs con testo RTL (caratteri invertiti per riflessione).
    x0 in portrait = posizione colonna landscape (giorno); top = riga dati landscape.
    Tutti i testi sono invertiti carattere per carattere.
    """
    words = page.extract_words(x_tolerance=3, y_tolerance=3)

    # '.vaL' = 'Lav.' invertito
    lav_h = next((w for w in words if w['text'] == '.vaL'), None)
    if not lav_h:
        return None
    lav_top = lav_h['top']

    # Mese/anno dalle righe di testo invertite
    rev_lines = [ln[::-1].strip() for ln in text.splitlines() if ln.strip()]
    month_name = None
    year = None
    for ln in rev_lines:
        if re.match(r'^\d{4}$', ln):
            y = int(ln)
            if 2000 <= y <= 2030:
                year = y
        if ln.lower() in MONTHS_IT:
            month_name = ln.lower()
    if not (month_name and year):
        return None
    month = MONTHS_IT.get(month_name)
    if not month:
        return None

    # Nome lavoratore: parole maiuscole vicino all'header 'ovitanimoN' (Nominativo inv.)
    nom_h = next((w for w in words if w['text'] == 'ovitanimoN'), None)
    nom_top = nom_h['top'] if nom_h else 720
    name_cands = sorted(
        [(w['top'], w['text'][::-1]) for w in words
         if 45 <= w['x0'] <= 70
         and (nom_top - 160) <= w['top'] <= (nom_top + 50)
         and w['text'][::-1].isupper()
         and not w['text'][::-1].isdigit()
         and len(w['text']) >= 3],
        key=lambda x: x[0], reverse=True
    )
    worker = ' '.join(r for _, r in name_cands[:2]) if name_cands else 'SCONOSCIUTO'

    # Erogati dal testo ricostruito
    erogati = 0
    erogati_found = False
    reconstructed = ' '.join(rev_lines)
    m_buoni = re.search(r'Buoni\s+Pasto\s+Salvo\s+Conguaglio\s+(\d+)', reconstructed, re.I)
    if m_buoni:
        erogati = int(m_buoni.group(1))
        erogati_found = True

    # Trova day_num_top: top più frequente tra interi 1-31 a x0 > 80
    from collections import Counter as _Counter
    cand_tops = []
    for w in words:
        if w['x0'] <= 80:
            continue
        rev = w['text'][::-1]
        if re.match(r'^\d{1,2}$', rev):
            try:
                d = int(rev)
                if 1 <= d <= 31:
                    cand_tops.append(round(w['top']))
            except Exception:
                pass
    if not cand_tops:
        return None
    day_num_top = _Counter(cand_tops).most_common(1)[0][0]

    # Parole numero-giorno
    day_words = sorted(
        [w for w in words
         if abs(round(w['top']) - day_num_top) <= 6
         and w['x0'] > 80
         and re.match(r'^\d{1,2}$', w['text'][::-1])
         and 1 <= int(w['text'][::-1]) <= 31],
        key=lambda w: w['x0']
    )

    LAV_TOP_TOL = 12
    maturati = 0
    details  = []

    for dw in day_words:
        day_num = dw['text'][::-1]
        day_x0  = dw['x0']

        # DOW: parola multi-char vicino a day_num_top - 20
        dow_cands = [
            w for w in words
            if abs(w['x0'] - day_x0) <= 5
            and (day_num_top - 33) <= w['top'] <= (day_num_top - 8)
            and len(w['text']) > 1
        ]
        dow_str = ''
        if dow_cands:
            dow_cands.sort(key=lambda w: abs(w['top'] - (day_num_top - 20)))
            dow_str = re.sub(r'[^A-Z]', '', dow_cands[0]['text'][::-1].upper())[:3]

        # Valore Lav. a lav_top
        lav_cands = [
            w for w in words
            if abs(w['x0'] - day_x0) <= 5
            and abs(w['top'] - lav_top) <= LAV_TOP_TOL
        ]
        lav_min = 0
        if lav_cands:
            lav_raw   = lav_cands[0]['text'][::-1]
            lav_clean = re.sub(r'[A-Z]+$', '', lav_raw)
            if re.match(r'^\d{1,2}\.\d{2}', lav_clean):
                lav_min = hhmm_to_min(lav_clean)

        worked    = lav_min > 0
        qualifies = worked and (lav_min >= THRESHOLD)
        if qualifies:
            maturati += 1

        details.append({
            'day': day_num, 'dow': dow_str,
            'worked': worked,
            'lav_min': lav_min,
            'lav_hm': f"{lav_min//60}h{lav_min%60:02d}m" if lav_min else '',
            'qualifies': qualifies,
        })

    return {
        'worker': worker,
        'year': year,
        'month': month,
        'maturati': maturati,
        'erogati': erogati,
        'erogati_found': erogati_found,
        'details': details,
    }


# ─── NEW FORMAT PARSER ────────────────────────────────────────────────────────

def parse_new_format(page, text: str):
    """
    Format: Cartellino Orario (new format)
    Line 1 (index 1): 'COGNOME NOME, MESE ANNO, Cartellino Orario'
    Uses x-column position to read 'Lavorato' value so that trailing Rp.DIU/Rp.NOF/Rp.NEF
    columns (present on reperibilità days) are not mistakenly used as Lavorato.
    """
    lines = text.splitlines()
    # Find the worker/date header line
    worker, month, year = None, None, None
    for ln in lines[:5]:
        m = re.match(r'^([A-ZÀ-ÿ\'\s]+),\s*([A-Za-z]+)\s*(\d{4}),\s*Cartellino', ln, re.I)
        if m:
            worker     = m.group(1).strip()
            month_name = m.group(2).strip().lower()
            year       = int(m.group(3))
            month      = MONTHS_IT.get(month_name)
            break

    if not (worker and month and year):
        return None

    # ── Buoni erogati ──
    erogati = 0
    erogati_found = False
    m_buoni = re.search(r'Buoni\s+pasto\s+(\d+)', text, re.I)
    if m_buoni:
        erogati = int(m_buoni.group(1))
        erogati_found = True

    # ── Day rows: use word extraction to find Lavorato by column position ──
    # Direct word clustering with 5px y-tolerance avoids the banker's-rounding split
    # that occurs when Lavorato words are ~0.18px above their DOW/entry-time companions
    # (round(top/4)*4 buckets sometimes put them in different buckets).
    raw_words = page.extract_words(x_tolerance=3, y_tolerance=3)
    lav_x = None
    for w in raw_words:
        if w['text'] == 'Lavorato':
            lav_x = w['x0']
            break

    DAY_NUM_RE = re.compile(r'^\d{1,2}$')
    DOW_RE     = re.compile(r'^(?:lun|mar|mer|gio|ven|sab|dom)$', re.I)
    TIME_RE    = re.compile(r'^\d{1,2}:\d{2}\*?$')

    maturati = 0
    details  = []

    if lav_x is not None:
        # ── Cluster words into visual rows using y-proximity (tol=5px) ──
        # Row spacing in new-format PDFs is ~9px; same-row words differ by ≤0.2px in y.
        row_clusters = []  # list of [word, ...]
        for w in sorted(raw_words, key=lambda x: x['top']):
            placed = False
            for cluster in row_clusters:
                if abs(w['top'] - cluster[0]['top']) <= 5:
                    cluster.append(w)
                    placed = True
                    break
            if not placed:
                row_clusters.append([w])

        # Sort each cluster by x so texts[0]=day_num, texts[1]=DOW
        for cluster in row_clusters:
            cluster.sort(key=lambda w: w['x0'])

        for cluster in row_clusters:
            texts = [w['text'] for w in cluster]

            # Day rows: first word = day number (1-31), second = DOW name
            if len(texts) < 2:
                continue
            if not (DAY_NUM_RE.match(texts[0]) and DOW_RE.match(texts[1])):
                continue

            day = texts[0]
            dow = texts[1].upper()

            # Smontante: first entry time (well before Lavorato column) is 00:00
            entry_words = [w for w in cluster if w['x0'] < lav_x - 50
                           and TIME_RE.match(w['text'])]
            if entry_words and entry_words[0]['text'].replace('*', '') == '00:00':
                details.append({'day': day, 'dow': dow, 'worked': False,
                                'lav_min': 0, 'lav_hm': '', 'qualifies': False})
                continue

            # Get the Lavorato value at its exact column (tol=20 keeps clear of Rp cols)
            lav_words = [w for w in cluster
                         if abs(w['x0'] - lav_x) <= 20
                         and TIME_RE.match(w['text'].replace('*', ''))]
            if lav_words:
                lav_raw = lav_words[0]['text']
                lav_min = hhmm_to_min(lav_raw.replace('*', '').replace(':', '.'))
            else:
                lav_min = 0

            worked    = lav_min > 0
            qualifies = worked and (lav_min >= THRESHOLD)
            if qualifies:
                maturati += 1

            details.append({
                'day': day, 'dow': dow,
                'worked': worked,
                'lav_min': lav_min,
                'lav_hm': f"{lav_min//60}h{lav_min%60:02d}m" if lav_min else '',
                'qualifies': qualifies,
            })

    else:
        # ── Fallback: text-based approach (original) ──
        DAY_RE   = re.compile(r'^\s*(\d{1,2})\s+(lun|mar|mer|gio|ven|sab|dom)\s*(.*)', re.I)
        TIME_RE2 = re.compile(r'\b(\d{1,2}:\d{2})\b')

        for ln in lines:
            dm = DAY_RE.match(ln.strip())
            if not dm:
                continue

            day  = dm.group(1)
            dow  = dm.group(2).upper()
            rest = dm.group(3)
            times = TIME_RE2.findall(rest)

            if not times:
                details.append({'day': day, 'dow': dow, 'worked': False,
                                'lav_min': 0, 'lav_hm': '', 'qualifies': False})
                continue
            if times[0] == '00:00':
                details.append({'day': day, 'dow': dow, 'worked': False,
                                'lav_min': 0, 'lav_hm': '', 'qualifies': False})
                continue
            if len(times) < 2:
                details.append({'day': day, 'dow': dow, 'worked': False,
                                'lav_min': 0, 'lav_hm': '', 'qualifies': False})
                continue

            if len(times) % 2 == 0:
                # Even count: all entry/exit pairs, no explicit Lavorato column.
                # Compute total worked = sum of (exit_i - entry_i)
                total = 0
                for i in range(0, len(times), 2):
                    t_in  = hhmm_to_min(times[i].replace(':', '.'))
                    t_out = hhmm_to_min(times[i + 1].replace(':', '.'))
                    diff  = t_out - t_in
                    if 0 < diff < 720:
                        total += diff
                lav_min = total
            else:
                # Odd count: last value is the explicit Lavorato column value
                lav_min = hhmm_to_min(times[-1].replace(':', '.'))
            worked  = lav_min > 0
            qualifies = worked and (lav_min >= THRESHOLD)
            if qualifies:
                maturati += 1

            details.append({
                'day': day, 'dow': dow,
                'worked': worked,
                'lav_min': lav_min,
                'lav_hm': f"{lav_min//60}h{lav_min%60:02d}m" if lav_min else '',
                'qualifies': qualifies,
            })

    return {
        'worker': worker,
        'year': year,
        'month': month,
        'maturati': maturati,
        'erogati': erogati,
        'erogati_found': erogati_found,
        'details': details,
    }

# ─── STAMPA CARTELLINO FORMAT PARSER ─────────────────────────────────────────

def parse_cart_format(page, text: str):
    """
    Format: STAMPA CARTELLINO (formato Rieti/cittadella)
    """
    # Worker name
    m_cog = re.search(r'Cognome\s+([A-ZÀ-ÿ\'\-]+)', text)
    m_nom = re.search(r'Nome\s+([A-ZÀ-ÿ\'\-]+)', text)
    if not (m_cog and m_nom):
        return None
    worker = f"{m_cog.group(1)} {m_nom.group(1)}"

    # Month/year
    m_date = re.search(r'Mese\s+(\d{2})/(\d{4})', text)
    if not m_date:
        return None
    month = int(m_date.group(1))
    year  = int(m_date.group(2))

    # Lavorate column x position
    lines_dict = words_by_line(page)
    lav_x = find_col_x(lines_dict, ['Lavorate'])
    if lav_x is None:
        lav_x = 490  # fallback

    # Buoni erogati (page 1 only - page 2 handled by caller)
    erogati = 0  # will be updated by caller from page 2

    # Day rows: '01 MA E07:28[01] U13:31[01] ...'
    DAY_RE   = re.compile(r'^\s*(\d{2})\s+([A-Z]{2}\*?)\s+(.*)', re.I)
    EU_TIME  = re.compile(r'[EU]\d{2}:\d{2}')   # E07:28, U13:31
    TIME_RE  = re.compile(r'\d{2}:\d{2}')

    maturati = 0
    details  = []

    for y_bucket in sorted(lines_dict):
        row_words = lines_dict[y_bucket]
        row_text  = ' '.join(w['text'] for w in row_words)
        dm = DAY_RE.match(row_text)
        if not dm:
            continue

        day = dm.group(1)
        dow = dm.group(2).upper().replace('*', '')

        # Check real work: E/U timbrature present
        has_eu = EU_TIME.search(row_text)

        # Lavorato at column x
        lav_raw = word_at_x(lines_dict, y_bucket, lav_x, tol=20)
        if lav_raw and TIME_RE.match(lav_raw) and has_eu:
            lav_min = hhmm_to_min(lav_raw.replace(':', '.'))
            worked  = True
        else:
            lav_min = 0
            worked  = False

        qualifies = worked and (lav_min >= THRESHOLD)
        if qualifies:
            maturati += 1

        details.append({
            'day': day, 'dow': dow,
            'worked': worked,
            'lav_min': lav_min,
            'lav_hm': f"{lav_min//60}h{lav_min%60:02d}m" if lav_min else '',
            'qualifies': qualifies,
        })

    return {
        'worker': worker,
        'year': year,
        'month': month,
        'maturati': maturati,
        'erogati': erogati,
        'erogati_found': False,
        'details': details,
    }

# ─── PDF PROCESSING ───────────────────────────────────────────────────────────

def process_pdf(pdf_path: str) -> list:
    """Process a single PDF file. Returns list of monthly result dicts."""
    results = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            n_pages = len(pdf.pages)
            # For STAMPA CARTELLINO: page 1 = data, page 2 = summary (erogati)
            cart_page2_text = ''
            try:
                if n_pages >= 2:
                    cart_page2_text = pdf.pages[1].extract_text() or ''
            except Exception:
                pass

            i = 0
            while i < n_pages:
                try:
                    page = pdf.pages[i]
                    text = page.extract_text() or ''
                except Exception as pe:
                    _log_unreadable(pdf_path, f"pagina {i+1}: {pe}")
                    i += 1
                    continue
                fmt  = detect_format(text)

                if fmt == 'OLD':
                    r = parse_old_format(page, text)
                    if r:
                        results.append(r)
                    i += 1

                elif fmt == 'NEW':
                    r = parse_new_format(page, text)
                    if r:
                        results.append(r)
                    i += 1

                elif fmt == 'OLD_REV':
                    r = parse_reversed_old_format(page, text)
                    if r:
                        results.append(r)
                    i += 1

                elif fmt == 'CART':
                    # Page 2 (summary) for this cartellino
                    pg2_text = ''
                    if i + 1 < n_pages:
                        pg2_text = pdf.pages[i+1].extract_text() or ''
                    elif cart_page2_text:
                        pg2_text = cart_page2_text

                    r = parse_cart_format(page, text)
                    if r:
                        # Extract erogati from page 2
                        m_pc = re.search(r'Pasti\s+Convenzionati\s+(\d+)', pg2_text, re.I)
                        m_pi = re.search(r'Pasti\s+Interi\s+(\d+)', pg2_text, re.I)
                        if m_pc or m_pi:
                            pc = int(m_pc.group(1)) if m_pc else 0
                            pi = int(m_pi.group(1)) if m_pi else 0
                            r['erogati'] = pc + pi
                            r['erogati_found'] = True
                        results.append(r)
                    # Skip page 2 (summary page)
                    i += 2

                else:
                    i += 1

    except Exception as e:
        print(f"    ERRORE leggendo {pdf_path}: {e}", flush=True)
        _log_unreadable(pdf_path, str(e))

    return results

def find_all_pdfs(base_dir: str) -> list:
    """Find all PDF files recursively."""
    pdfs = []
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith('.pdf'):
                pdfs.append(os.path.join(root, f))
    return pdfs

# ─── EXCEL STYLES ─────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HDR_FONT   = Font(bold=True, color='FFFFFF')
TOT_FILL   = PatternFill('solid', fgColor='BDD7EE')
TOT_FONT   = Font(bold=True)
SUB_FILL   = PatternFill('solid', fgColor='D6E4F0')
ALT_FILL   = PatternFill('solid', fgColor='EEF4FB')
MISS_FILL  = PatternFill('solid', fgColor='FFE699')   # erogati non trovati
NEG_FILL   = PatternFill('solid', fgColor='FCE4D6')   # delta negativo
THIN       = Side(style='thin', color='CCCCCC')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER     = Alignment(horizontal='center', vertical='center')
LEFT       = Alignment(horizontal='left', vertical='center')

def _h(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER
    return cell

def _c(ws, r, c, v, bold=False, fill=None, align=CENTER):
    cell = ws.cell(row=r, column=c, value=v)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    cell.alignment = align
    cell.border = BORDER
    return cell

# ─── WORKER EXCEL ─────────────────────────────────────────────────────────────

def create_worker_excel(worker: str, records: list, out_dir: str):
    """Create per-worker Excel with summary (Riepilogo) sheet + per-year detail sheets.

    Returns (out_path, year_row_map) where year_row_map = {year: row_in_riepilogo}.
    Riepilogo month cells use intra-workbook formula references to TOTALE MESE rows
    in each year's detail sheet.  All totals (Δ, €, ore) use Excel formulas so they
    auto-update if the user edits maturati/erogati in the detail sheets.
    """
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = 'Riepilogo'

    records = sorted(records, key=lambda r: (r['year'], r['month']))
    years   = sorted({r['year'] for r in records})

    ALL_MONTHS   = list(range(1, 13))
    month_labels = [MONTH_SHORT[m] for m in ALL_MONTHS]

    # ── Step 1: write detail sheets and track TOTALE MESE row positions ──
    # totale_mese_rows[year][month] = row number in that year's sheet
    totale_mese_rows = {}

    for y in years:
        year_records = [r for r in records if r['year'] == y]
        if not year_records:
            continue
        ws = wb.create_sheet(title=str(y))
        headers = ['Mese', 'Giorno', 'GG', 'Lavorato', 'Qualifica',
                   'Maturato', 'Erogato', 'Δ', 'Note']
        for ci, h in enumerate(headers, 1):
            _h(ws, 1, ci, h)

        ri2 = 2
        totale_mese_rows[y] = {}
        tm_rows = []   # list of TOTALE MESE row numbers (for TOTALE ANNO SUM)

        for rec in sorted(year_records, key=lambda r: r['month']):
            delta = rec['maturati'] - rec['erogati']
            # Daily rows (worked days only)
            for d in rec['details']:
                if not d['worked']:
                    continue
                fill = ALT_FILL if ri2 % 2 == 0 else None
                _c(ws, ri2, 1, MONTH_SHORT.get(rec['month'], ''), fill=fill)
                _c(ws, ri2, 2, d['day'], fill=fill)
                _c(ws, ri2, 3, d['dow'], fill=fill)
                _c(ws, ri2, 4, d['lav_hm'], fill=fill)
                _c(ws, ri2, 5, 'Sì' if d['qualifies'] else '', fill=fill)
                _c(ws, ri2, 6, '', fill=fill)
                _c(ws, ri2, 7, '', fill=fill)
                _c(ws, ri2, 8, '', fill=fill)
                _c(ws, ri2, 9, '', fill=fill)
                ri2 += 1

            # TOTALE MESE row
            note = '' if rec['erogati_found'] else '⚠ erogati non in PDF'
            f = (MISS_FILL if not rec['erogati_found'] and rec['maturati'] > 0
                 else SUB_FILL)
            if delta < 0:
                f = NEG_FILL
            mn = MONTH_SHORT.get(rec['month'], '')
            _c(ws, ri2, 1, mn, bold=True, fill=f)
            _c(ws, ri2, 2, 'TOTALE MESE', bold=True, fill=f)
            for ci in range(3, 6):
                _c(ws, ri2, ci, '', fill=f)
            _c(ws, ri2, 6, rec['maturati'], bold=True, fill=f)   # F = Maturato
            _c(ws, ri2, 7, rec['erogati'],  bold=True, fill=f)   # G = Erogato
            # H = Δ as formula, clamped to 0 (can't be negative)
            cell_h = ws.cell(row=ri2, column=8, value=f'=MAX(0,F{ri2}-G{ri2})')
            cell_h.font = TOT_FONT; cell_h.fill = f
            cell_h.alignment = CENTER; cell_h.border = BORDER
            _c(ws, ri2, 9, note, fill=f, align=LEFT)

            totale_mese_rows[y][rec['month']] = ri2
            tm_rows.append(ri2)
            ri2 += 1

        # TOTALE ANNO row
        # H = SUM of all TOTALE MESE rows (daily rows have '' in H → ignored by SUM)
        h_sum = f'=SUM(H2:H{ri2 - 1})'
        mat_sum = '+'.join(f'F{r}' for r in tm_rows) if tm_rows else '0'
        ero_sum = '+'.join(f'G{r}' for r in tm_rows) if tm_rows else '0'

        _c(ws, ri2, 1, 'TOTALE ANNO', bold=True, fill=TOT_FILL)
        for ci in range(2, 6):
            _c(ws, ri2, ci, '', fill=TOT_FILL)

        for ci, formula in [(6, f'={mat_sum}'), (7, f'={ero_sum}'), (8, h_sum)]:
            cell = ws.cell(row=ri2, column=ci, value=formula)
            cell.font = TOT_FONT; cell.fill = TOT_FILL
            cell.alignment = CENTER; cell.border = BORDER

        # Note col: euro (from delta H) and ore (from erogati G) as TEXT formula (informational)
        note_formula = (f'=TEXT(H{ri2}*{BUONO_EURO},"0.00")&" €  /  "'
                        f'&TEXT(G{ri2}*{BUONO_ORE},"0.0")&" h"')
        cell_n = ws.cell(row=ri2, column=9, value=note_formula)
        cell_n.font = TOT_FONT; cell_n.fill = TOT_FILL
        cell_n.alignment = LEFT; cell_n.border = BORDER
        ri2 += 1

        for ci, w in enumerate([8, 7, 5, 10, 8, 8, 8, 5, 30], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # ── Step 2: write Riepilogo sheet with formula references to detail sheets ──
    _h(ws_sum, 1, 1, 'Anno')
    for ci, ml in enumerate(month_labels, 2):
        _h(ws_sum, 1, ci, ml)
    _h(ws_sum, 1, 14, 'TOTALE Δ')
    _h(ws_sum, 1, 15, '€ da recuperare')
    _h(ws_sum, 1, 16, 'Ore da recuperare')

    year_row_map = {}   # year -> row in this Riepilogo (for external references)

    for ri, y in enumerate(years, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _c(ws_sum, ri, 1, y, bold=True, fill=fill)
        year_row_map[y] = ri

        for ci, m in enumerate(ALL_MONTHS, 2):
            tm_row = totale_mese_rows.get(y, {}).get(m)
            if tm_row is not None:
                # Reference TOTALE MESE Δ in the year detail sheet
                # Sheet name = str(y); always a 4-digit number → no special quoting issues
                formula = f"='{y}'!$H${tm_row}"
                cell = ws_sum.cell(row=ri, column=ci, value=formula)
                cell.fill = fill if fill else PatternFill()
                cell.alignment = CENTER
                cell.border = BORDER
            else:
                _c(ws_sum, ri, ci, '', fill=fill)

        # TOTALE Δ = SUM of all month columns (formula)
        cell_n = ws_sum.cell(row=ri, column=14, value=f'=SUM(B{ri}:M{ri})')
        cell_n.font = TOT_FONT; cell_n.fill = SUB_FILL
        cell_n.alignment = CENTER; cell_n.border = BORDER

        # € and Ore as formulas
        cell_e = ws_sum.cell(row=ri, column=15, value=f'=N{ri}*{BUONO_EURO}')
        cell_e.font = TOT_FONT; cell_e.fill = TOT_FILL
        cell_e.alignment = CENTER; cell_e.border = BORDER
        cell_e.number_format = '0.00 "€"'

        # Ore: erogati × 0.5 (non delta × 0.5)
        _ero_rows = sorted(totale_mese_rows.get(y, {}).values())
        if _ero_rows:
            _ero_sum = '+'.join(f"'{y}'!G{r}" for r in _ero_rows)
            _ore_formula = f'=({_ero_sum})*{BUONO_ORE}'
        else:
            _ore_formula = '=0'
        cell_o = ws_sum.cell(row=ri, column=16, value=_ore_formula)
        cell_o.font = TOT_FONT; cell_o.fill = TOT_FILL
        cell_o.alignment = CENTER; cell_o.border = BORDER
        cell_o.number_format = '0.0 "h"'

    # Grand total row
    tr = len(years) + 2
    _c(ws_sum, tr, 1, 'TOTALE', bold=True, fill=TOT_FILL)
    for ci in range(2, 14):
        _c(ws_sum, tr, ci, '', fill=TOT_FILL)

    cell_nt = ws_sum.cell(row=tr, column=14, value=f'=SUM(N2:N{tr - 1})')
    cell_nt.font = TOT_FONT; cell_nt.fill = TOT_FILL
    cell_nt.alignment = CENTER; cell_nt.border = BORDER

    cell_et = ws_sum.cell(row=tr, column=15, value=f'=N{tr}*{BUONO_EURO}')
    cell_et.font = TOT_FONT; cell_et.fill = TOT_FILL
    cell_et.alignment = CENTER; cell_et.border = BORDER
    cell_et.number_format = '0.00 "€"'

    cell_ot = ws_sum.cell(row=tr, column=16, value=f'=SUM(P2:P{tr - 1})')
    cell_ot.font = TOT_FONT; cell_ot.fill = TOT_FILL
    cell_ot.alignment = CENTER; cell_ot.border = BORDER
    cell_ot.number_format = '0.0 "h"'

    ws_sum.column_dimensions['A'].width = 8
    for ci in range(2, 14):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 7
    ws_sum.column_dimensions['N'].width = 10
    ws_sum.column_dimensions['O'].width = 16
    ws_sum.column_dimensions['P'].width = 16

    # Save
    safe = re.sub(r'[\\/*?:\[\]]', '_', worker)
    out_path = os.path.join(out_dir, f"{safe}.xlsx")
    try:
        wb.save(out_path)
    except PermissionError:
        _log_unreadable(out_path,
            "PermissionError: file aperto in un'altra applicazione - chiuderlo e rieseguire lo script")
        print(f"    ⚠ File in uso, impossibile salvare {safe}.xlsx - annotato nel log")
    return out_path, year_row_map

# ─── CUMULATIVE EXCEL ─────────────────────────────────────────────────────────

def create_cumulative_excel(all_results: dict, out_dir: str,
                            pre_agg: dict = None,
                            worker_year_rows: dict = None):
    """
    all_results:      worker -> list of result dicts (newly processed)
    pre_agg:          worker -> {year -> delta} (pre-loaded from existing xlsx)
    worker_year_rows: worker -> {year -> row_in_worker_riepilogo}

    Per-year/worker cells use external formula references to the individual worker
    xlsx files, so editing a worker detail file propagates here when links are updated.
    All totals use Excel SUM formulas; € and ore use multiplication formulas.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Riepilogo Generale'

    if worker_year_rows is None:
        worker_year_rows = {}

    # Aggregate live results to know which workers/years exist
    agg = defaultdict(lambda: defaultdict(int))   # worker -> year -> delta
    for worker, recs in all_results.items():
        for r in recs:
            agg[worker][r['year']] += (r['maturati'] - r['erogati'])

    # Merge pre-loaded data (from skipped folders in resume mode)
    if pre_agg:
        for worker, year_deltas in pre_agg.items():
            if worker not in agg:
                for year, delta in year_deltas.items():
                    agg[worker][year] += delta
                # Compute year_row_map for pre_agg workers (same sorting logic as create_worker_excel)
                if worker not in worker_year_rows:
                    w_years = sorted(year_deltas.keys())
                    worker_year_rows[worker] = {y: i + 2 for i, y in enumerate(w_years)}

    workers = sorted(agg.keys())
    years   = sorted({y for w in agg.values() for y in w})

    _h(ws, 1, 1, 'Lavoratore')
    for ci, y in enumerate(years, 2):
        _h(ws, 1, ci, y)
    tot_col = len(years) + 2
    eur_col = len(years) + 3
    ore_col = len(years) + 4
    tot_col_l = openpyxl.utils.get_column_letter(tot_col)
    eur_col_l = openpyxl.utils.get_column_letter(eur_col)
    ore_col_l = openpyxl.utils.get_column_letter(ore_col)
    last_yr_l = openpyxl.utils.get_column_letter(len(years) + 1)  # last year column
    _h(ws, 1, tot_col, 'TOTALE Δ')
    _h(ws, 1, eur_col, '€ da recuperare')
    _h(ws, 1, ore_col, 'Ore da recuperare')

    for ri, worker in enumerate(workers, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _c(ws, ri, 1, worker, fill=fill, align=LEFT)

        safe_name = re.sub(r'[\\/*?:\[\]]', '_', worker)
        yr_rows = worker_year_rows.get(worker, {})

        for ci, y in enumerate(years, 2):
            yr_row = yr_rows.get(y)
            if yr_row is not None:
                # External reference: points to TOTALE Δ column (N=14) in worker's Riepilogo
                formula = f"='[{safe_name}.xlsx]Riepilogo'!$N${yr_row}"
                cell = ws.cell(row=ri, column=ci, value=formula)
                cell.fill = fill if fill else PatternFill()
                cell.alignment = CENTER
                cell.border = BORDER
            else:
                _c(ws, ri, ci, '', fill=fill)

        # TOTALE Δ: SUM of all year columns (formula — picks up external refs)
        cell_t = ws.cell(row=ri, column=tot_col, value=f'=SUM(B{ri}:{last_yr_l}{ri})')
        cell_t.font = TOT_FONT; cell_t.fill = TOT_FILL
        cell_t.alignment = CENTER; cell_t.border = BORDER

        # € and ore as formulas
        cell_e = ws.cell(row=ri, column=eur_col, value=f'={tot_col_l}{ri}*{BUONO_EURO}')
        cell_e.font = TOT_FONT; cell_e.fill = TOT_FILL
        cell_e.alignment = CENTER; cell_e.border = BORDER
        cell_e.number_format = '0.00 "€"'

        cell_o = ws.cell(row=ri, column=ore_col, value=f'={tot_col_l}{ri}*{BUONO_ORE}')
        cell_o.font = TOT_FONT; cell_o.fill = TOT_FILL
        cell_o.alignment = CENTER; cell_o.border = BORDER
        cell_o.number_format = '0.0 "h"'

    # Totals row (SUM of each column)
    tr = len(workers) + 2
    _c(ws, tr, 1, 'TOTALE', bold=True, fill=SUB_FILL)
    for ci, y in enumerate(years, 2):
        col_l = openpyxl.utils.get_column_letter(ci)
        cell = ws.cell(row=tr, column=ci, value=f'=SUM({col_l}2:{col_l}{tr - 1})')
        cell.font = TOT_FONT; cell.fill = SUB_FILL
        cell.alignment = CENTER; cell.border = BORDER

    cell_gt = ws.cell(row=tr, column=tot_col, value=f'=SUM({tot_col_l}2:{tot_col_l}{tr - 1})')
    cell_gt.font = TOT_FONT; cell_gt.fill = TOT_FILL
    cell_gt.alignment = CENTER; cell_gt.border = BORDER

    cell_ge = ws.cell(row=tr, column=eur_col, value=f'={tot_col_l}{tr}*{BUONO_EURO}')
    cell_ge.font = TOT_FONT; cell_ge.fill = TOT_FILL
    cell_ge.alignment = CENTER; cell_ge.border = BORDER
    cell_ge.number_format = '0.00 "€"'

    cell_go = ws.cell(row=tr, column=ore_col, value=f'={tot_col_l}{tr}*{BUONO_ORE}')
    cell_go.font = TOT_FONT; cell_go.fill = TOT_FILL
    cell_go.alignment = CENTER; cell_go.border = BORDER
    cell_go.number_format = '0.0 "h"'

    ws.column_dimensions['A'].width = 30
    for ci in range(2, len(years) + 5):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    out_path = os.path.join(out_dir, 'RIEPILOGO_BUONI_PASTO_TIVOLI.xlsx')
    try:
        wb.save(out_path)
    except PermissionError:
        _log_unreadable(out_path, "PermissionError: file riepilogo aperto - chiuderlo e rieseguire")
        print("  ⚠ Riepilogo in uso - impossibile salvare, annotato nel log")
    return out_path

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print('=' * 65)
    print('Analisi Buoni Pasto TIVOLI')
    print('=' * 65)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Reset log file
    with open(LOG_FILE, 'w', encoding='utf-8') as f:
        f.write('PDF NON LEGGIBILI - Analisi Buoni Pasto TIVOLI\n')
        f.write('=' * 60 + '\n\n')

    # Get worker folders
    worker_folders = sorted([
        d for d in os.listdir(BASE_PATH)
        if os.path.isdir(os.path.join(BASE_PATH, d))
        and d not in ('buonipasto tivoli', '_estratti')
    ])

    if MAX_WORKERS:
        worker_folders = worker_folders[:MAX_WORKERS]
        print(f"\n⚠  MODALITÀ TEST: elaboro solo i primi {MAX_WORKERS} lavoratori\n")

    all_results = {}
    worker_year_rows = {}   # worker -> {year -> row_in_riepilogo}

    # ── Resume mode: pre-load deltas from existing xlsx for skipped folders ──
    pre_agg = {}   # worker_name -> {year -> delta}
    if START_FROM:
        print(f"⚠  RIPRESA DA: {START_FROM}")
        print(f"   Carico risultati precedenti da xlsx esistenti...\n")
        for f in sorted(os.listdir(OUTPUT_DIR)):
            if (f.endswith('.xlsx') and
                    f != 'RIEPILOGO_BUONI_PASTO_TIVOLI.xlsx'):
                worker_name = f[:-5]
                deltas = read_worker_deltas_from_xlsx(
                    os.path.join(OUTPUT_DIR, f))
                if deltas:
                    pre_agg[worker_name] = deltas
        print(f"   Lavoratori precedenti caricati: {len(pre_agg)}\n")

    for wf in worker_folders:
        # Skip folders before START_FROM
        if START_FROM and wf < START_FROM:
            continue

        wf_path = os.path.join(BASE_PATH, wf)
        pdfs    = find_all_pdfs(wf_path)
        print(f"\n[{wf}]  ({len(pdfs)} PDF)")

        worker_records = []
        for pdf_path in pdfs:
            # Use thread with timeout to avoid hanging on corrupt PDFs
            executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
            fut = executor.submit(process_pdf, pdf_path)
            try:
                recs = fut.result(timeout=PDF_TIMEOUT)
            except concurrent.futures.TimeoutError:
                _log_unreadable(pdf_path,
                    f"TIMEOUT dopo {PDF_TIMEOUT}s - PDF probabilmente corrotto/bloccante")
                print(f"    ⚠ TIMEOUT ({PDF_TIMEOUT}s): {os.path.basename(pdf_path)}")
                recs = []
            except Exception as e:
                _log_unreadable(pdf_path, str(e))
                recs = []
            finally:
                executor.shutdown(wait=False)
            worker_records.extend(recs)

        if not worker_records:
            print('  Nessun dato estratto.')
            continue

        # Group by actual worker name from PDF (may differ from folder)
        by_worker = defaultdict(list)
        for r in worker_records:
            by_worker[r['worker']].append(r)

        for worker_name, recs in by_worker.items():
            # Skip SCONOSCIUTO entries (unidentified pages with no data)
            if worker_name == 'SCONOSCIUTO':
                total_mat = sum(r['maturati'] for r in recs)
                total_ero = sum(r['erogati'] for r in recs)
                if total_mat == 0 and total_ero == 0:
                    print(f"  [SCONOSCIUTO] {len(recs)} pagine non identificate - ignorate")
                    continue
                else:
                    print(f"  ⚠ SCONOSCIUTO con dati: maturati={total_mat} erogati={total_ero} - incluso nel log")
                    _log_unreadable('SCONOSCIUTO', f"Pagine non identificate con dati: mat={total_mat} ero={total_ero}")
                    continue
            total_mat = sum(r['maturati'] for r in recs)
            total_ero = sum(r['erogati'] for r in recs)
            total_dlt = total_mat - total_ero
            print(f"  {worker_name}: maturati={total_mat} erogati={total_ero} "
                  f"delta={total_dlt} ({total_dlt*BUONO_EURO:.2f}€ / {total_dlt*BUONO_ORE:.1f}h)")

            out_path, yr_row_map = create_worker_excel(worker_name, recs, OUTPUT_DIR)
            print(f"  -> {os.path.basename(out_path)}")

            if worker_name not in all_results:
                all_results[worker_name] = []
            all_results[worker_name].extend(recs)
            worker_year_rows[worker_name] = yr_row_map

    # Cumulative Excel — combine live results + pre-loaded data from xlsx
    print('\n' + '='*65)
    print('Generazione riepilogo generale...')
    cum_path = create_cumulative_excel(all_results, OUTPUT_DIR,
                                       pre_agg=pre_agg,
                                       worker_year_rows=worker_year_rows)
    print(f'Salvato: {cum_path}')

    grand_live = sum(
        sum(r['maturati'] - r['erogati'] for r in recs)
        for recs in all_results.values()
    )
    grand_pre = sum(
        sum(d for d in year_deltas.values())
        for w, year_deltas in pre_agg.items()
        if w not in all_results
    )
    grand = grand_live + grand_pre
    print(f'\nTOTALE DELTA: {grand} buoni pasto')
    print(f'VALORE:       {grand * BUONO_EURO:.2f} €')
    print(f'ORE:          {grand * BUONO_ORE:.1f} h')
    print('='*65)

    # Cleanup temp debug files
    for f in ['debug_cols.py', 'debug_cols2.py', 'debug_samples.py']:
        try: os.remove(os.path.join(BASE_PATH, f))
        except: pass

if __name__ == '__main__':
    main()
