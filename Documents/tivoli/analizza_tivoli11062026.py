#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Processa TIVOLI 11062026: estrazione ricorsiva ZIP, lettura PDF,
Excel per lavoratore, DEFINITIVO, ZIP mail.
"""
import os, sys, re, shutil, zipfile, io
from collections import defaultdict
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\Gianmarco\Documents\tivoli")
import analisi_tivoli, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DOWNLOADS  = r"C:\Users\Gianmarco\Downloads"
ZIP_DIR    = os.path.join(DOWNLOADS, "TIVOLI 11062026")
OUT_ROOT   = os.path.join(DOWNLOADS, "tivoli11062026")
SOURCE_DIR = os.path.join(OUT_ROOT, "buonipasto")
DEFINITIVO = os.path.join(OUT_ROOT, "buonipasto DEFINITIVO")
STAGING    = os.path.join(OUT_ROOT, "_staging")
OUT_ZIP    = os.path.join(DOWNLOADS, "tivoli11062026_DEFINITIVO.zip")

YEAR_SHEETS = {str(y) for y in range(2010, 2030)}
BUONO_ORE   = 0.5
BUONO_EURO  = 4.13

for d in [SOURCE_DIR, DEFINITIVO, STAGING]:
    os.makedirs(d, exist_ok=True)

analisi_tivoli.OUTPUT_DIR = SOURCE_DIR
analisi_tivoli.LOG_FILE   = os.path.join(OUT_ROOT, "pdf_non_leggibili.log")


# ── FASE 1: Estrai ricorsivamente tutti i ZIP ────────────────────────────────

def extract_recursive(zip_path, dest_dir):
    """Estrae ZIP e qualsiasi ZIP annidato trovato dentro."""
    try:
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(dest_dir)
    except Exception as e:
        print(f"  ERRORE estrazione {os.path.basename(zip_path)}: {e}")
        return
    # Cerca inner ZIP e li estrae nella stessa cartella
    for root, dirs, files in os.walk(dest_dir):
        for fname in sorted(files):
            if fname.lower().endswith(".zip") and not fname.startswith("__MACOSX"):
                inner_path = os.path.join(root, fname)
                inner_dest = os.path.join(root, os.path.splitext(fname)[0])
                os.makedirs(inner_dest, exist_ok=True)
                try:
                    with zipfile.ZipFile(inner_path, "r") as iz:
                        iz.extractall(inner_dest)
                    os.remove(inner_path)   # rimuovi il .zip dopo l'estrazione
                except Exception as e:
                    print(f"    ERRORE inner zip {fname}: {e}")


print("=" * 70)
print("FASE 1: Estrazione ZIP (ricorsiva)")
print("=" * 70)

outer_zips = sorted(f for f in os.listdir(ZIP_DIR) if f.endswith(".zip"))
print(f"ZIP trovati: {len(outer_zips)}")
for zname in outer_zips:
    zpath = os.path.join(ZIP_DIR, zname)
    print(f"  Estraggo: {zname}")
    extract_recursive(zpath, STAGING)

print(f"Staging: {STAGING}")


# ── FASE 2: Trova tutti i PDF ─────────────────────────────────────────────────

print("\n" + "=" * 70)
print("FASE 2: Ricerca PDF")
print("=" * 70)

all_pdfs = analisi_tivoli.find_all_pdfs(STAGING)
print(f"PDF trovati: {len(all_pdfs)}")


# ── FASE 3: Processa PDF ──────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("FASE 3: Lettura PDF")
print("=" * 70)

worker_records = []
errori = 0
for i, pdf_path in enumerate(sorted(all_pdfs), 1):
    if i % 50 == 0:
        print(f"  {i}/{len(all_pdfs)}...")
    try:
        recs = analisi_tivoli.process_pdf(pdf_path)
        worker_records.extend(recs)
    except Exception as e:
        print(f"  ERRORE {os.path.basename(pdf_path)}: {e}")
        errori += 1

by_worker = defaultdict(list)
for r in worker_records:
    by_worker[r["worker"]].append(r)

print(f"\nLavoratori trovati: {len(by_worker)}")
for w, recs in sorted(by_worker.items()):
    mat = sum(r["maturati"] for r in recs)
    ero = sum(r["erogati"]  for r in recs)
    print(f"  {w}: mat={mat} ero={ero} delta={mat-ero} ({(mat-ero)*BUONO_EURO:.2f}€)")
if errori:
    print(f"PDF con errori di lettura: {errori}")


# ── FASE 4: Crea Excel singoli (SOURCE) ──────────────────────────────────────

print("\n" + "=" * 70)
print("FASE 4: Creazione Excel singoli")
print("=" * 70)

PAT_H = re.compile(r"H(\d+)\*0\.5")

def fix_worker(src, dst):
    """Copia src → dst correggendo le formule del foglio Riepilogo."""
    wb = openpyxl.load_workbook(src)
    ytm = {}
    for sn in wb.sheetnames:
        if sn not in YEAR_SHEETS:
            continue
        ws2 = wb[sn]
        tm = []
        for ri in range(2, ws2.max_row + 1):
            ca = ws2.cell(ri, 1).value
            cb = ws2.cell(ri, 2).value
            if cb == "TOTALE MESE":
                tm.append(ri)
            elif ca == "TOTALE ANNO":
                ci = ws2.cell(ri, 9)
                if ci.value and isinstance(ci.value, str):
                    ci.value = PAT_H.sub(r"G\1*0.5", ci.value)
        ytm[sn] = tm
    if "Riepilogo" in wb.sheetnames:
        ws3 = wb["Riepilogo"]
        for ri in range(2, ws3.max_row + 1):
            yv = ws3.cell(ri, 1).value
            if isinstance(yv, int):
                trs = ytm.get(str(yv), [])
                parts = "+".join(f"'{yv}'!G{r}" for r in sorted(trs))
                ws3.cell(ri, 16).value = f"=({parts})*{BUONO_ORE}" if trs else "=0"
            elif yv == "TOTALE":
                ws3.cell(ri, 16).value = f"=SUM(P2:P{ri - 1})"
    wb.save(dst)
    wb.close()


source_paths = {}
for worker_name, recs in sorted(by_worker.items()):
    if worker_name == "SCONOSCIUTO":
        print(f"  Ignorate {len(recs)} pagine SCONOSCIUTO")
        continue
    mat = sum(r["maturati"] for r in recs)
    ero = sum(r["erogati"]  for r in recs)
    dlt = mat - ero
    print(f"  {worker_name}: mat={mat} ero={ero} delta={dlt} ({dlt*BUONO_EURO:.2f}€)")
    try:
        out_path, _ = analisi_tivoli.create_worker_excel(worker_name, recs, SOURCE_DIR)
        source_paths[worker_name] = out_path
    except Exception as e:
        print(f"    ERRORE Excel: {e}")


# ── FASE 5: Fix formule → DEFINITIVO ─────────────────────────────────────────

print("\n" + "=" * 70)
print("FASE 5: Fix formule → DEFINITIVO")
print("=" * 70)

for worker_name, src_path in sorted(source_paths.items()):
    safe    = re.sub(r"[\\/*?:\[\]]", "_", worker_name)
    dst_def = os.path.join(DEFINITIVO, f"{safe}.xlsx")
    try:
        fix_worker(src_path, dst_def)
        print(f"  OK: {safe}.xlsx")
    except Exception as e:
        print(f"  ERRORE fix {safe}: {e}")


# ── FASE 6: RIEPILOGO GENERALE ───────────────────────────────────────────────

print("\n" + "=" * 70)
print("FASE 6: Costruzione RIEPILOGO")
print("=" * 70)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
TOT_FILL = PatternFill("solid", fgColor="BDD7EE")
TOT_FONT = Font(bold=True)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
NEG_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN     = Side(style="thin", color="CCCCCC")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center")


def _h(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER


def _c(ws, r, c, v, bold=False, fill=None, align=CENTER):
    cell = ws.cell(row=r, column=c, value=v)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    cell.alignment = align
    cell.border = BORDER
    return cell


def _f(ws, r, c, formula, bold=False, fill=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=formula)
    if bold: cell.font = TOT_FONT
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    cell.alignment = CENTER
    cell.border = BORDER
    return cell


def build_riepilogo(src_dir, out_path):
    agg     = defaultdict(lambda: defaultdict(int))
    agg_ero = defaultdict(lambda: defaultdict(int))
    xlsx_files = sorted(
        f for f in os.listdir(src_dir)
        if f.endswith(".xlsx") and not f.startswith("RIEPILOGO")
    )
    print(f"  Lettura {len(xlsx_files)} file da {os.path.basename(src_dir)}...")
    for fname in xlsx_files:
        worker = fname[:-5]
        path   = os.path.join(src_dir, fname)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name == "Riepilogo":
                    continue
                try:
                    year = int(sheet_name)
                except ValueError:
                    continue
                ws_y = wb[sheet_name]
                for row in ws_y.iter_rows(min_row=2, values_only=True):
                    if row[1] == "TOTALE MESE":
                        mat = row[5] if isinstance(row[5], (int, float)) else 0
                        ero = row[6] if isinstance(row[6], (int, float)) else 0
                        agg[worker][year]     += max(0, int(mat) - int(ero))
                        agg_ero[worker][year] += int(ero)
            wb.close()
        except Exception as e:
            print(f"    ERRORE {fname}: {e}")

    workers = sorted(agg.keys())
    years   = sorted({y for w in agg.values() for y in w})
    if not workers:
        print("  Nessun dato per il riepilogo.")
        return
    print(f"  Lavoratori: {len(workers)},  Anni: {years[0]}–{years[-1]}")

    wb2 = openpyxl.Workbook()
    ws  = wb2.active
    ws.title = "Riepilogo Generale"
    _h(ws, 1, 1, "Lavoratore")
    for ci, y in enumerate(years, 2):
        _h(ws, 1, ci, y)
    tot_col   = len(years) + 2
    eur_col   = len(years) + 3
    ore_col   = len(years) + 4
    tot_col_l = openpyxl.utils.get_column_letter(tot_col)
    last_yr_l = openpyxl.utils.get_column_letter(len(years) + 1)
    ore_col_l = openpyxl.utils.get_column_letter(ore_col)
    _h(ws, 1, tot_col, "TOTALE Δ")
    _h(ws, 1, eur_col, "€ da recuperare")
    _h(ws, 1, ore_col, "Ore da recuperare")

    for ri, worker in enumerate(workers, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _c(ws, ri, 1, worker, fill=fill, align=LEFT)
        for ci, y in enumerate(years, 2):
            v = agg[worker].get(y, 0)
            f = fill if v >= 0 else NEG_FILL
            _c(ws, ri, ci, v if v != 0 else "", fill=f)
        _f(ws, ri, tot_col, f"=SUM(B{ri}:{last_yr_l}{ri})", bold=True, fill=TOT_FILL)
        _f(ws, ri, eur_col, f"={tot_col_l}{ri}*{BUONO_EURO}", bold=True, fill=TOT_FILL,
           fmt='0.00 "€"')
        ero_tot  = sum(agg_ero[worker].get(y, 0) for y in years)
        ore_cell = _c(ws, ri, ore_col, ero_tot * BUONO_ORE, bold=True, fill=TOT_FILL)
        ore_cell.number_format = '0.0 "h"'

    tr = len(workers) + 2
    _c(ws, tr, 1, "TOTALE", bold=True, fill=SUB_FILL)
    for ci, y in enumerate(years, 2):
        col_l = openpyxl.utils.get_column_letter(ci)
        _f(ws, tr, ci, f"=SUM({col_l}2:{col_l}{tr-1})", bold=True, fill=SUB_FILL)
    _f(ws, tr, tot_col, f"=SUM({tot_col_l}2:{tot_col_l}{tr-1})", bold=True, fill=TOT_FILL)
    _f(ws, tr, eur_col, f"={tot_col_l}{tr}*{BUONO_EURO}",        bold=True, fill=TOT_FILL,
       fmt='0.00 "€"')
    _f(ws, tr, ore_col, f"=SUM({ore_col_l}2:{ore_col_l}{tr-1})", bold=True, fill=TOT_FILL,
       fmt='0.0 "h"')

    ws.column_dimensions["A"].width = 35
    for ci in range(2, len(years) + 5):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    wb2.save(out_path)
    print(f"  Salvato: {out_path}")

    grand_delta = sum(sum(yd.values()) for yd in agg.values())
    grand_ero   = sum(sum(yd.values()) for yd in agg_ero.values())
    print(f"  TOTALE DELTA  : {grand_delta:,} buoni pasto")
    print(f"  VALORE EURO   : {grand_delta * BUONO_EURO:,.2f} €")
    print(f"  TOTALE EROGATI: {grand_ero:,} buoni pasto")


build_riepilogo(SOURCE_DIR,  os.path.join(SOURCE_DIR, "RIEPILOGO_BUONI_PASTO.xlsx"))
build_riepilogo(DEFINITIVO,  os.path.join(DEFINITIVO, "RIEPILOGO_BUONI_PASTO.xlsx"))


# ── FASE 7: Impacchetta DEFINITIVO in ZIP per mail ────────────────────────────

print("\n" + "=" * 70)
print("FASE 7: Creazione ZIP per mail")
print("=" * 70)

xlsx_files = sorted(f for f in os.listdir(DEFINITIVO) if f.endswith(".xlsx"))
with zipfile.ZipFile(OUT_ZIP, "w", zipfile.ZIP_DEFLATED) as zout:
    for fname in xlsx_files:
        fpath = os.path.join(DEFINITIVO, fname)
        zout.write(fpath, fname)
        print(f"  Aggiunto: {fname}")

print(f"\nZIP creato: {OUT_ZIP}  ({len(xlsx_files)} file)")

print("\n" + "=" * 70)
print("COMPLETATO.")
print(f"Risultati in: {OUT_ROOT}")
print(f"ZIP mail:     {OUT_ZIP}")
print("=" * 70)
