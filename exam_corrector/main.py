#!/usr/bin/env python3
"""
ExamCorrector v2.0 — Applicazione Docente
Correzione automatica esami A/B/C, report PDF/Excel, statistiche.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog
import json
import csv
import os
import re
import io
import math
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Optional, Dict
from datetime import datetime

# ── Optional deps with feature flags ─────────────────────────────────────────
try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    import fitz            # PyMuPDF
    from PIL import Image
    HAS_FITZ = HAS_NUMPY
except ImportError:
    HAS_FITZ = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    HAS_PLOT = HAS_NUMPY
except ImportError:
    HAS_PLOT = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    import windnd
    HAS_DND = True
except ImportError:
    HAS_DND = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
        HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# ── Color palette ─────────────────────────────────────────────────────────────
C_NAVY   = "#1F4E79"
C_BLUE   = "#2E75B6"
C_BG     = "#f0f4f8"
C_WHITE  = "#ffffff"
C_GREEN  = "#C6EFCE"
C_RED    = "#FFC7CE"
C_AMBER  = "#FFEB9C"
C_BORDER = "#d1d9e0"
C_TEXT   = "#1a1a2e"
C_ACCENT = "#0078d4"

FONT_UI   = ("Segoe UI", 10)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_MONO = ("Consolas", 9)

APP_VERSION = "2.0"

# =============================================================================
# HTML TEMPLATE — student answer form
# =============================================================================

HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>__EXAM_NAME__ — Foglio Risposte</title>
<style>
  body { font-family:'Segoe UI',Arial,sans-serif; background:#f0f4f8; margin:0; padding:20px; color:#1a1a2e; }
  .container { max-width:800px; margin:auto; background:#fff; border-radius:12px; padding:30px; box-shadow:0 4px 20px rgba(0,0,0,.12); }
  h1 { color:#1F4E79; margin-bottom:4px; font-size:1.6em; }
  .subtitle { color:#555; margin-bottom:20px; font-size:.95em; }
  .phase { display:none; }
  .phase.active { display:block; }
  .student-section { background:#f0f4f8; border-radius:8px; padding:16px; margin-bottom:20px; }
  label { font-weight:600; display:block; margin-bottom:6px; }
  input[type=text] { width:100%; padding:10px; border:2px solid #d1d9e0; border-radius:6px; font-size:1em; box-sizing:border-box; }
  input[type=text]:focus { outline:none; border-color:#1F4E79; }
  .btn-start { display:block; width:100%; padding:14px; background:#1e5928; color:#fff; border:none; border-radius:8px; font-size:1.1em; font-weight:700; cursor:pointer; margin-top:16px; letter-spacing:.5px; }
  .btn-start:hover { background:#2a7a38; }
  #timer-bar { background:#dce6f0; border-radius:8px; padding:10px 16px; margin-bottom:16px; display:flex; align-items:center; justify-content:space-between; transition:background .4s; }
  #timer-bar.warn { background:#FFEB9C; }
  #timer-bar.alert { background:#FFC7CE; }
  #timer-val { font-size:1.6em; font-weight:700; color:#1F4E79; font-family:monospace; letter-spacing:2px; }
  .progress-bar { background:#f0f4f8; border-radius:6px; padding:7px 12px; margin-bottom:12px; font-size:.9em; color:#444; }
  .questions { column-count:3; column-gap:16px; }
  @media(max-width:600px) { .questions { column-count:2; } }
  .q-row { break-inside:avoid; background:#fff; border:1px solid #d1d9e0; border-radius:8px; padding:10px 14px; margin-bottom:10px; display:flex; align-items:center; gap:10px; }
  .q-num { font-weight:700; color:#1F4E79; min-width:28px; font-size:.95em; }
  .radio-group { display:flex; gap:14px; }
  .radio-group label { font-weight:normal; display:flex; align-items:center; gap:5px; cursor:pointer; margin:0; }
  input[type=radio] { accent-color:#1F4E79; width:16px; height:16px; }
  .btn { display:block; width:100%; padding:14px; background:#1F4E79; color:#fff; border:none; border-radius:8px; font-size:1.1em; font-weight:700; cursor:pointer; margin-top:24px; letter-spacing:.5px; }
  .btn:hover { background:#2E75B6; }
  #status { margin-top:14px; text-align:center; font-weight:600; min-height:22px; color:#1e5928; }
  .saved { background:#C6EFCE; border-radius:6px; padding:8px 14px; }
</style>
</head>
<body>
<div class="container">
  <h1>\U0001f4cb __EXAM_NAME__</h1>

  <!-- Fase 1: dati studente -->
  <div id="phase-setup" class="phase active">
    <p class="subtitle">__START_SUBTITLE__</p>
    <div class="student-section">
      <label for="sName">\U0001f464 Nome e Cognome</label>
      <input type="text" id="sName" placeholder="es. Rossi Mario" autocomplete="name"/>
    </div>
    <button class="btn-start" onclick="startExam()">▶ Avvia Esame</button>
  </div>

  <!-- Fase 2: domande -->
  <div id="phase-exam" class="phase">
    __TIMER_HTML__
    <div class="progress-bar" id="prog">0 / __NUM_QUESTIONS__ risposte</div>
    <div class="questions" id="questions"></div>
    <button class="btn" onclick="saveAnswers()">\U0001f4be Salva Risposte</button>
    <div id="status"></div>
  </div>
</div>

<script>
const N = __NUM_QUESTIONS__;
const CHOICES = __CHOICES_JSON__;
const DURATION = __DURATION_SECONDS__;
const EXAM_NAME = __EXAM_NAME_JS__;

let studentName = "";
let secondsLeft = DURATION;
let timerInterval = null;

function updateProgress() {
  let count = 0;
  for (let i = 1; i <= N; i++) {
    if (document.querySelector('input[name="q' + i + '"]:checked')) count++;
  }
  const pb = document.getElementById('prog');
  pb.textContent = count + ' / ' + N + ' risposte' + (count === N ? '  ✓ Tutte completate' : '');
  pb.style.color = count === N ? '#1e5928' : '#444';
}

function buildQuestions() {
  const qDiv = document.getElementById('questions');
  const frag = document.createDocumentFragment();
  for (let i = 1; i <= N; i++) {
    const row = document.createElement('div');
    row.className = 'q-row';
    const numSpan = document.createElement('span');
    numSpan.className = 'q-num';
    numSpan.textContent = i + '.';
    const rg = document.createElement('div');
    rg.className = 'radio-group';
    CHOICES.forEach(function(c) {
      const lbl = document.createElement('label');
      const inp = document.createElement('input');
      inp.type = 'radio';
      inp.name = 'q' + i;
      inp.value = c;
      inp.addEventListener('change', updateProgress);
      lbl.appendChild(inp);
      lbl.appendChild(document.createTextNode(' ' + c));
      rg.appendChild(lbl);
    });
    row.appendChild(numSpan);
    row.appendChild(rg);
    frag.appendChild(row);
  }
  qDiv.appendChild(frag);
}

function startExam() {
  const inp = document.getElementById('sName');
  const name = inp.value.trim();
  if (!name) {
    inp.style.border = '2px solid #c0392b';
    inp.focus();
    return;
  }
  inp.style.border = '';
  studentName = name;
  document.getElementById('phase-setup').style.display = 'none';
  document.getElementById('phase-exam').style.display = 'block';
  buildQuestions();
  if (DURATION > 0) startTimer();
}

function startTimer() {
  const bar = document.getElementById('timer-bar');
  const el  = document.getElementById('timer-val');
  function tick() {
    if (secondsLeft <= 0) {
      clearInterval(timerInterval);
      bar.className = 'alert';
      el.textContent = 'TEMPO SCADUTO';
      saveAnswers(true);
      return;
    }
    const m = Math.floor(secondsLeft / 60);
    const s = secondsLeft % 60;
    el.textContent = String(m).padStart(2, '0') + ':' + String(s).padStart(2, '0');
    if (secondsLeft <= 60) bar.className = 'alert';
    else if (secondsLeft <= 300) bar.className = 'warn';
    else bar.className = '';
    secondsLeft--;
  }
  tick();
  timerInterval = setInterval(tick, 1000);
}

function saveAnswers(auto) {
  const name = studentName || document.getElementById('sName').value.trim();
  if (!name && !auto) { alert('Avvia prima l\'esame!'); return; }
  const answers = [];
  for (let i = 1; i <= N; i++) {
    const r = document.querySelector(`input[name="q${i}"]:checked`);
    answers.push(r ? r.value : '');
  }
  if (!auto && answers.some(a => a === '')) {
    if (!confirm('Non tutte le domande hanno una risposta. Continuare?')) return;
  }
  if (timerInterval) clearInterval(timerInterval);
  const data = { name, exam_name: EXAM_NAME, num_questions: N, answers, timestamp: new Date().toISOString() };
  const blob = new Blob([JSON.stringify(data, null, 2)], { type: 'application/json' });
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = (name || 'studente').replace(/\s+/g, '_') + '_risposte.json';
  a.click();
  const st = document.getElementById('status');
  st.textContent = (auto ? '⏰ Tempo scaduto — ' : '✅ ') + 'Risposte salvate in ' + a.download;
  st.className = 'saved';
}
</script>
</body>
</html>
"""


def generate_student_html(exam_name: str, num_questions: int, output_path: str,
                          num_choices: int = 3, duration_minutes: int = 0):
    """Generate the student HTML answer form."""
    choices = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:max(2, min(num_choices, 26))])
    if duration_minutes > 0:
        start_subtitle = (f"Durata: <strong>{duration_minutes} minuti</strong>. "
                          "Inserisci il tuo nome e clicca <strong>Avvia Esame</strong>.")
        timer_html = ('<div id="timer-bar">'
                      '<span>&#x23F1; Tempo rimanente</span>'
                      '<span id="timer-val">--:--</span>'
                      '</div>')
    else:
        start_subtitle = "Inserisci il tuo nome e clicca <strong>Avvia Esame</strong>."
        timer_html = ""
    html = (HTML_TEMPLATE
            .replace("__EXAM_NAME__", exam_name)
            .replace("__EXAM_NAME_JS__", json.dumps(exam_name))
            .replace("__NUM_QUESTIONS__", str(num_questions))
            .replace("__CHOICES_JSON__", json.dumps(choices))
            .replace("__DURATION_SECONDS__", str(duration_minutes * 60))
            .replace("__TIMER_HTML__", timer_html)
            .replace("__START_SUBTITLE__", start_subtitle)
            )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)


# =============================================================================
# DOMAIN MODEL
# =============================================================================

@dataclass
class ExamConfig:
    exam_name: str
    num_questions: int
    correct_answers: List[str]    # one letter per question (index 0 = Q1)
    course_name: str   = ""
    scale: float       = 30.0
    passing_pct: float = 75.0      # percentage (0-100)
    decimal_places: int = 1
    num_choices: int   = 3
    duration_minutes: int = 0

    @property
    def passing_score(self) -> float:
        """Raw (unrounded) passing threshold — used only for display."""
        return self.scale * self.passing_pct / 100

    def score(self, num_correct: int) -> float:
        return round(num_correct / self.num_questions * self.scale, self.decimal_places)

    def fmt_score(self, s: float) -> str:
        return f"{s:.{self.decimal_places}f}"

    def score_label(self) -> str:
        sc = int(self.scale) if self.scale == int(self.scale) else self.scale
        return f"/{sc}"


@dataclass
class StudentResult:
    student_name: str
    student_file: str
    answers: List[str]
    correct_mask: List[bool] = field(default_factory=list)
    num_correct: int = 0
    score: float = 0.0
    passed: bool = False

    def compute(self, config: "ExamConfig"):
        self.correct_mask = [
            a.upper() == c.upper()
            for a, c in zip(self.answers, config.correct_answers)
        ]
        self.num_correct = sum(self.correct_mask)
        self.score = config.score(self.num_correct)
        # Compare raw ratio to avoid rounding edge cases near the cutoff
        raw_ratio = self.num_correct / config.num_questions if config.num_questions else 0
        self.passed = raw_ratio >= config.passing_pct / 100


def correct_student(data: Dict, config: ExamConfig) -> "StudentResult":
    """Grade a single student\'s answers."""
    answers = [str(a).upper() if a else "" for a in data.get("answers", [])]
    while len(answers) < config.num_questions:
        answers.append("")
    answers = answers[:config.num_questions]
    result = StudentResult(
        student_name=data.get("name") or data.get("student_name") or Path(data.get("file", "?")).stem,
        student_file=data.get("file", ""),
        answers=answers,
    )
    result.compute(config)
    return result


def compute_stats(results: List["StudentResult"], config: ExamConfig) -> dict:
    """Compute exam-level statistics."""
    if not results:
        return {}
    n = len(results)
    scores = [r.score for r in results]
    passed = [r for r in results if r.passed]

    q_errors = []
    for qi in range(config.num_questions):
        wrong = sum(1 for r in results if qi < len(r.correct_mask) and not r.correct_mask[qi])
        q_errors.append(wrong / n * 100)

    p = [100.0 - e for e in q_errors]

    k = max(1, int(n * 0.27))
    sorted_results = sorted(results, key=lambda r: r.score)
    bottom = sorted_results[:k]
    top    = sorted_results[-k:]
    disc = []
    for qi in range(config.num_questions):
        p_top    = sum(1 for r in top    if qi < len(r.correct_mask) and r.correct_mask[qi]) / k
        p_bottom = sum(1 for r in bottom if qi < len(r.correct_mask) and r.correct_mask[qi]) / k
        disc.append(round(p_top - p_bottom, 3))

    mean = sum(scores) / n
    return {
        "n":          n,
        "mean":       round(mean, 2),
        "median":     round(sorted(scores)[n // 2], 2),
        "std":        round(math.sqrt(sum((s - mean)**2 for s in scores) / n), 2),
        "min":        min(scores),
        "max":        max(scores),
        "n_passed":   len(passed),
        "pct_passed": round(len(passed) / n * 100, 1),
        "q_errors":   q_errors,
        "p":          p,
        "disc":       disc,
        "scores":     scores,
    }


# =============================================================================
# PDF ANSWER EXTRACTION (PyMuPDF + PIL + pdfplumber)
# =============================================================================

def _parse_qb_meta_only(pdf_path: str) -> Optional[Dict]:
    """Fallback: extract exam name, course name and question count."""
    if not HAS_PDFPLUMBER:
        return None
    try:
        with pdfplumber.open(pdf_path) as pdf_doc:
            page = pdf_doc.pages[0]
            text = page.extract_text() or ""
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        exam_name = lines[0] if lines else Path(pdf_path).stem
        course_name = ""
        n_questions = None
        for i, line in enumerate(lines[1:], 1):
            m = re.search(r"/(\d+)\s*$", line)
            if m:
                n_questions = int(m.group(1))
                break
            # Heuristic: second non-empty line that isn't a question-count → course name
            if i == 1 and len(line) < 60 and not re.search(r"^\d+[./]", line):
                course_name = line
        return {"exam_name": exam_name, "course_name": course_name,
                "num_questions": n_questions, "answers": {}}
    except Exception:
        return None


def parse_qb_marksheet(pdf_path: str) -> Optional[Dict]:
    """Extract correct answers from a QBGenerator PDF mark sheet."""
    if not HAS_FITZ or not HAS_PDFPLUMBER:
        return _parse_qb_meta_only(pdf_path)
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]

        imgs = page.get_images(full=True)
        img_brightness: Dict[int, float] = {}
        for img in imgs:
            xref = img[0]
            base_image = doc.extract_image(xref)
            pil_img = Image.open(io.BytesIO(base_image["image"])).convert("L")
            if HAS_NUMPY:
                avg = float(np.array(pil_img).mean())
            else:
                pixels = list(pil_img.getdata())
                avg = sum(pixels) / len(pixels) if pixels else 255.0
            img_brightness[xref] = avg

        brightnesses = sorted(img_brightness.values())
        if len(brightnesses) < 2:
            doc.close()
            return _parse_qb_meta_only(pdf_path)

        threshold = (brightnesses[0] + brightnesses[-1]) / 2
        is_filled = {xref: b < threshold for xref, b in img_brightness.items()}

        img_placements = page.get_image_info(xrefs=True)
        doc.close()

        with pdfplumber.open(pdf_path) as pdf_doc:
            plumb_page = pdf_doc.pages[0]
            words = plumb_page.extract_words()
            text  = plumb_page.extract_text() or ""

        lines = [l.strip() for l in text.split("\n") if l.strip()]
        exam_name   = lines[0] if lines else Path(pdf_path).stem
        course_name = ""
        n_questions = None
        for i, line in enumerate(lines[1:], 1):
            m = re.search(r"/(\d+)\s*$", line)
            if m:
                n_questions = int(m.group(1))
                break
            if i == 1 and len(line) < 60 and not re.search(r"^\d+[./]", line):
                course_name = line

        q_positions: Dict[int, Dict] = {}
        for w in words:
            try:
                q = int(w["text"])
                if 1 <= q <= 200:
                    q_y = (w["top"] + w.get("bottom", w["top"] + 10)) / 2
                    if q_y > 90 and q not in q_positions:
                        q_positions[q] = {"x": w["x0"], "y": q_y}
            except ValueError:
                pass

        answers: Dict[int, str] = {}
        for q_num, q_pos in sorted(q_positions.items()):
            q_y, q_x = q_pos["y"], q_pos["x"]
            row_imgs = []
            for ip in img_placements:
                bbox = ip.get("bbox", [])
                if len(bbox) < 4:
                    continue
                img_y = (bbox[1] + bbox[3]) / 2
                img_x = (bbox[0] + bbox[2]) / 2
                if abs(img_y - q_y) < 14 and q_x < img_x < q_x + 120:
                    row_imgs.append({"xref": ip["xref"], "x": img_x})
            row_imgs.sort(key=lambda x: x["x"])
            for i, ri in enumerate(row_imgs[:3]):
                if is_filled.get(ri["xref"], False):
                    answers[q_num] = ["A", "B", "C"][i]
                    break

        return {
            "exam_name":     exam_name,
            "course_name":   course_name,
            "num_questions": n_questions,
            "answers":       answers,
        }
    except Exception:
        return _parse_qb_meta_only(pdf_path)


# =============================================================================
# EXCEL EXPORTS
# =============================================================================

def _apply_cell_style(ws, row, col, value=None, bold=False, fill_hex=None,
                      align="center", border=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, name="Calibri", size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    if border:
        side = Side(style="thin")
        cell.border = Border(left=side, right=side, top=side, bottom=side)
    return cell


def export_csv(results: List[StudentResult], config: ExamConfig, output_path: str):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Studente", "File", "Corrette", "Totale",
                         f"Voto{config.score_label()}", "Esito"])
        for r in results:
            writer.writerow([
                r.student_name, r.student_file,
                r.num_correct, config.num_questions,
                r.score, "Idoneo" if r.passed else "Non idoneo",
            ])


def export_excel(results: List[StudentResult], config: ExamConfig, output_path: str):
    if not HAS_EXCEL:
        raise ImportError("openpyxl non installato")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Voti"
    headers = ["Studente", "File", "Corrette", "Totale",
               f"Voto{config.score_label()}", "Esito"]
    for ci, h in enumerate(headers, 1):
        cell = _apply_cell_style(ws, 1, ci, h, bold=True, fill_hex="1F4E79",
                                  align="center", border=True)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    num_fmt = f"0.{'0' * config.decimal_places}" if config.decimal_places > 0 else "0"
    for ri, r in enumerate(results, 2):
        fill = "C6EFCE" if r.passed else "FFC7CE"
        vals = [r.student_name, r.student_file, r.num_correct,
                config.num_questions, r.score, "Idoneo" if r.passed else "Non idoneo"]
        for ci, v in enumerate(vals, 1):
            cell = _apply_cell_style(ws, ri, ci, v, fill_hex=fill, border=True)
            if ci == 5:
                cell.number_format = num_fmt
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    wb.save(output_path)


def export_student_reports(results: List[StudentResult], config: ExamConfig, output_path: str):
    if not HAS_EXCEL:
        raise ImportError("openpyxl non installato")
    wb = openpyxl.Workbook()
    for r in results:
        ws = wb.create_sheet(title=r.student_name[:30])
        ws.merge_cells("A1:F1")
        header_text = f"Report esame: {config.exam_name}"
        if config.course_name:
            header_text += f" — {config.course_name}"
        ws["A1"] = header_text
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = f"Studente: {r.student_name}"
        ws["A3"] = (f"Voto: {config.fmt_score(r.score)}{config.score_label()}  |  "
                    f"Corrette: {r.num_correct}/{config.num_questions}"
                    f"  |  Esito: {'Idoneo' if r.passed else 'Non idoneo'}")
        ws["A3"].font = Font(bold=True)
        headers = ["Domanda", "Risposta Studente", "Risposta Corretta", "Esito"]
        for ci, h in enumerate(headers, 1):
            cell = _apply_cell_style(ws, 5, ci, h, bold=True, fill_hex="1F4E79", border=True)
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        for qi in range(config.num_questions):
            correct = qi < len(r.correct_mask) and r.correct_mask[qi]
            fill = "C6EFCE" if correct else "FFC7CE"
            vals = [
                qi + 1,
                r.answers[qi] if qi < len(r.answers) else "",
                config.correct_answers[qi] if qi < len(config.correct_answers) else "",
                "✓" if correct else "✗",
            ]
            for ci, v in enumerate(vals, 1):
                _apply_cell_style(ws, qi + 6, ci, v, fill_hex=fill, border=True)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(output_path)


def export_item_analysis(results: List[StudentResult], config: ExamConfig,
                          stats: dict, output_path: str, session_name: str = "",
                          append_to: Optional[str] = None):
    if not HAS_EXCEL:
        raise ImportError("openpyxl non installato")
    if append_to and Path(append_to).exists():
        wb = openpyxl.load_workbook(append_to)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    sheet_name = (session_name or config.exam_name)[:30]
    ws = wb.create_sheet(title=sheet_name)
    ws["A1"] = f"Item Analysis — {config.exam_name}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = (f"Data: {datetime.now().strftime('%d/%m/%Y')}   "
                f"N={stats.get('n', len(results))}   "
                f"Media={stats.get('mean', '')}   Pass={stats.get('pct_passed', '')}%")
    choices = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:config.num_choices])
    headers = ["Dom.", "Difficolta p (%)", "Disc. D", "% Errori"] + [f"{c} (%)" for c in choices]
    for ci, h in enumerate(headers, 1):
        cell = _apply_cell_style(ws, 4, ci, h, bold=True, fill_hex="1F4E79", border=True)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    n = len(results)
    p_list    = stats.get("p", [])
    disc_list = stats.get("disc", [])
    q_errors  = stats.get("q_errors", [])
    for qi in range(config.num_questions):
        row = qi + 5
        counts = {c: 0 for c in choices}
        for r in results:
            ans = r.answers[qi].upper() if qi < len(r.answers) else ""
            if ans in counts:
                counts[ans] += 1
        p_val    = round(p_list[qi], 1)    if qi < len(p_list)    else ""
        disc_val = round(disc_list[qi], 3) if qi < len(disc_list) else ""
        err_val  = round(q_errors[qi], 1)  if qi < len(q_errors)  else ""
        if isinstance(p_val, float):
            fill = "C6EFCE" if p_val >= 70 else ("FFEB9C" if p_val >= 40 else "FFC7CE")
        else:
            fill = None
        vals = [qi + 1, p_val, disc_val, err_val] + [
            round(counts[c] / n * 100, 1) if n else 0 for c in choices
        ]
        for ci, v in enumerate(vals, 1):
            _apply_cell_style(ws, row, ci, v, fill_hex=fill, border=True)
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 18
    wb.save(output_path)


def export_kpi(results: List[StudentResult], config: ExamConfig,
               stats: dict, output_path: str, session_name: str = "",
               append_to: Optional[str] = None):
    if not HAS_EXCEL:
        raise ImportError("openpyxl non installato")
    if append_to and Path(append_to).exists():
        wb = openpyxl.load_workbook(append_to)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    sheet_name = (session_name or config.exam_name)[:30]
    ws = wb.create_sheet(title=sheet_name)
    ws["A1"] = f"KPI Esame — {config.exam_name}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    p_list = stats.get("p", [])
    d_list = stats.get("disc", [])
    kpi_rows = [
        ("Sessione",         session_name or config.exam_name),
        ("Corso",            config.course_name),
        ("Data",             datetime.now().strftime("%d/%m/%Y")),
        ("N studenti",       stats.get("n", len(results))),
        ("Media voto",       stats.get("mean", "")),
        ("Mediana voto",     stats.get("median", "")),
        ("Dev std",          stats.get("std", "")),
        ("Voto minimo",      stats.get("min", "")),
        ("Voto massimo",     stats.get("max", "")),
        ("N idonei",         stats.get("n_passed", "")),
        ("% Idonei",         stats.get("pct_passed", "")),
        ("Soglia suff.",     f"{config.fmt_score(config.passing_score)}{config.score_label()} ({config.passing_pct:.0f}%)"),
        ("N domande",        config.num_questions),
        ("Diff. media (%)",  round(sum(p_list) / len(p_list), 1) if p_list else ""),
        ("Disc. media",      round(sum(d_list) / len(d_list), 3) if d_list else ""),
    ]
    for ri, (k, v) in enumerate(kpi_rows, 3):
        ws.cell(row=ri, column=1, value=k).font = Font(bold=True, name="Calibri")
        ws.cell(row=ri, column=2, value=v)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    wb.save(output_path)


# =============================================================================
# PDF PRINTING (reportlab)
# =============================================================================

def print_student_reports_pdf(results: List[StudentResult], config: ExamConfig,
                               output_path: str):
    """Generate a PDF with one page per student showing their answers."""
    if not HAS_REPORTLAB:
        raise ImportError("reportlab non installato")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                  textColor=colors.HexColor("#1F4E79"),
                                  fontSize=14, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                                fontSize=11, spaceAfter=4)
    doc_pdf = SimpleDocTemplate(output_path, pagesize=A4,
                                 rightMargin=1.5*cm, leftMargin=1.5*cm,
                                 topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    for idx, r in enumerate(results):
        if idx > 0:
            story.append(PageBreak())
        exam_header = f"Esame: {config.exam_name}"
        if config.course_name:
            exam_header += f" — {config.course_name}"
        story.append(Paragraph(exam_header, title_style))
        story.append(Paragraph(f"<b>Studente:</b> {r.student_name}", sub_style))
        story.append(Paragraph(
            f"<b>Voto:</b> {config.fmt_score(r.score)}{config.score_label()} &nbsp;&nbsp; "
            f"<b>Corrette:</b> {r.num_correct}/{config.num_questions} &nbsp;&nbsp; "
            f"<b>Esito:</b> {'IDONEO' if r.passed else 'NON IDONEO'}",
            sub_style,
        ))
        story.append(Spacer(1, 0.3*cm))
        story.append(HRFlowable(width="100%", thickness=1,
                                 color=colors.HexColor("#1F4E79")))
        story.append(Spacer(1, 0.3*cm))
        half = math.ceil(config.num_questions / 2)
        col_hdr = ["Dom.", "Studente", "Corretta", "Esito"]
        header_row = col_hdr + [""] + col_hdr
        col_widths = [1.2*cm, 1.8*cm, 1.8*cm, 1.2*cm, 0.4*cm,
                      1.2*cm, 1.8*cm, 1.8*cm, 1.2*cm]
        table_data = [header_row]
        table_style_cmds = [
            ("BACKGROUND", (0, 0), (3, 0), colors.HexColor("#1F4E79")),
            ("BACKGROUND", (5, 0), (8, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR",  (0, 0), (3, 0), colors.white),
            ("TEXTCOLOR",  (5, 0), (8, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1,  0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",       (0, 0), (3, -1), 0.5, colors.grey),
            ("GRID",       (5, 0), (8, -1), 0.5, colors.grey),
        ]
        for row_i in range(half):
            qi_left  = row_i
            qi_right = row_i + half

            def make_cells(qi):
                if qi >= config.num_questions:
                    return ["", "", "", ""]
                correct = qi < len(r.correct_mask) and r.correct_mask[qi]
                ans = r.answers[qi] if qi < len(r.answers) else ""
                ca  = config.correct_answers[qi] if qi < len(config.correct_answers) else ""
                return [str(qi + 1), ans, ca, "OK" if correct else "X"]

            left  = make_cells(qi_left)
            right = make_cells(qi_right)
            table_data.append(left + [""] + right)
            data_row = row_i + 1
            if qi_left < config.num_questions:
                c_left = qi_left < len(r.correct_mask) and r.correct_mask[qi_left]
                fill_l = colors.HexColor("#C6EFCE" if c_left else "#FFC7CE")
                table_style_cmds.append(("BACKGROUND", (0, data_row), (3, data_row), fill_l))
            if qi_right < config.num_questions:
                c_right = qi_right < len(r.correct_mask) and r.correct_mask[qi_right]
                fill_r = colors.HexColor("#C6EFCE" if c_right else "#FFC7CE")
                table_style_cmds.append(("BACKGROUND", (5, data_row), (8, data_row), fill_r))
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle(table_style_cmds))
        story.append(tbl)
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph(
            f"<i>Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')} "
            f"— ExamCorrector v{APP_VERSION}</i>",
            styles["Normal"],
        ))
    doc_pdf.build(story)


def print_statistics_pdf(results: List[StudentResult], config: ExamConfig,
                          stats: dict, output_path: str):
    """Generate a PDF with exam statistics."""
    if not HAS_REPORTLAB:
        raise ImportError("reportlab non installato")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title2", parent=styles["Heading1"],
                                  textColor=colors.HexColor("#1F4E79"),
                                  fontSize=16, spaceAfter=8)
    h2_style = ParagraphStyle("h2b", parent=styles["Heading2"],
                               textColor=colors.HexColor("#2E75B6"),
                               fontSize=12, spaceAfter=6)
    normal = styles["Normal"]
    doc_pdf = SimpleDocTemplate(output_path, pagesize=A4,
                                 rightMargin=1.5*cm, leftMargin=1.5*cm,
                                 topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    exam_title = f"Statistiche Esame: {config.exam_name}"
    if config.course_name:
        exam_title += f" — {config.course_name}"
    story.append(Paragraph(exam_title, title_style))
    story.append(Paragraph(
        f"Data: {datetime.now().strftime('%d/%m/%Y')}   "
        f"N studenti: {stats.get('n', len(results))}",
        normal,
    ))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("KPI Principali", h2_style))
    kpi_data = [
        ["Indicatore", "Valore"],
        ["Media voto",        str(stats.get("mean", ""))],
        ["Mediana",           str(stats.get("median", ""))],
        ["Dev. std.",         str(stats.get("std", ""))],
        ["Minimo",            str(stats.get("min", ""))],
        ["Massimo",           str(stats.get("max", ""))],
        ["Idonei",            f"{stats.get('n_passed', '')} ({stats.get('pct_passed', '')}%)"],
        ["Soglia sufficienza", f"{config.fmt_score(config.passing_score)}{config.score_label()} ({config.passing_pct:.0f}%)"],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[7*cm, 5*cm])
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#f0f4f8"), colors.white]),
        ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph("Item Analysis", h2_style))
    ia_header = ["Dom.", "Diff. p(%)", "Disc. D", "% Errori"]
    ia_data   = [ia_header]
    p_list    = stats.get("p", [])
    disc_list = stats.get("disc", [])
    err_list  = stats.get("q_errors", [])
    ia_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
    ]
    for qi in range(config.num_questions):
        p_val = round(p_list[qi], 1)    if qi < len(p_list)    else ""
        d_val = round(disc_list[qi], 3) if qi < len(disc_list) else ""
        e_val = round(err_list[qi], 1)  if qi < len(err_list)  else ""
        ia_data.append([str(qi + 1), str(p_val), str(d_val), str(e_val)])
        if isinstance(p_val, float):
            bg = colors.HexColor(
                "#C6EFCE" if p_val >= 70 else ("#FFEB9C" if p_val >= 40 else "#FFC7CE")
            )
            ia_style_cmds.append(("BACKGROUND", (0, qi + 1), (-1, qi + 1), bg))
    ia_tbl = Table(ia_data, colWidths=[2*cm, 4*cm, 3*cm, 3*cm], repeatRows=1)
    ia_tbl.setStyle(TableStyle(ia_style_cmds))
    story.append(ia_tbl)
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"<i>Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')} "
        f"— ExamCorrector v{APP_VERSION}</i>",
        normal,
    ))
    doc_pdf.build(story)


# =============================================================================
# TEACHER APPLICATION GUI
# =============================================================================

class TeacherApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"ExamCorrector v{APP_VERSION} — Applicazione Docente")
        self.geometry("1120x760")
        self.configure(bg=C_BG)
        self.minsize(900, 620)
        self._config:  Optional[ExamConfig]  = None
        self._results: List[StudentResult]   = []
        self._stats:   dict                  = {}
        self._answer_vars: List[tk.StringVar] = []
        self._build_ui()
        self.after(400, self._setup_dnd)

    # ── Style ─────────────────────────────────────────────────────────────────

    def _apply_style(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("TNotebook",       background=C_BG,      borderwidth=0)
        s.configure("TNotebook.Tab",   background="#dce6f0",  foreground=C_TEXT,
                    padding=[16, 7],   font=FONT_BOLD)
        s.map("TNotebook.Tab",
              background=[("selected", C_NAVY), ("active", C_BLUE)],
              foreground=[("selected", "white"), ("active", "white")])
        s.configure("TFrame",          background=C_BG)
        s.configure("TLabel",          background=C_BG, foreground=C_TEXT, font=FONT_UI)
        s.configure("TButton",         font=FONT_UI, padding=[8, 5])
        s.configure("TEntry",          fieldbackground=C_WHITE, font=FONT_UI)
        s.configure("TSpinbox",        fieldbackground=C_WHITE, font=FONT_UI)
        s.configure("Treeview",        font=FONT_UI, rowheight=24,
                    fieldbackground=C_WHITE, background=C_WHITE)
        s.configure("Treeview.Heading", font=FONT_BOLD, background="#dce6f0")
        s.configure("Accent.TButton",  background=C_NAVY, foreground="white", font=FONT_BOLD)
        s.map("Accent.TButton",
              background=[("active", C_BLUE), ("pressed", "#163a5a")])

    # ── UI Build ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._apply_style()
        header = tk.Frame(self, bg=C_NAVY, height=58)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)
        hf = tk.Frame(header, bg=C_NAVY)
        hf.pack(side="left", fill="y", padx=14, pady=8)
        tk.Label(hf, text=f"ExamCorrector  v{APP_VERSION}",
                 bg=C_NAVY, fg="white", font=("Segoe UI", 13, "bold")).pack(anchor="w")
        self._header_sub = tk.StringVar(value="Applicazione Docente")
        tk.Label(hf, textvariable=self._header_sub,
                 bg=C_NAVY, fg="#7db3e0", font=("Segoe UI", 8)).pack(anchor="w")
        self._status_var = tk.StringVar(value="Pronto")
        tk.Label(header, textvariable=self._status_var,
                 bg=C_NAVY, fg="#a0c4ff", font=FONT_UI).pack(side="right", padx=18)
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill="both", expand=True, padx=8, pady=8)
        self._tab1 = ttk.Frame(self._nb)
        self._tab2 = ttk.Frame(self._nb)
        self._tab3 = ttk.Frame(self._nb)
        self._nb.add(self._tab1, text="  1 \xb7 Configura Esame  ")
        self._nb.add(self._tab2, text="  2 \xb7 Correggi Esami  ")
        self._nb.add(self._tab3, text="  3 \xb7 Statistiche & Report  ")
        self._build_tab1()
        self._build_tab2()
        self._build_tab3()

    # ── TAB 1 ─────────────────────────────────────────────────────────────────

    def _build_tab1(self):
        tab = self._tab1
        tab.columnconfigure(0, weight=1)

        info = tk.LabelFrame(tab, text="  \U0001f4cb  Informazioni Esame  ",
                              bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        info.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))
        info.columnconfigure(1, weight=1)
        tk.Label(info, text="Nome esame:", bg=C_BG, font=FONT_UI).grid(
            row=0, column=0, sticky="w", padx=10, pady=6)
        self._name_var = tk.StringVar(value="Esame")
        ttk.Entry(info, textvariable=self._name_var, width=44).grid(
            row=0, column=1, sticky="ew", padx=8, pady=6)
        tk.Label(info, text="N\xb0 domande:", bg=C_BG, font=FONT_UI).grid(
            row=1, column=0, sticky="w", padx=10, pady=6)
        self._nq_var = tk.IntVar(value=10)
        ttk.Spinbox(info, from_=1, to=200, textvariable=self._nq_var, width=7).grid(
            row=1, column=1, sticky="w", padx=8)
        ttk.Button(info, text="\u21bb Aggiorna griglia",
                   command=self._refresh_grid).grid(row=1, column=1, sticky="e", padx=8)
        tk.Label(info, text="Nome corso:", bg=C_BG, font=FONT_UI).grid(
            row=2, column=0, sticky="w", padx=10, pady=6)
        self._course_var = tk.StringVar(value="")
        ttk.Entry(info, textvariable=self._course_var, width=44).grid(
            row=2, column=1, sticky="ew", padx=8, pady=6)

        params_card = tk.LabelFrame(tab, text="  \u2696  Parametri Valutazione  ",
                                     bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        params_card.grid(row=1, column=0, sticky="ew", padx=12, pady=4)
        # Scale
        tk.Label(params_card, text="Scala voto:", bg=C_BG, font=FONT_UI).grid(
            row=0, column=0, sticky="w", padx=10, pady=6)
        self._scale_mode_var = tk.StringVar(value="30simi")
        scale_cb = ttk.Combobox(params_card, textvariable=self._scale_mode_var,
                                 values=["30simi", "10simi", "100simi", "Personalizzato"],
                                 state="readonly", width=14)
        scale_cb.grid(row=0, column=1, padx=4, pady=6, sticky="w")
        tk.Label(params_card, text="Valore:", bg=C_BG, font=FONT_UI).grid(
            row=0, column=2, sticky="w", padx=(12, 2))
        self._scale_custom_var = tk.StringVar(value="30")
        self._scale_custom_entry = ttk.Entry(params_card, textvariable=self._scale_custom_var,
                                              width=7, state="disabled")
        self._scale_custom_entry.grid(row=0, column=3, padx=4, sticky="w")
        scale_cb.bind("<<ComboboxSelected>>", self._on_scale_mode_changed)
        # Passing %
        tk.Label(params_card, text="Soglia (%):", bg=C_BG, font=FONT_UI).grid(
            row=0, column=4, sticky="w", padx=(20, 2))
        self._passing_pct_var = tk.DoubleVar(value=75.0)
        ttk.Spinbox(params_card, from_=1, to=100, increment=0.5, format="%.1f",
                    textvariable=self._passing_pct_var, width=7).grid(
            row=0, column=5, padx=4, sticky="w")
        # Decimal places
        tk.Label(params_card, text="Cifre decimali:", bg=C_BG, font=FONT_UI).grid(
            row=1, column=0, sticky="w", padx=10, pady=6)
        self._decimal_var = tk.IntVar(value=1)
        ttk.Spinbox(params_card, from_=0, to=4, textvariable=self._decimal_var,
                    width=5).grid(row=1, column=1, padx=4, sticky="w")
        # Num choices
        tk.Label(params_card, text="Alternative rsp.:", bg=C_BG, font=FONT_UI).grid(
            row=1, column=2, sticky="w", padx=(12, 2))
        self._num_choices_var = tk.IntVar(value=3)
        choices_cb = ttk.Combobox(params_card, textvariable=self._num_choices_var,
                                   values=[2, 3, 4, 5], state="readonly", width=5)
        choices_cb.grid(row=1, column=3, padx=4, sticky="w")
        choices_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_grid())
        tk.Label(params_card, text="(A-B, A-B-C, A-B-C-D, A-B-C-D-E) — auto da PDF",
                 bg=C_BG, fg="#666", font=("Segoe UI", 8)).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=4)
        # Duration
        tk.Label(params_card, text="Durata (min):", bg=C_BG, font=FONT_UI).grid(
            row=2, column=0, sticky="w", padx=10, pady=6)
        self._duration_var = tk.IntVar(value=0)
        ttk.Spinbox(params_card, from_=0, to=300, increment=5,
                    textvariable=self._duration_var, width=7).grid(
            row=2, column=1, padx=4, sticky="w")
        tk.Label(params_card, text="0 = nessun timer  —  incluso in exam_info.json e nel modulo HTML",
                 bg=C_BG, fg="#666", font=("Segoe UI", 8)).grid(
            row=2, column=2, columnspan=4, sticky="w", padx=4)

        pdf_card = tk.LabelFrame(tab, text="  \U0001f4c4  Importa PDF Marksheet  ",
                                  bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        pdf_card.grid(row=2, column=0, sticky="ew", padx=12, pady=4)
        pdf_card.columnconfigure(0, weight=1)
        self._pdf_drop_frame = tk.Frame(pdf_card, bg="#e8f0fe", relief="groove", bd=2,
                                         cursor="hand2", height=70)
        self._pdf_drop_frame.grid(row=0, column=0, sticky="ew", padx=8, pady=6)
        self._pdf_drop_frame.columnconfigure(0, weight=1)
        self._pdf_drop_frame.grid_propagate(False)
        self._pdf_drop_label = tk.Label(
            self._pdf_drop_frame,
            text="\u2b07  Trascina qui il PDF marksheet oppure clicca per sfogliare",
            bg="#e8f0fe", fg=C_NAVY, font=FONT_BOLD, cursor="hand2",
        )
        self._pdf_drop_label.grid(row=0, column=0, pady=20)
        self._pdf_drop_frame.bind("<Button-1>", lambda e: self._browse_pdf())
        self._pdf_drop_label.bind("<Button-1>", lambda e: self._browse_pdf())
        btn_row = tk.Frame(pdf_card, bg=C_BG)
        btn_row.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 6))
        ttk.Button(btn_row, text="\U0001f4c2 Sfoglia PDF",
                   command=self._browse_pdf).pack(side="left", padx=4)
        ttk.Button(btn_row, text="\u26a1 Estrai risposte", style="Accent.TButton",
                   command=self._import_pdf_answers).pack(side="left", padx=4)
        self._pdf_path_var = tk.StringVar()
        tk.Label(btn_row, textvariable=self._pdf_path_var,
                 bg=C_BG, fg="#555", font=("Segoe UI", 9)).pack(side="left", padx=8)

        grid_card = tk.LabelFrame(tab, text="  \u270f  Risposte Corrette  ",
                                   bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        grid_card.grid(row=3, column=0, sticky="nsew", padx=12, pady=4)
        tab.rowconfigure(3, weight=1)
        grid_card.columnconfigure(0, weight=1)
        grid_card.rowconfigure(0, weight=1)
        self._grid_canvas = tk.Canvas(grid_card, bg=C_BG, highlightthickness=0)
        self._grid_canvas.grid(row=0, column=0, sticky="nsew")
        grid_sb = ttk.Scrollbar(grid_card, orient="vertical",
                                 command=self._grid_canvas.yview)
        grid_sb.grid(row=0, column=1, sticky="ns")
        self._grid_canvas.configure(yscrollcommand=grid_sb.set)
        self._grid_inner = tk.Frame(self._grid_canvas, bg=C_BG)
        self._grid_canvas.create_window((0, 0), window=self._grid_inner, anchor="nw")
        self._grid_inner.bind("<Configure>",
                              lambda e: self._grid_canvas.configure(
                                  scrollregion=self._grid_canvas.bbox("all")))

        act = tk.Frame(tab, bg=C_BG)
        act.grid(row=4, column=0, sticky="ew", padx=12, pady=(4, 8))
        ttk.Button(act, text="\U0001f4be Salva configurazione",
                   command=self._save_config).pack(side="left", padx=4)
        ttk.Button(act, text="\U0001f4c2 Carica configurazione",
                   command=self._load_config).pack(side="left", padx=4)
        tk.Frame(act, bg=C_BORDER, width=2, height=28).pack(side="left", fill="y", padx=8)
        ttk.Button(act, text="\U0001f310 Genera modulo studenti (HTML)",
                   style="Accent.TButton",
                   command=self._generate_html).pack(side="left", padx=4)
        self._tab1_status = tk.StringVar()
        tk.Label(act, textvariable=self._tab1_status,
                 fg="#1e5928", bg=C_BG, font=FONT_UI).pack(side="left", padx=12)
        self._refresh_grid()

    def _on_scale_mode_changed(self, event=None):
        mode = self._scale_mode_var.get()
        if mode == "Personalizzato":
            self._scale_custom_entry.config(state="normal")
        else:
            mapping = {"30simi": "30", "10simi": "10", "100simi": "100"}
            self._scale_custom_var.set(mapping.get(mode, "30"))
            self._scale_custom_entry.config(state="disabled")

    def _refresh_grid(self):
        try:
            nq = int(self._nq_var.get())
        except Exception:
            nq = 10
        try:
            nc = int(self._num_choices_var.get())
        except Exception:
            nc = 3
        nc = max(2, min(nc, 26))
        letters = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:nc])
        # Preserve existing valid answers before destroying widgets
        old_vars = list(self._answer_vars)
        for w in self._grid_inner.winfo_children():
            w.destroy()
        self._answer_vars.clear()
        cols = 5
        for qi in range(nq):
            r, c = divmod(qi, cols)
            cell = tk.Frame(self._grid_inner, bg=C_WHITE,
                            highlightbackground=C_BORDER, highlightthickness=1,
                            padx=4, pady=3)
            cell.grid(row=r, column=c, padx=3, pady=3, sticky="nsew")
            tk.Label(cell, text=str(qi + 1), bg=C_NAVY, fg="white",
                     font=("Segoe UI", 8, "bold"), width=3, pady=1).pack(side="left", padx=(0, 5))
            old_val = old_vars[qi].get() if qi < len(old_vars) else letters[0]
            default_val = old_val if old_val in letters else letters[0]
            var = tk.StringVar(value=default_val)
            self._answer_vars.append(var)
            for letter in letters:
                tk.Radiobutton(
                    cell, text=letter, variable=var, value=letter,
                    indicatoron=False,
                    bg="#4a86c1", fg="white",
                    selectcolor=C_NAVY,
                    activebackground=C_BLUE, activeforeground="white",
                    relief="flat", bd=0,
                    font=("Segoe UI", 9, "bold"),
                    width=2, pady=2, cursor="hand2",
                ).pack(side="left", padx=1)
        for c in range(cols):
            self._grid_inner.columnconfigure(c, weight=1)

    def _get_config(self) -> Optional[ExamConfig]:
        try:
            nq = int(self._nq_var.get())
        except Exception:
            messagebox.showerror("Errore", "Numero domande non valido.")
            return None
        answers = [v.get() for v in self._answer_vars[:nq]]
        if len(answers) < nq:
            messagebox.showerror("Errore", "Griglia incompleta. Aggiorna prima la griglia.")
            return None
        course_name = self._course_var.get().strip()
        mode = self._scale_mode_var.get()
        if mode == "10simi":
            scale = 10.0
        elif mode == "100simi":
            scale = 100.0
        elif mode == "Personalizzato":
            try:
                scale = float(self._scale_custom_var.get())
            except Exception:
                scale = 30.0
        else:
            scale = 30.0
        try:
            passing_pct = float(self._passing_pct_var.get())
        except Exception:
            passing_pct = 75.0
        try:
            decimal_places = int(self._decimal_var.get())
        except Exception:
            decimal_places = 1
        try:
            num_choices = int(self._num_choices_var.get())
        except Exception:
            num_choices = 3
        try:
            duration_minutes = int(self._duration_var.get())
        except Exception:
            duration_minutes = 0
        return ExamConfig(
            exam_name=self._name_var.get().strip() or "Esame",
            num_questions=nq,
            correct_answers=answers,
            course_name=course_name,
            scale=scale,
            passing_pct=passing_pct,
            decimal_places=decimal_places,
            num_choices=num_choices,
            duration_minutes=duration_minutes,
        )

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="Seleziona PDF marksheet",
            filetypes=[("PDF", "*.pdf"), ("Tutti", "*.*")],
        )
        if path:
            self._pdf_path_var.set(path)

    def _import_pdf_answers(self):
        path = self._pdf_path_var.get()
        if not path:
            path = filedialog.askopenfilename(
                title="Seleziona PDF marksheet",
                filetypes=[("PDF", "*.pdf"), ("Tutti", "*.*")],
            )
            if not path:
                return
            self._pdf_path_var.set(path)
        self._status_var.set("Estrazione risposte in corso…")
        self.update_idletasks()
        result = parse_qb_marksheet(path)
        if not result:
            messagebox.showerror("Errore", "Impossibile estrarre dati dal PDF.")
            self._status_var.set("Errore estrazione")
            return
        if result.get("exam_name"):
            self._name_var.set(result["exam_name"])
        if result.get("course_name"):
            self._course_var.set(result["course_name"])
        # QBGenerator PDFs always use A/B/C
        self._num_choices_var.set(3)
        if result.get("num_questions"):
            self._nq_var.set(result["num_questions"])
            self._refresh_grid()
        answers = result.get("answers", {})
        if answers:
            for q_num, letter in answers.items():
                idx = q_num - 1
                if 0 <= idx < len(self._answer_vars):
                    self._answer_vars[idx].set(letter.upper())
            n_total = result.get("num_questions") or self._nq_var.get()
            msg = f"\u2713 Estratte {len(answers)}/{n_total} risposte"
            self._tab1_status.set(msg)
            self._status_var.set(msg)
        else:
            n = result.get("num_questions")
            if n:
                self._tab1_status.set(
                    f"\u2713 Metadati estratti: {result['exam_name']} ({n}Q). Nessuna risposta trovata.")
            else:
                self._tab1_status.set("\u26a0 Nessuna risposta trovata nel PDF.")
            self._status_var.set("Estrazione parziale")

    def _on_pdf_drop(self, files):
        decoded = [f.decode("utf-8") if isinstance(f, bytes) else str(f) for f in files]
        pdfs = [p for p in decoded if p.lower().endswith(".pdf")]
        if pdfs:
            self._pdf_path_var.set(pdfs[0])
            self._import_pdf_answers()

    def _save_config(self):
        cfg = self._get_config()
        if not cfg:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".json", filetypes=[("Config JSON", "*.json")],
            initialfile=f"{cfg.exam_name}_config.json",
        )
        if path:
            data = {"exam_name": cfg.exam_name, "num_questions": cfg.num_questions,
                    "correct_answers": cfg.correct_answers,
                    "course_name": cfg.course_name,
                    "scale": cfg.scale, "passing_pct": cfg.passing_pct,
                    "decimal_places": cfg.decimal_places, "num_choices": cfg.num_choices,
                    "duration_minutes": cfg.duration_minutes}
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            self._tab1_status.set(f"\u2713 Config salvata: {Path(path).name}")

    def _load_config(self):
        path = filedialog.askopenfilename(
            title="Carica configurazione",
            filetypes=[("Config JSON", "*.json"), ("Tutti", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            self._name_var.set(data.get("exam_name", "Esame"))
            self._course_var.set(data.get("course_name", ""))
            nq = data.get("num_questions", 10)
            self._nq_var.set(nq)
            # Scale
            scale = data.get("scale", 30.0)
            if scale == 30.0:
                self._scale_mode_var.set("30simi")
            elif scale == 10.0:
                self._scale_mode_var.set("10simi")
            elif scale == 100.0:
                self._scale_mode_var.set("100simi")
            else:
                self._scale_mode_var.set("Personalizzato")
                self._scale_custom_var.set(str(scale))
            self._on_scale_mode_changed()
            self._passing_pct_var.set(data.get("passing_pct", 75.0))
            self._decimal_var.set(data.get("decimal_places", 1))
            # Set num_choices BEFORE _refresh_grid
            self._num_choices_var.set(data.get("num_choices", 3))
            self._duration_var.set(data.get("duration_minutes", 0))
            self._refresh_grid()
            for i, a in enumerate(data.get("correct_answers", [])):
                if i < len(self._answer_vars):
                    self._answer_vars[i].set(a)
            self._config = ExamConfig(
                exam_name=data["exam_name"],
                num_questions=nq,
                correct_answers=data.get("correct_answers", []),
                course_name=data.get("course_name", ""),
                scale=scale,
                passing_pct=data.get("passing_pct", 75.0),
                decimal_places=data.get("decimal_places", 1),
                num_choices=data.get("num_choices", 3),
                duration_minutes=data.get("duration_minutes", 0),
            )
            self._status_var.set(f"\u2713 Config: {self._config.exam_name} ({nq}Q)")
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile caricare config:\n{e}")

    def _generate_html(self):
        cfg = self._get_config()
        if not cfg:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".html", filetypes=[("HTML", "*.html")],
            initialfile=f"{cfg.exam_name}_risposte.html",
        )
        if path:
            generate_student_html(cfg.exam_name, cfg.num_questions, path,
                                  num_choices=cfg.num_choices,
                                  duration_minutes=cfg.duration_minutes)
            self._tab1_status.set(f"\u2713 Modulo HTML: {Path(path).name}")
            import webbrowser
            webbrowser.open(path)

    # ── TAB 2 ─────────────────────────────────────────────────────────────────

    def _build_tab2(self):
        tab = self._tab2
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)

        cfg_card = tk.LabelFrame(tab, text="  \u2699  Configurazione Esame  ",
                                  bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        cfg_card.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 4))
        cfg_card.columnconfigure(2, weight=1)
        ttk.Button(cfg_card, text="\U0001f4cb Usa config Tab 1",
                   command=self._use_tab1_config).grid(row=0, column=0, padx=8, pady=8)
        ttk.Button(cfg_card, text="\U0001f4c2 Carica config da file",
                   command=self._load_config_tab2).grid(row=0, column=1, padx=4)
        self._cfg_info = tk.StringVar(value="Nessuna configurazione caricata")
        tk.Label(cfg_card, textvariable=self._cfg_info,
                 bg=C_BG, fg=C_NAVY, font=FONT_BOLD).grid(row=0, column=2, padx=12)

        folder_card = tk.LabelFrame(tab, text="  \U0001f4c1  Risposte Studenti (JSON)  ",
                                     bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        folder_card.grid(row=1, column=0, sticky="ew", padx=12, pady=4)
        folder_card.columnconfigure(1, weight=1)
        self._drop_label = tk.Label(
            folder_card,
            text="\u2b07  Trascina qui i file JSON oppure scegli la cartella",
            bg="#e8f0fe", fg=C_NAVY, font=FONT_BOLD,
            relief="groove", bd=2, cursor="hand2", pady=12,
        )
        self._drop_label.grid(row=0, column=0, columnspan=4, sticky="ew", padx=8, pady=6)
        self._drop_label.bind("<Button-1>", lambda e: self._browse_folder())
        tk.Label(folder_card, text="Cartella:", bg=C_BG, font=FONT_UI).grid(
            row=1, column=0, padx=8, pady=6, sticky="w")
        self._folder_var = tk.StringVar()
        ttk.Entry(folder_card, textvariable=self._folder_var).grid(
            row=1, column=1, sticky="ew", padx=4)
        ttk.Button(folder_card, text="\U0001f4c2 Sfoglia",
                   command=self._browse_folder).grid(row=1, column=2, padx=4)
        ttk.Button(folder_card, text="\u25b6 Correggi", style="Accent.TButton",
                   command=self._run_correction).grid(row=1, column=3, padx=(4, 8))

        res_card = tk.LabelFrame(tab, text="  \U0001f4ca  Risultati  ",
                                  bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        res_card.grid(row=2, column=0, sticky="nsew", padx=12, pady=4)
        res_card.columnconfigure(0, weight=1)
        res_card.rowconfigure(0, weight=1)
        cols = ("name", "file", "correct", "score", "result")
        self._tree = ttk.Treeview(res_card, columns=cols, show="headings", selectmode="browse")
        for col, hd, w, anc in [
            ("name",    "Studente",  230, "w"),
            ("file",    "File",      200, "w"),
            ("correct", "Corrette",   90, "center"),
            ("score",   "Voto /30",   90, "center"),
            ("result",  "Esito",     120, "center"),
        ]:
            self._tree.heading(col, text=hd)
            self._tree.column(col, width=w, anchor=anc)
        self._tree.tag_configure("passed", background=C_GREEN)
        self._tree.tag_configure("failed", background=C_RED)
        vsb = ttk.Scrollbar(res_card, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._tree.pack(fill="both", expand=True, padx=(4, 0), pady=4)

        exp_frm = tk.Frame(tab, bg=C_BG)
        exp_frm.grid(row=3, column=0, sticky="ew", padx=12, pady=(4, 8))
        ttk.Button(exp_frm, text="\U0001f4ca CSV",
                   command=self._export_csv).pack(side="left", padx=4)
        ttk.Button(exp_frm, text="\U0001f4d7 Excel voti",
                   command=self._export_excel).pack(side="left", padx=4)
        ttk.Button(exp_frm, text="\U0001f4c4 Report studenti (Excel)",
                   command=self._export_student_reports).pack(side="left", padx=4)
        ttk.Button(exp_frm, text="\U0001f4c8 Vai a Statistiche",
                   command=self._go_to_stats).pack(side="left", padx=4)
        self._tab2_status = tk.StringVar()
        tk.Label(exp_frm, textvariable=self._tab2_status,
                 fg="#1e5928", bg=C_BG, font=FONT_UI).pack(side="left", padx=12)

    def _on_json_drop(self, files):
        decoded = [f.decode("utf-8") if isinstance(f, bytes) else str(f) for f in files]
        dirs = [p for p in decoded if Path(p).is_dir()]
        if dirs:
            self._folder_var.set(dirs[0])
            self._run_correction()
            return
        jsons = [p for p in decoded if p.lower().endswith(".json")]
        if jsons:
            self._folder_var.set(str(Path(jsons[0]).parent))
            self._run_correction()

    def _use_tab1_config(self):
        cfg = self._get_config()
        if cfg:
            self._config = cfg
            info = f"\u2713  {cfg.exam_name}  \u2014  {cfg.num_questions} domande  |  {cfg.score_label()}  |  soglia {cfg.passing_pct:.0f}%"
            if cfg.course_name:
                info += f"  |  {cfg.course_name}"
            self._cfg_info.set(info)
            self._header_sub.set(f"{cfg.exam_name}  \u2014  {cfg.num_questions} domande")

    def _load_config_tab2(self):
        self._load_config()
        if self._config:
            c = self._config
            info = f"\u2713  {c.exam_name}  \u2014  {c.num_questions} domande  |  {c.score_label()}  |  soglia {c.passing_pct:.0f}%"
            if c.course_name:
                info += f"  |  {c.course_name}"
            self._cfg_info.set(info)
            self._header_sub.set(f"{c.exam_name}  \u2014  {c.num_questions} domande")

    def _browse_folder(self):
        f = filedialog.askdirectory(title="Cartella con le risposte degli studenti")
        if f:
            self._folder_var.set(f)

    def _run_correction(self):
        if not self._config:
            messagebox.showinfo("Config mancante",
                                "Prima carica o configura l\'esame nella scheda 1.")
            return
        folder = self._folder_var.get()
        if not folder or not Path(folder).is_dir():
            messagebox.showerror("Errore", "Seleziona una cartella valida.")
            return
        json_files = list(Path(folder).glob("*.json"))
        student_files = []
        for jf in json_files:
            try:
                with open(jf, encoding="utf-8") as f:
                    d = json.load(f)
                if "answers" in d and ("name" in d or "student_name" in d):
                    student_files.append(jf)
            except Exception:
                pass
        if not student_files:
            messagebox.showinfo("Nessun file", "Nessun file JSON studente trovato.")
            return
        self._results.clear()
        for jf in student_files:
            try:
                with open(jf, encoding="utf-8") as f:
                    data = json.load(f)
                data["file"] = str(jf)
                self._results.append(correct_student(data, self._config))
            except Exception as e:
                messagebox.showwarning("Avviso", f"Errore in {jf.name}:\n{e}")
        self._results.sort(key=lambda r: r.score, reverse=True)
        self._stats = compute_stats(self._results, self._config)
        self._tree.heading("score", text=f"Voto{self._config.score_label()}")
        for item in self._tree.get_children():
            self._tree.delete(item)
        for r in self._results:
            tag = "passed" if r.passed else "failed"
            self._tree.insert("", "end", values=(
                r.student_name, Path(r.student_file).name,
                f"{r.num_correct}/{self._config.num_questions}",
                self._config.fmt_score(r.score),
                "\u2713 Idoneo" if r.passed else "\u2717 Non idoneo",
            ), tags=(tag,))
        n  = len(self._results)
        np_ = self._stats.get("n_passed", 0)
        msg = f"\u2713 {n} studenti — {np_} idonei ({self._stats.get('pct_passed', 0):.1f}%)"
        self._tab2_status.set(msg)
        self._status_var.set(f"Correzione completata: {n} studenti")

    def _export_csv(self):
        if not self._results:
            messagebox.showinfo("Info", "Esegui prima la correzione.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialfile=f"{self._config.exam_name}_voti.csv",
        )
        if path:
            try:
                export_csv(self._results, self._config, path)
                self._tab2_status.set(f"\u2713 CSV: {Path(path).name}")
            except Exception as e:
                messagebox.showerror("Errore", str(e))

    def _export_excel(self):
        if not self._results:
            messagebox.showinfo("Info", "Esegui prima la correzione.")
            return
        if not HAS_EXCEL:
            messagebox.showinfo("Info", "Installa openpyxl:\n  pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"{self._config.exam_name}_voti.xlsx",
        )
        if path:
            try:
                export_excel(self._results, self._config, path)
                self._tab2_status.set(f"\u2713 Excel: {Path(path).name}")
            except Exception as e:
                messagebox.showerror("Errore", str(e))

    def _export_student_reports(self):
        if not self._results:
            messagebox.showinfo("Info", "Esegui prima la correzione.")
            return
        if not HAS_EXCEL:
            messagebox.showinfo("Info", "Installa openpyxl:\n  pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"{self._config.exam_name}_report_studenti.xlsx",
        )
        if path:
            try:
                export_student_reports(self._results, self._config, path)
                self._tab2_status.set(f"\u2713 Report: {Path(path).name}")
            except Exception as e:
                messagebox.showerror("Errore", str(e))

    def _go_to_stats(self):
        self._update_stats()
        self._nb.select(2)

    # ── TAB 3 ─────────────────────────────────────────────────────────────────

    def _build_tab3(self):
        tab = self._tab3
        tab.columnconfigure(0, weight=1)
        tab.columnconfigure(1, weight=2)
        tab.rowconfigure(2, weight=1)

        top_bar = tk.Frame(tab, bg=C_BG)
        top_bar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=12, pady=(10, 4))
        ttk.Button(top_bar, text="\U0001f504 Aggiorna",
                   command=self._update_stats).pack(side="left", padx=4)
        tk.Frame(top_bar, bg=C_BORDER, width=2, height=28).pack(side="left", fill="y", padx=6)
        ttk.Button(top_bar, text="\U0001f4c4 Report studenti (Excel)",
                   command=self._export_student_reports).pack(side="left", padx=4)
        ttk.Button(top_bar, text="\U0001f4ca Item Analysis (Excel)",
                   command=self._export_item_analysis).pack(side="left", padx=4)
        ttk.Button(top_bar, text="\U0001f4c8 KPI (Excel)",
                   command=self._export_kpi).pack(side="left", padx=4)
        tk.Frame(top_bar, bg=C_BORDER, width=2, height=28).pack(side="left", fill="y", padx=6)
        ttk.Button(
            top_bar, text="\U0001f5a8 Stampa PDF studenti",
            command=self._print_student_reports_pdf,
            state="normal" if HAS_REPORTLAB else "disabled",
        ).pack(side="left", padx=4)
        ttk.Button(
            top_bar, text="\U0001f5a8 Stampa PDF statistiche",
            command=self._print_statistics_pdf,
            state="normal" if HAS_REPORTLAB else "disabled",
        ).pack(side="left", padx=4)
        if not HAS_REPORTLAB:
            tk.Label(top_bar, text="(installa reportlab per PDF)",
                     fg="gray", bg=C_BG, font=("Segoe UI", 8)).pack(side="left", padx=4)

        # KPI summary cards
        self._kpi_row = tk.Frame(tab, bg=C_BG)
        self._kpi_row.grid(row=1, column=0, columnspan=2, sticky="ew", padx=12, pady=(0, 4))

        txt_card = tk.LabelFrame(tab, text="  \U0001f4ca  Riepilogo Statistico  ",
                                  bg=C_BG, fg=C_NAVY, font=FONT_BOLD, relief="groove", bd=2)
        txt_card.grid(row=2, column=0, sticky="nsew", padx=(12, 5), pady=4)
        txt_card.rowconfigure(0, weight=1)
        txt_card.columnconfigure(0, weight=1)
        self._stats_text = scrolledtext.ScrolledText(
            txt_card, width=44, state="disabled",
            font=FONT_MONO, wrap="word", bg=C_WHITE, fg=C_TEXT,
            relief="flat", bd=0,
        )
        self._stats_text.grid(row=0, column=0, sticky="nsew", padx=4, pady=4)

        self._chart_frm = tk.LabelFrame(tab, text="  \U0001f4c8  Grafici  ",
                                         bg=C_BG, fg=C_NAVY, font=FONT_BOLD,
                                         relief="groove", bd=2)
        self._chart_frm.grid(row=2, column=1, sticky="nsew", padx=(5, 12), pady=4)
        self._chart_frm.rowconfigure(0, weight=1)
        self._chart_frm.columnconfigure(0, weight=1)

    def _draw_kpi_cards(self, stats: dict, config: "ExamConfig"):
        for w in self._kpi_row.winfo_children():
            w.destroy()
        pct = stats.get("pct_passed", 0)
        pass_fg  = "#1e5928" if pct >= 70 else ("#7d4e00" if pct >= 50 else "#9c1c1c")
        pass_bg  = "#C6EFCE" if pct >= 70 else ("#FFEB9C" if pct >= 50 else "#FFC7CE")
        cards = [
            (str(stats.get("n", 0)), "Studenti",         "#444",    "#f0f4f8"),
            (f"{config.fmt_score(stats.get('mean', 0))}{config.score_label()}", "Media",
             C_NAVY, "#dce6f0"),
            (f"{config.fmt_score(stats.get('median', 0))}{config.score_label()}", "Mediana",
             C_NAVY, "#dce6f0"),
            (f"{config.fmt_score(stats.get('std', 0))}", "Dev. std.", "#7d4e00", "#FFEB9C"),
            (f"{stats.get('n_passed', 0)}  ({pct:.1f}%)", "Idonei", pass_fg, pass_bg),
        ]
        for val, label, fg, bg in cards:
            card = tk.Frame(self._kpi_row, bg=bg, padx=14, pady=6,
                            highlightbackground=C_BORDER, highlightthickness=1)
            card.pack(side="left", padx=4, pady=2)
            tk.Label(card, text=val, bg=bg, fg=fg,
                     font=("Segoe UI", 14, "bold")).pack()
            tk.Label(card, text=label, bg=bg, fg="#666",
                     font=("Segoe UI", 8)).pack()

    def _update_stats(self):
        if not self._results or not self._stats:
            self._set_stats_text(
                "Nessun risultato disponibile.\nEsegui prima la correzione.")
            for w in self._kpi_row.winfo_children():
                w.destroy()
            return
        s = self._stats
        cfg = self._config
        lbl = cfg.score_label()
        self._draw_kpi_cards(s, cfg)
        lines = [
            "=" * 38,
            f"  ESAME: {cfg.exam_name}",
        ]
        if cfg.course_name:
            lines.append(f"  Corso: {cfg.course_name}")
        lines += [
            "=" * 38,
            f"  Studenti         : {s['n']}",
            f"  Media voto       : {cfg.fmt_score(s['mean'])}{lbl}",
            f"  Mediana          : {cfg.fmt_score(s['median'])}{lbl}",
            f"  Dev. std.        : {cfg.fmt_score(s['std'])}",
            f"  Minimo           : {cfg.fmt_score(s['min'])}{lbl}",
            f"  Massimo          : {cfg.fmt_score(s['max'])}{lbl}",
            "-" * 38,
            f"  Idonei           : {s['n_passed']} ({s['pct_passed']:.1f}%)",
            f"  Non idonei       : {s['n'] - s['n_passed']} ({100 - s['pct_passed']:.1f}%)",
            f"  Soglia suff.     : {cfg.fmt_score(cfg.passing_score)}{lbl} ({cfg.passing_pct:.0f}%)",
            "-" * 38,
            "  TOP 5 errori per domanda:",
        ]
        q_err = sorted(enumerate(s.get("q_errors", []), 1),
                       key=lambda x: x[1], reverse=True)
        for qi, err in q_err[:5]:
            lines.append(f"    Q{qi:>3}: {err:.1f}% errori")
        self._set_stats_text("\n".join(lines))
        self._draw_charts()

    def _set_stats_text(self, text: str):
        self._stats_text.config(state="normal")
        self._stats_text.delete("1.0", "end")
        self._stats_text.insert("1.0", text)
        self._stats_text.config(state="disabled")

    def _draw_charts(self):
        for w in self._chart_frm.winfo_children():
            w.destroy()
        if not HAS_PLOT or not self._stats:
            tk.Label(self._chart_frm,
                     text="Installa matplotlib:\n  pip install matplotlib",
                     bg=C_BG, fg="gray", font=FONT_UI).pack(pady=30)
            return
        scores = self._stats.get("scores", [])
        q_err  = self._stats.get("q_errors", [])
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(8, 3.8), facecolor=C_BG)
        fig.subplots_adjust(wspace=0.35)
        scale = self._config.scale
        step = max(scale / 15, 1)
        bins = [i * step for i in range(int(scale / step) + 2)]
        ax1.hist(scores, bins=bins, color=C_BLUE, edgecolor="white", alpha=0.85)
        ax1.axvline(x=self._config.passing_score, color="#C00000",
                    linestyle="--", lw=1.5, label="Sufficienza")
        ax1.set_title("Distribuzione voti", fontsize=10, fontweight="bold")
        ax1.set_xlabel(f"Voto{self._config.score_label()}")
        ax1.set_ylabel("N studenti")
        ax1.legend(fontsize=8)
        ax1.set_facecolor("#F8FBFF")
        if q_err:
            xs = list(range(1, len(q_err) + 1))
            ax2.bar(xs, q_err,
                    color=["#C00000" if e >= 50 else C_BLUE for e in q_err],
                    alpha=0.85, width=0.7)
            ax2.axhline(y=50, color="#C00000", linestyle="--", lw=1, label="50%")
            ax2.set_title("% errori per domanda", fontsize=10, fontweight="bold")
            ax2.set_xlabel("Domanda")
            ax2.set_ylabel("% errati")
            ax2.set_ylim(0, 105)
            ax2.legend(fontsize=8)
            ax2.tick_params(axis="x", labelsize=7, rotation=45)
            ax2.set_facecolor("#F8FBFF")
        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self._chart_frm)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    def _export_item_analysis(self):
        if not self._results:
            messagebox.showinfo("Info", "Esegui prima la correzione.")
            return
        if not HAS_EXCEL:
            messagebox.showinfo("Info", "Installa openpyxl:\n  pip install openpyxl")
            return
        session_name = simpledialog.askstring(
            "Nome sessione", "Inserisci il nome della sessione (es. Giugno 2025):",
            parent=self,
        )
        if session_name is None:
            return
        append_to = None
        if messagebox.askyesno("Aggiungi a file esistente?",
                               "Aggiungere a un file Item Analysis esistente?\n"
                               "SI -> seleziona file   NO -> crea nuovo"):
            append_to = filedialog.askopenfilename(
                title="Seleziona file Item Analysis esistente",
                filetypes=[("Excel", "*.xlsx")],
            )
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=(Path(append_to).name if append_to
                         else f"{self._config.exam_name}_item_analysis.xlsx"),
            initialdir=str(Path(append_to).parent) if append_to else None,
        )
        if path:
            try:
                export_item_analysis(self._results, self._config, self._stats,
                                     path, session_name, append_to=append_to)
                messagebox.showinfo("Esportato", f"Item Analysis salvato:\n{path}")
            except Exception as e:
                messagebox.showerror("Errore", str(e))

    def _export_kpi(self):
        if not self._results:
            messagebox.showinfo("Info", "Esegui prima la correzione.")
            return
        if not HAS_EXCEL:
            messagebox.showinfo("Info", "Installa openpyxl:\n  pip install openpyxl")
            return
        session_name = simpledialog.askstring(
            "Nome sessione", "Inserisci il nome della sessione (es. Giugno 2025):",
            parent=self,
        )
        if session_name is None:
            return
        append_to = None
        if messagebox.askyesno("Aggiungi a file KPI esistente?",
                               "Aggiungere a un file KPI esistente?\n"
                               "SI -> seleziona file   NO -> crea nuovo"):
            append_to = filedialog.askopenfilename(
                title="Seleziona file KPI esistente",
                filetypes=[("Excel", "*.xlsx")],
            )
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=(Path(append_to).name if append_to
                         else f"{self._config.exam_name}_KPI.xlsx"),
            initialdir=str(Path(append_to).parent) if append_to else None,
        )
        if path:
            try:
                export_kpi(self._results, self._config, self._stats,
                           path, session_name, append_to=append_to)
                messagebox.showinfo("Esportato", f"KPI salvato:\n{path}")
            except Exception as e:
                messagebox.showerror("Errore", str(e))

    def _print_student_reports_pdf(self):
        if not self._results:
            messagebox.showinfo("Info", "Esegui prima la correzione.")
            return
        if not HAS_REPORTLAB:
            messagebox.showinfo("Info", "Installa reportlab:\n  pip install reportlab")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF", "*.pdf")],
            initialfile=f"{self._config.exam_name}_report_studenti.pdf",
        )
        if not path:
            return
        try:
            print_student_reports_pdf(self._results, self._config, path)
            messagebox.showinfo("PDF generato", f"Report studenti:\n{path}")
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Errore PDF", str(e))

    def _print_statistics_pdf(self):
        if not self._results:
            messagebox.showinfo("Info", "Esegui prima la correzione.")
            return
        if not HAS_REPORTLAB:
            messagebox.showinfo("Info", "Installa reportlab:\n  pip install reportlab")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF", "*.pdf")],
            initialfile=f"{self._config.exam_name}_statistiche.pdf",
        )
        if not path:
            return
        try:
            if not self._stats:
                self._stats = compute_stats(self._results, self._config)
            print_statistics_pdf(self._results, self._config, self._stats, path)
            messagebox.showinfo("PDF generato", f"Statistiche:\n{path}")
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Errore PDF", str(e))

    # ── Drag & drop ───────────────────────────────────────────────────────────

    def _setup_dnd(self):
        if not HAS_DND:
            return
        try:
            windnd.hook_dropfiles(self._pdf_drop_frame, func=self._on_pdf_drop)
            windnd.hook_dropfiles(self._pdf_drop_label, func=self._on_pdf_drop)
            windnd.hook_dropfiles(self._drop_label,     func=self._on_json_drop)
            windnd.hook_dropfiles(self._tree,           func=self._on_json_drop)
        except Exception:
            pass


if __name__ == "__main__":
    TeacherApp().mainloop()
